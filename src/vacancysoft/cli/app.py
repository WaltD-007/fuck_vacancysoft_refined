from __future__ import annotations

import asyncio
import os
from pathlib import Path

import typer
from dotenv import load_dotenv
from sqlalchemy import select

load_dotenv()


def _load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            os.environ.setdefault(key, value)

from vacancysoft.adapters import ADAPTER_REGISTRY, derive_workday_candidate_endpoints
from vacancysoft.adapters.base import DiscoveredJobRecord, SourceAdapter
from vacancysoft.db.base import Base
from vacancysoft.db.engine import build_engine
from vacancysoft.db.models import ClassificationResult, EnrichedJob, RawJob, ReviewQueueItem, ScoreResult, Source, SourceRun
from vacancysoft.db.session import SessionLocal
from vacancysoft.exporters.excel_exporter import export_profile_to_excel, export_segment_to_excel
from vacancysoft.exporters.json_exporter import export_profile_to_json, export_segment_to_json
from vacancysoft.exporters.views import (
    accepted_only_query,
    accepted_plus_review_query,
    client_segment_query,
    fetch_rows,
    grouped_by_taxonomy_query,
    load_exporter_config,
)
from vacancysoft.exporters.webhook_sender import send_new_leads_to_webhook, send_profile_to_webhook, send_segment_to_webhook
from vacancysoft.pipelines.classification_persistence import classify_enriched_jobs
from vacancysoft.pipelines.enrichment_persistence import enrich_raw_jobs
from vacancysoft.pipelines.maintenance import cleanup_orphaned_classification_results
from vacancysoft.pipelines.persistence import persist_discovery_batch
from vacancysoft.pipelines.detail_backfill import backfill_detail_for_enriched_jobs
from vacancysoft.pipelines.scoring_persistence import score_enriched_jobs
from vacancysoft.source_registry.seed_loader import seed_sources_from_yaml

app = typer.Typer(help="Prospero — Recruitment Intelligence Platform")
pipeline_app = typer.Typer(help="Pipeline commands")
export_app = typer.Typer(help="Export helpers")
db_app = typer.Typer(help="Database helpers")
app.add_typer(pipeline_app, name="pipeline")
app.add_typer(export_app, name="export")
app.add_typer(db_app, name="db")


@app.callback()
def _load_env(ctx: typer.Context) -> None:
    root = Path.cwd()
    _load_env_file(root / ".env")
    _load_env_file(root / "alembic" / "env")


@db_app.command("init")
def init_db() -> None:
    engine = build_engine()
    Base.metadata.create_all(bind=engine)
    typer.echo("Database schema initialised")


@db_app.command("seed-sources")
def seed_sources(config_path: str = typer.Option("configs/seeds/employers.yaml", "--config-path")) -> None:
    with SessionLocal() as session:
        created, updated = seed_sources_from_yaml(session, config_path)
    typer.echo(f"Seeded sources. created={created} updated={updated}")


@db_app.command("seed-config-boards")
def seed_config_boards() -> None:
    from vacancysoft.source_registry.config_seed_loader import seed_sources_from_config
    with SessionLocal() as session:
        created, updated = seed_sources_from_config(session)
    typer.echo(f"Seeded config boards. created={created} updated={updated}")


@db_app.command("stats")
def db_stats() -> None:
    with SessionLocal() as session:
        source_count = len(list(session.execute(select(Source)).scalars()))
        run_count = len(list(session.execute(select(SourceRun)).scalars()))
        raw_job_count = len(list(session.execute(select(RawJob)).scalars()))
        enriched_job_count = len(list(session.execute(select(EnrichedJob)).scalars()))
        classification_count = len(list(session.execute(select(ClassificationResult)).scalars()))
        score_count = len(list(session.execute(select(ScoreResult)).scalars()))
    typer.echo(
        f"sources={source_count} source_runs={run_count} raw_jobs={raw_job_count} enriched_jobs={enriched_job_count} classification_results={classification_count} score_results={score_count}"
    )


@db_app.command("cleanup-classifications")
def cleanup_classifications() -> None:
    with SessionLocal() as session:
        removed = cleanup_orphaned_classification_results(session)
    typer.echo(f"Removed orphaned classification results: {removed}")


@db_app.command("reset-pipeline")
def reset_pipeline() -> None:
    """Delete all enriched jobs, classifications, and scores so the pipeline can re-run from raw jobs."""
    from sqlalchemy import delete, text
    # Delete dossiers first (FK to enriched_jobs) — use separate session so FK errors don't poison the main txn
    d = 0
    try:
        with SessionLocal() as session:
            d = session.execute(text("DELETE FROM intelligence_dossiers")).rowcount
            session.commit()
    except Exception:
        pass
    # Also clear queue items that reference enriched jobs
    try:
        with SessionLocal() as session:
            session.execute(text("DELETE FROM review_queue_items"))
            session.commit()
    except Exception:
        pass
    with SessionLocal() as session:
        s = session.execute(delete(ScoreResult)).rowcount
        session.commit()
    with SessionLocal() as session:
        c = session.execute(delete(ClassificationResult)).rowcount
        session.commit()
    with SessionLocal() as session:
        e = session.execute(delete(EnrichedJob)).rowcount
        session.commit()
    typer.echo(f"Cleared pipeline data. dossiers={d} enriched={e} classifications={c} scores={s}")


@db_app.command("fix-adapters")
def fix_adapters() -> None:
    """Scan all generic_site sources and reassign to correct adapter based on URL patterns."""
    from vacancysoft.source_registry.config_seed_loader import detect_adapter_from_url, PLATFORM_REGISTRY
    fixed = 0
    with SessionLocal() as session:
        generics = list(session.execute(
            select(Source).where(Source.adapter_name == "generic_site", Source.active.is_(True))
        ).scalars())
        for src in generics:
            override = detect_adapter_from_url(src.base_url)
            if override:
                adapter, platform_key = override
                meta = PLATFORM_REGISTRY.get(platform_key)
                if meta:
                    src.adapter_name = meta["adapter"]
                    src.source_type = meta["source_type"]
                    src.ats_family = meta["ats_family"]
                    src.board_name = meta["board_name"]
                    typer.echo(f"  Fixed: {src.employer_name} → {meta['adapter']}")
                    fixed += 1
        session.commit()
    typer.echo(f"Fixed {fixed} sources")


@db_app.command("add-source")
def add_source(
    url: str = typer.Argument(..., help="Careers page URL to add"),
    company: str | None = typer.Option(None, "--company", help="Company name (auto-detected if not provided)"),
    scrape: bool = typer.Option(True, "--scrape/--no-scrape", help="Immediately scrape after adding (default: yes)"),
) -> None:
    """Auto-detect platform from a URL and add it as a source."""
    import hashlib
    from urllib.parse import urlparse as _urlparse
    from vacancysoft.api.source_detector import detect_and_validate
    from vacancysoft.source_registry.config_seed_loader import PLATFORM_REGISTRY

    result = asyncio.run(detect_and_validate(url))

    if result["error"] and not result["reachable"]:
        typer.echo(f"Warning: {result['error']} — adding as generic browser source anyway")
        result["adapter"] = "generic_site"
        result["slug"] = None

    adapter = result["adapter"]
    slug = result["slug"]
    company_name = company or result["company_guess"] or "Unknown"
    job_count = result["job_count"]

    typer.echo(f"Detected: {adapter} | {company_name} | {f'{job_count} jobs' if job_count is not None else 'reachable'}")

    # Map detector adapter names to PLATFORM_REGISTRY keys
    adapter_map = {
        "greenhouse": "greenhouse", "workday": "workday", "lever": "lever",
        "icims": "icims", "ashby": "ashby", "smartrecruiters": "smartrecruiters",
        "workable": "workable", "oracle_cloud": "oracle", "successfactors": "successfactors",
        "eightfold": "eightfold", "pinpoint": "pinpoint", "hibob": "hibob",
        "taleo": "taleo", "teamtailor": "teamtailor", "generic_site": "generic_browser",
    }
    platform_key = adapter_map.get(adapter, "generic_browser")
    meta = PLATFORM_REGISTRY.get(platform_key, PLATFORM_REGISTRY["generic_browser"])

    # Build config blob with adapter-specific fields
    config_blob = {"job_board_url": url}
    if slug:
        config_blob["slug"] = slug

    from urllib.parse import urlparse as _wp
    _p = _wp(url)
    _host_parts = _p.netloc.lower().split(".")
    _path_parts = [p for p in _p.path.split("/") if p and p.lower() not in ("en-us", "en-gb", "en", "en_us", "en_gb", "jobs", "job")]

    if adapter == "workday":
        _tenant = _host_parts[0]
        _shard = _host_parts[1] if len(_host_parts) > 2 else _host_parts[0]
        _site_path = _path_parts[-1] if _path_parts else _tenant
        config_blob["endpoint_url"] = f"https://{_tenant}.{_shard}.myworkdayjobs.com/wday/cxs/{_tenant}/{_site_path}/jobs"
        config_blob["tenant"] = _tenant
        config_blob["shard"] = _shard
        config_blob["site_path"] = _site_path

    elif adapter == "greenhouse":
        # API just needs the slug
        config_blob["slug"] = slug

    elif adapter == "lever":
        config_blob["slug"] = slug

    elif adapter == "icims":
        # iCIMS uses the subdomain as slug, needs in_iframe=1
        config_blob["slug"] = slug or _host_parts[0]

    elif adapter == "ashby":
        config_blob["slug"] = slug

    elif adapter == "smartrecruiters":
        config_blob["slug"] = slug

    elif adapter == "workable":
        config_blob["slug"] = slug

    elif adapter == "oracle_cloud":
        # Oracle Cloud just needs the full URL
        config_blob["job_board_url"] = url

    elif adapter == "successfactors":
        config_blob["job_board_url"] = url

    elif adapter == "eightfold":
        config_blob["slug"] = slug or _host_parts[0]

    elif adapter == "pinpoint":
        config_blob["slug"] = slug

    elif adapter == "taleo":
        config_blob["job_board_url"] = url

    elif adapter == "hibob":
        config_blob["job_board_url"] = url

    # Build source key
    def _slugify(v):
        return "_".join("".join(ch.lower() if ch.isalnum() else " " for ch in v).split())

    url_hash = hashlib.md5(url.encode()).hexdigest()[:8]
    source_key = f"{meta['adapter']}_{_slugify(company_name)}_{url_hash}"
    parsed = _urlparse(url)
    hostname = parsed.hostname or "unknown"

    with SessionLocal() as session:
        existing = session.execute(
            select(Source).where((Source.source_key == source_key) | (Source.base_url == url))
        ).scalars().first()

        if existing:
            typer.echo(f"Source already exists: {existing.employer_name} ({existing.adapter_name})")
            raise typer.Exit(0)

        session.add(Source(
            source_key=source_key,
            employer_name=company_name,
            board_name=meta["board_name"],
            base_url=url,
            hostname=hostname,
            source_type=meta["source_type"],
            ats_family=meta["ats_family"],
            adapter_name=meta["adapter"],
            active=True,
            seed_type="manual_add",
            discovery_method="url_auto_detect",
            fingerprint=f"{hostname}|{meta['ats_family'] or meta['adapter']}",
            canonical_company_key=_slugify(company_name),
            config_blob=config_blob,
            capability_blob={},
        ))
        session.commit()

    typer.echo(f"Added: {company_name} ({meta['adapter']})")

    if scrape:
        typer.echo(f"Scraping {company_name}...")
        from vacancysoft.cli.app import run_pipeline
        run_pipeline(
            adapter=None,
            source_key=source_key,
            discover_limit=0,
            per_adapter=0,
            export_profile="accepted_plus_review",
            output="leads_output.xlsx",
            dry_run=True,
            snapshot_every=1,
        )
    else:
        typer.echo("Will be included in next pipeline run")


@pipeline_app.command("discover")
def discover(
    adapter: str | None = typer.Option(None, "--adapter", help="Only run sources for this adapter"),
    source_key: str | None = typer.Option(None, "--source-key", help="Run a single source by key"),
    limit: int = typer.Option(0, "--limit", help="Max sources to process (0 = all)"),
) -> None:
    typer.echo("Discovering...")
    ok, failed, total_jobs = _run_discovery(adapter, source_key, limit)
    typer.echo(f"Discovery complete. ok={ok} failed={failed} total_jobs={total_jobs}")


@pipeline_app.command("discover-demo")
def discover_demo(source_key: str | None = typer.Option(None, "--source-key")) -> None:
    with SessionLocal() as session:
        if source_key:
            source = session.execute(select(Source).where(Source.source_key == source_key)).scalar_one_or_none()
        else:
            source = session.execute(select(Source).where(Source.active.is_(True)).limit(1)).scalar_one_or_none()

        if source is None:
            raise typer.BadParameter("No source found. Run 'prospero db seed-sources' first.")

        sample_jobs = [
            DiscoveredJobRecord(
                external_job_id=f"demo-{source.source_key}-001",
                title_raw=f"Senior Risk Manager at {source.employer_name}",
                location_raw="London, UK",
                posted_at_raw="2026-04-05",
                summary_raw="Demo discovery record persisted through the new pipeline.",
                discovered_url=f"{source.base_url.rstrip('/')}/jobs/demo-001",
                apply_url=f"{source.base_url.rstrip('/')}/jobs/demo-001/apply",
                listing_payload={"demo": True, "source_key": source.source_key},
                completeness_score=0.82,
                extraction_confidence=0.88,
                provenance={"mode": "demo", "source_key": source.source_key},
            )
        ]
        source_run, count = persist_discovery_batch(session=session, source=source, records=sample_jobs, trigger="manual")
        source_run_id = source_run.id
    typer.echo(f"Demo discovery persisted. source_run_id={source_run_id} raw_jobs={count}")


@pipeline_app.command("discover-workday")
def discover_workday(
    endpoint_url: str | None = typer.Option(None, "--endpoint-url"),
    job_board_url: str | None = typer.Option(None, "--job-board-url"),
    limit: int = typer.Option(20, "--limit"),
    persist: bool = typer.Option(False, "--persist"),
    source_key: str | None = typer.Option(None, "--source-key"),
) -> None:
    adapter = WorkdayAdapter()
    if not endpoint_url and not job_board_url:
        raise typer.BadParameter("Provide either --endpoint-url or --job-board-url")

    if job_board_url and not endpoint_url:
        candidates = derive_workday_candidate_endpoints(job_board_url)
        typer.echo(f"Derived Workday endpoint candidates: {candidates}")
        resolved_endpoint_url, page = asyncio.run(
            adapter.discover_from_board_url(job_board_url=job_board_url, limit=limit)
        )
    else:
        source_config = {
            "endpoint_url": endpoint_url,
            "job_board_url": job_board_url,
            "limit": limit,
        }
        page = asyncio.run(adapter.discover(source_config=source_config))
        resolved_endpoint_url = str(endpoint_url)

    typer.echo(
        f"Workday discovery returned jobs={len(page.jobs)} next_cursor={page.next_cursor} endpoint={resolved_endpoint_url}"
    )
    for record in page.jobs[:5]:
        typer.echo(f"- {record.title_raw} | {record.location_raw} | {record.discovered_url}")

    if not persist:
        return
    if not source_key:
        raise typer.BadParameter("--source-key is required when using --persist")

    with SessionLocal() as session:
        source = session.execute(select(Source).where(Source.source_key == source_key)).scalar_one_or_none()
        if source is None:
            raise typer.BadParameter(f"No source found for source_key={source_key}")
        source_run, count = persist_discovery_batch(session=session, source=source, records=page.jobs, trigger="manual")
        typer.echo(f"Persisted Workday discovery. source_run_id={source_run.id} raw_jobs={count}")


@pipeline_app.command("discover-adzuna")
def discover_adzuna(
    search_term: list[str] = typer.Option([], "--search-term"),
    country: list[str] = typer.Option([], "--country"),
    max_pages: int = typer.Option(5, "--max-pages"),
    results_per_page: int = typer.Option(50, "--results-per-page"),
) -> None:
    adapter = AdzunaAdapter()
    page = asyncio.run(
        adapter.discover(
            source_config={
                "search_terms": search_term,
                "countries": country,
                "max_pages": max_pages,
                "results_per_page": results_per_page,
            }
        )
    )
    typer.echo(f"Adzuna discovery returned jobs={len(page.jobs)}")
    for record in page.jobs[:10]:
        company = record.provenance.get("company", "")
        typer.echo(f"- {record.title_raw} | {company} | {record.location_raw} | {record.discovered_url}")


@pipeline_app.command("discover-workable")
def discover_workable(
    slug: str = typer.Option(..., "--slug"),
    company: str | None = typer.Option(None, "--company"),
    job_board_url: str | None = typer.Option(None, "--job-board-url"),
) -> None:
    adapter = WorkableAdapter()
    page = asyncio.run(
        adapter.discover(
            source_config={
                "slug": slug,
                "company": company,
                "job_board_url": job_board_url,
            }
        )
    )
    typer.echo(f"Workable discovery returned jobs={len(page.jobs)}")
    for record in page.jobs[:10]:
        company_name = record.provenance.get("company", "")
        typer.echo(f"- {record.title_raw} | {company_name} | {record.location_raw} | {record.discovered_url}")


@pipeline_app.command("discover-greenhouse")
def discover_greenhouse(
    slug: str = typer.Option(..., "--slug"),
    company: str | None = typer.Option(None, "--company"),
    job_board_url: str | None = typer.Option(None, "--job-board-url"),
) -> None:
    adapter = GreenhouseAdapter()
    page = asyncio.run(
        adapter.discover(
            source_config={
                "slug": slug,
                "company": company,
                "job_board_url": job_board_url,
            }
        )
    )
    typer.echo(f"Greenhouse discovery returned jobs={len(page.jobs)}")
    for record in page.jobs[:10]:
        company_name = record.provenance.get("company", "")
        typer.echo(f"- {record.title_raw} | {company_name} | {record.location_raw} | {record.discovered_url}")


@pipeline_app.command("discover-eightfold")
def discover_eightfold(
    job_board_url: str = typer.Option(..., "--job-board-url"),
    company: str | None = typer.Option(None, "--company"),
    search_term: list[str] = typer.Option([], "--search-term"),
) -> None:
    adapter = EightfoldAdapter()
    page = asyncio.run(
        adapter.discover(
            source_config={
                "job_board_url": job_board_url,
                "company": company,
                "search_terms": search_term or None,
            }
        )
    )
    typer.echo(f"Eightfold discovery returned jobs={len(page.jobs)}")
    for record in page.jobs[:10]:
        company_name = record.provenance.get("company", "")
        typer.echo(f"- {record.title_raw} | {company_name} | {record.location_raw} | {record.discovered_url}")


@pipeline_app.command("enrich")
def enrich(limit: int = typer.Option(0, "--limit", help="Max records to enrich (0 = all pending)")) -> None:
    with SessionLocal() as session:
        count = enrich_raw_jobs(session, limit=limit or None)
    typer.echo(f"Enrichment complete. enriched_jobs={count}")


@pipeline_app.command("enrich-demo")
def enrich_demo(limit: int = typer.Option(10, "--limit")) -> None:
    with SessionLocal() as session:
        count = enrich_raw_jobs(session, limit=limit)
    typer.echo(f"Demo enrichment persisted. enriched_jobs={count}")


@pipeline_app.command("backfill")
def backfill(
    limit: int = typer.Option(100, "--limit", help="Max records to backfill (0 = all pending)"),
    concurrency: int = typer.Option(5, "--concurrency"),
) -> None:
    """Fetch missing location/date from individual job pages for already-enriched jobs."""
    with SessionLocal() as session:
        count = backfill_detail_for_enriched_jobs(session, limit=limit or None, concurrency=concurrency)
    typer.echo(f"Backfill complete. updated={count}")


@pipeline_app.command("classify")
def classify(limit: int = typer.Option(0, "--limit", help="Max records to classify (0 = all pending)")) -> None:
    with SessionLocal() as session:
        count = classify_enriched_jobs(session, limit=limit or None)
    typer.echo(f"Classification complete. classification_results={count}")


@pipeline_app.command("classify-demo")
def classify_demo(limit: int = typer.Option(10, "--limit")) -> None:
    with SessionLocal() as session:
        count = classify_enriched_jobs(session, limit=limit)
    typer.echo(f"Demo classification persisted. classification_results={count}")


@pipeline_app.command("score")
def score(limit: int = typer.Option(0, "--limit", help="Max records to score (0 = all pending)")) -> None:
    with SessionLocal() as session:
        count = score_enriched_jobs(session, limit=limit or None)
    typer.echo(f"Scoring complete. score_results={count}")


@pipeline_app.command("score-demo")
def score_demo(limit: int = typer.Option(10, "--limit")) -> None:
    with SessionLocal() as session:
        count = score_enriched_jobs(session, limit=limit)
    typer.echo(f"Demo scoring persisted. score_results={count}")


def _run_discovery(
    adapter_filter: str | None,
    source_key_filter: str | None,
    discover_limit: int,
) -> tuple[int, int, int]:
    """Shared discovery logic used by both `discover` and `run` commands."""
    with SessionLocal() as session:
        query = select(Source).where(Source.active.is_(True))
        if source_key_filter:
            query = select(Source).where(Source.source_key == source_key_filter)
        elif adapter_filter:
            query = query.where(Source.adapter_name == adapter_filter)
        sources = list(session.execute(query).scalars())
        if discover_limit > 0:
            sources = sources[:discover_limit]

    ok_count = 0
    fail_count = 0
    total_jobs = 0

    for source in sources:
        adapter_cls = ADAPTER_REGISTRY.get(source.adapter_name)
        if adapter_cls is None:
            typer.echo(f"  SKIP {source.source_key}: unknown adapter '{source.adapter_name}'")
            fail_count += 1
            continue

        config = dict(source.config_blob or {})
        config.setdefault("company", source.employer_name)

        try:
            adapter_instance = adapter_cls()

            # Persist each page of results as it's scraped (for paginating adapters)
            _page_jobs_counted = 0

            def _page_cb(page_num: int, new_records: list, all_recs: list) -> None:
                nonlocal _page_jobs_counted
                _page_jobs_counted += len(new_records)
                with SessionLocal() as session:
                    source_obj = session.execute(
                        select(Source).where(Source.source_key == source.source_key)
                    ).scalar_one()
                    persist_discovery_batch(
                        session=session, source=source_obj, records=new_records, trigger="pipeline_discover",
                    )
                typer.echo(f"    page {page_num}: +{len(new_records)} jobs persisted")

            page = asyncio.run(adapter_instance.discover(source_config=config, on_page_scraped=_page_cb))

            job_count = len(page.jobs)
            total_jobs += job_count

            remaining = job_count - _page_jobs_counted
            if remaining > 0 and page.jobs:
                with SessionLocal() as session:
                    source_obj = session.execute(
                        select(Source).where(Source.source_key == source.source_key)
                    ).scalar_one()
                    _run, raw_count = persist_discovery_batch(
                        session=session, source=source_obj, records=page.jobs, trigger="pipeline_discover",
                    )
                typer.echo(f"  OK {source.source_key}: {job_count} jobs, {raw_count} persisted")
            elif _page_jobs_counted > 0:
                typer.echo(f"  OK {source.source_key}: {job_count} jobs (persisted page-by-page)")
            else:
                typer.echo(f"  OK {source.source_key}: 0 jobs")
            ok_count += 1
        except Exception as exc:
            typer.echo(f"  FAIL {source.source_key}: {type(exc).__name__}: {exc}")
            fail_count += 1

    return ok_count, fail_count, total_jobs


def _snapshot_excel(
    output_path: str,
    export_profile: str,
    limit: int = 50000,
    scrape_log: list[tuple] | None = None,
    source_results: list[dict] | None = None,
) -> None:
    """Write current scored leads to Excel (incremental checkpoint), with optional Log + Source Results sheets."""
    try:
        with SessionLocal() as session:
            export_profile_to_excel(session, profile_name=export_profile, output_path=output_path, limit=limit)

        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = load_workbook(output_path)
        hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hfill = PatternFill("solid", start_color="1F3864")
        rfont = Font(name="Arial", size=10)

        # --- Scrape Log sheet (page-level activity) ---
        if scrape_log:
            if "Log" in wb.sheetnames:
                del wb["Log"]
            ws = wb.create_sheet("Log")
            headers = ["Timestamp", "Company", "Page", "New Jobs", "Running Total", "Duration (s)"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = hfont
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 35
            ws.column_dimensions["C"].width = 8
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 14
            for row_idx, entry in enumerate(scrape_log, 2):
                ts, company, page_num, new_count, running_total, duration = entry
                for col, val in enumerate([ts, company, page_num, new_count, running_total, duration], 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.font = rfont
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:F{ws.max_row}"

        # --- Source Results sheet (rebuilt from full DB every time) ---
        from sqlalchemy import func as sa_func
        if "Source Results" in wb.sheetnames:
            del wb["Source Results"]
        ws2 = wb.create_sheet("Source Results")
        headers2 = ["Company", "Adapter", "Source Key", "Verdict", "Jobs", "Enriched", "Scored", "Failure Type", "Error", "Board URL"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")

        verdict_fills = {
            "ok":      PatternFill("solid", start_color="C6EFCE"),
            "empty":   PatternFill("solid", start_color="FFEB9C"),
            "FAIL":    PatternFill("solid", start_color="FFC7CE"),
            "PARTIAL": PatternFill("solid", start_color="FFD699"),
        }

        # Build lookup from source_results for verdicts/errors
        _sr_lookup: dict[str, dict] = {}
        if source_results:
            for sr in source_results:
                _sr_lookup[sr.get("source_key", "")] = sr

        with SessionLocal() as db_session:
            all_sources = list(db_session.execute(
                select(Source).where(Source.active.is_(True)).order_by(Source.adapter_name, Source.employer_name)
            ).scalars())

            row_idx = 2
            for src in all_sources:
                raw_count = db_session.execute(
                    select(sa_func.count()).select_from(RawJob).where(RawJob.source_id == src.id)
                ).scalar() or 0
                enriched_count = db_session.execute(
                    select(sa_func.count()).select_from(EnrichedJob)
                    .join(RawJob, EnrichedJob.raw_job_id == RawJob.id)
                    .where(RawJob.source_id == src.id)
                    .where(EnrichedJob.detail_fetch_status.notin_(["geo_filtered", "agency_filtered", "title_filtered"]))
                ).scalar() or 0
                scored_count = db_session.execute(
                    select(sa_func.count()).select_from(ScoreResult)
                    .join(EnrichedJob, ScoreResult.enriched_job_id == EnrichedJob.id)
                    .join(RawJob, EnrichedJob.raw_job_id == RawJob.id)
                    .where(RawJob.source_id == src.id)
                ).scalar() or 0

                sr = _sr_lookup.get(src.source_key, {})
                sr_verdict = sr.get("verdict", "")
                error_msg = sr.get("error", "")
                if sr_verdict in ("FAIL", "PARTIAL"):
                    verdict = sr_verdict
                elif raw_count > 0:
                    verdict = "ok"
                else:
                    verdict = "empty"

                # Classify failure type from error message
                failure_type = ""
                if error_msg:
                    err_lower = error_msg.lower()
                    if "cloudflare" in err_lower or "just a moment" in err_lower:
                        failure_type = "Bot Detection (Cloudflare)"
                    elif "incapsula" in err_lower or "imperva" in err_lower:
                        failure_type = "Bot Detection (Incapsula)"
                    elif "captcha" in err_lower or "recaptcha" in err_lower or "hcaptcha" in err_lower:
                        failure_type = "Bot Detection (CAPTCHA)"
                    elif "access denied" in err_lower or "bot" in err_lower or "blocked" in err_lower or "forbidden" in err_lower and "403" not in err_lower:
                        failure_type = "Bot Detection"
                    elif "timeout" in err_lower:
                        failure_type = "Timeout"
                    elif "404" in err_lower:
                        failure_type = "HTTP 404"
                    elif "403" in err_lower:
                        failure_type = "HTTP 403"
                    elif "401" in err_lower:
                        failure_type = "HTTP 401"
                    elif "429" in err_lower:
                        failure_type = "Rate Limited"
                    elif "500" in err_lower or "502" in err_lower or "503" in err_lower:
                        failure_type = "Server Error"
                    elif "connect" in err_lower or "dns" in err_lower or "resolve" in err_lower:
                        failure_type = "Connection Error"
                    elif "ssl" in err_lower or "certificate" in err_lower:
                        failure_type = "SSL Error"
                    elif error_msg:
                        failure_type = "Other"

                board_url = src.base_url or ""

                ws2.cell(row=row_idx, column=1, value=src.employer_name).font = rfont
                ws2.cell(row=row_idx, column=2, value=src.adapter_name).font = rfont
                ws2.cell(row=row_idx, column=3, value=src.source_key).font = rfont
                vc = ws2.cell(row=row_idx, column=4, value=verdict)
                vc.font = rfont
                vc.fill = verdict_fills.get(verdict, PatternFill())
                ws2.cell(row=row_idx, column=5, value=raw_count).font = rfont
                ws2.cell(row=row_idx, column=6, value=enriched_count).font = rfont
                ws2.cell(row=row_idx, column=7, value=scored_count).font = rfont
                ft_cell = ws2.cell(row=row_idx, column=8, value=failure_type)
                ft_cell.font = rfont
                if failure_type:
                    _amber = PatternFill("solid", start_color="FFD699")
                    _red = PatternFill("solid", start_color="FFC7CE")
                    ft_fills = {
                        "Timeout": _amber,
                        "HTTP 404": _red,
                        "HTTP 403": _red,
                        "HTTP 401": _red,
                        "Rate Limited": _amber,
                        "Server Error": _red,
                        "Connection Error": _red,
                        "SSL Error": _red,
                        "Bot Detection": _amber,
                        "Bot Detection (Cloudflare)": _amber,
                        "Bot Detection (Incapsula)": _amber,
                        "Bot Detection (CAPTCHA)": _amber,
                    }
                    ft_cell.fill = ft_fills.get(failure_type, PatternFill())
                ws2.cell(row=row_idx, column=9, value=error_msg).font = rfont
                ws2.cell(row=row_idx, column=10, value=board_url).font = rfont
                row_idx += 1

        for col, w in {"A": 35, "B": 18, "C": 45, "D": 10, "E": 8, "F": 10, "G": 10, "H": 18, "I": 60, "J": 60}.items():
            ws2.column_dimensions[col].width = w
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:J{row_idx - 1}"

        # --- Adapter Coverage sheet (one row per adapter: did it run this session?) ---
        # Rebuilt every snapshot so the final save of the run shows the final picture.
        if "Adapter Coverage" in wb.sheetnames:
            del wb["Adapter Coverage"]
        ws3 = wb.create_sheet("Adapter Coverage")
        headers3 = [
            "Adapter",
            "Active Sources",
            "Sources Ran This Session",
            "Sources Ran Ever",
            "Never Ran",
            "Last Run (any source)",
            "Raw Jobs This Session",
            "Success Rate",
            "Status",
        ]
        for col, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")

        # Status colour legend
        _status_fills = {
            "✓ Ran":           PatternFill("solid", start_color="C6EFCE"),   # green
            "⚠ Partial":      PatternFill("solid", start_color="FFEB9C"),   # amber
            "✗ Did Not Run":   PatternFill("solid", start_color="FFC7CE"),   # red
            "— No Sources":    PatternFill("solid", start_color="D9D9D9"),   # grey
        }

        # Session-window start: treat this run as "everything since the earliest
        # SourceRun that matches any source_key in the source_results input".
        # Fallback: anything in the last 12 hours.
        from datetime import datetime as _dt, timedelta as _td, timezone as _tz
        session_start = _dt.now(_tz.utc) - _td(hours=12)
        if source_results:
            # Use the oldest SourceRun timestamp among the source_keys we just touched
            source_keys_in_session = [sr.get("source_key") for sr in source_results if sr.get("source_key")]
            if source_keys_in_session:
                from vacancysoft.db.models import SourceRun as _SR
                with SessionLocal() as _sess:
                    earliest = _sess.execute(
                        select(sa_func.min(_SR.created_at))
                        .join(Source, Source.id == _SR.source_id)
                        .where(Source.source_key.in_(source_keys_in_session))
                    ).scalar()
                    if earliest:
                        # Small cushion so boundary runs aren't excluded
                        session_start = earliest - _td(seconds=30)

        with SessionLocal() as db_session:
            from vacancysoft.db.models import SourceRun as _SR

            # Active-source counts per adapter
            active_by_adapter = dict(db_session.execute(
                select(Source.adapter_name, sa_func.count())
                .where(Source.active.is_(True))
                .group_by(Source.adapter_name)
            ).all())

            # Ever-ran counts (distinct source IDs with at least one SourceRun)
            ever_ran = dict(db_session.execute(
                select(Source.adapter_name, sa_func.count(sa_func.distinct(Source.id)))
                .join(_SR, _SR.source_id == Source.id)
                .where(Source.active.is_(True))
                .group_by(Source.adapter_name)
            ).all())

            # This-session ran counts + raw-jobs created
            session_ran = dict(db_session.execute(
                select(Source.adapter_name, sa_func.count(sa_func.distinct(Source.id)))
                .join(_SR, _SR.source_id == Source.id)
                .where(Source.active.is_(True), _SR.created_at >= session_start)
                .group_by(Source.adapter_name)
            ).all())

            session_raw = dict(db_session.execute(
                select(Source.adapter_name, sa_func.sum(sa_func.coalesce(_SR.raw_jobs_created, 0)))
                .join(_SR, _SR.source_id == Source.id)
                .where(Source.active.is_(True), _SR.created_at >= session_start)
                .group_by(Source.adapter_name)
            ).all())

            session_success = dict(db_session.execute(
                select(Source.adapter_name, sa_func.count())
                .join(_SR, _SR.source_id == Source.id)
                .where(Source.active.is_(True), _SR.created_at >= session_start, _SR.status == "success")
                .group_by(Source.adapter_name)
            ).all())

            session_error = dict(db_session.execute(
                select(Source.adapter_name, sa_func.count())
                .join(_SR, _SR.source_id == Source.id)
                .where(Source.active.is_(True), _SR.created_at >= session_start, _SR.status == "error")
                .group_by(Source.adapter_name)
            ).all())

            last_run = dict(db_session.execute(
                select(Source.adapter_name, sa_func.max(_SR.created_at))
                .join(_SR, _SR.source_id == Source.id)
                .where(Source.active.is_(True))
                .group_by(Source.adapter_name)
            ).all())

        row_idx = 2
        adapters_sorted = sorted(active_by_adapter.keys())
        for adapter_name in adapters_sorted:
            active_n = active_by_adapter.get(adapter_name, 0)
            session_n = session_ran.get(adapter_name, 0)
            ever_n = ever_ran.get(adapter_name, 0)
            never_n = active_n - ever_n
            last_ts = last_run.get(adapter_name)
            raw_n = int(session_raw.get(adapter_name) or 0)
            ok_n = session_success.get(adapter_name, 0)
            err_n = session_error.get(adapter_name, 0)
            total_n = ok_n + err_n
            success_rate = f"{(ok_n / total_n) * 100:.0f}%" if total_n > 0 else "—"

            if active_n == 0:
                status = "— No Sources"
            elif session_n == 0:
                status = "✗ Did Not Run"
            elif session_n < active_n:
                status = "⚠ Partial"
            else:
                status = "✓ Ran"

            ws3.cell(row=row_idx, column=1, value=adapter_name).font = rfont
            ws3.cell(row=row_idx, column=2, value=active_n).font = rfont
            ws3.cell(row=row_idx, column=3, value=session_n).font = rfont
            ws3.cell(row=row_idx, column=4, value=ever_n).font = rfont
            ws3.cell(row=row_idx, column=5, value=never_n).font = rfont
            ws3.cell(row=row_idx, column=6, value=last_ts.strftime("%Y-%m-%d %H:%M") if last_ts else "never").font = rfont
            ws3.cell(row=row_idx, column=7, value=raw_n).font = rfont
            ws3.cell(row=row_idx, column=8, value=success_rate).font = rfont
            sc = ws3.cell(row=row_idx, column=9, value=status)
            sc.font = rfont
            sc.fill = _status_fills.get(status, PatternFill())
            row_idx += 1

        for col, w in {"A": 20, "B": 16, "C": 22, "D": 16, "E": 12, "F": 20, "G": 20, "H": 14, "I": 18}.items():
            ws3.column_dimensions[col].width = w
        ws3.freeze_panes = "A2"
        if row_idx > 2:
            ws3.auto_filter.ref = f"A1:I{row_idx - 1}"

        wb.save(output_path)
    except Exception as exc:
        typer.echo(f"  [snapshot] Excel write failed: {type(exc).__name__}: {exc}")


# Distinct colour per adapter so progress bars are visually distinguishable
_ADAPTER_COLOURS: dict[str, str] = {
    "workday":          "bright_blue",
    "greenhouse":       "bright_green",
    "workable":         "bright_yellow",
    "ashby":            "bright_magenta",
    "smartrecruiters":  "bright_cyan",
    "lever":            "bright_red",
    "icims":            "dark_orange",
    "oracle":           "medium_purple1",
    "successfactors":   "deep_sky_blue1",
    "eightfold":        "spring_green1",
    "generic_site":     "gold1",
    "adzuna":           "salmon1",
    "efinancialcareers":"turquoise2",
    "reed":             "orchid1",
    "google_jobs":      "dodger_blue1",
}


@pipeline_app.command("run")
def run_pipeline(
    adapter: str | None = typer.Option(None, "--adapter", help="Only discover from this adapter type"),
    source_key: str | None = typer.Option(None, "--source-key", help="Only discover from this source"),
    seed_type: str | None = typer.Option(None, "--seed-type", help="Only discover from sources with this seed_type (e.g. psl_import, manual_add)"),
    unscraped: bool = typer.Option(False, "--unscraped", help="Only discover from sources that have never been scraped"),
    discover_limit: int = typer.Option(0, "--discover-limit", help="Max total sources to discover from (0 = all)"),
    per_adapter: int = typer.Option(0, "--per-adapter", help="Max sources per adapter (0 = all). Picks N from each adapter for even coverage"),
    export_profile: str = typer.Option("accepted_plus_review", "--export-profile"),
    output: str | None = typer.Option("leads_output.xlsx", "--output", help="Excel output path"),
    dry_run: bool = typer.Option(False, "--dry-run", help="Skip webhook send"),
    snapshot_every: int = typer.Option(1, "--snapshot-every", help="Save Excel every N new scored leads"),
) -> None:
    """Run the full pipeline: discover → enrich → classify → score → export."""
    from collections import defaultdict
    from rich.console import Console
    from rich.live import Live
    from rich.table import Table
    from rich.progress import Progress, BarColumn, TextColumn, SpinnerColumn, TimeElapsedColumn, TaskID
    from rich.panel import Panel
    from rich.text import Text

    console = Console()

    # Load sources
    with SessionLocal() as session:
        query = select(Source).where(Source.active.is_(True))
        if source_key:
            query = select(Source).where(Source.source_key == source_key)
        elif adapter:
            query = query.where(Source.adapter_name == adapter)
        if seed_type:
            query = query.where(Source.seed_type == seed_type)
        all_sources = list(session.execute(query).scalars())

        # Filter to unscraped only
        if unscraped:
            scraped_ids = set(r[0] for r in session.execute(select(SourceRun.source_id).distinct()).all())
            all_sources = [s for s in all_sources if s.id not in scraped_ids]

    # Apply --per-adapter: pick N sources from each adapter for even spread
    if per_adapter > 0:
        by_adapter: dict[str, list] = defaultdict(list)
        for s in all_sources:
            by_adapter[s.adapter_name].append(s)
        sources = []
        for adapter_name in sorted(by_adapter.keys()):
            sources.extend(by_adapter[adapter_name][:per_adapter])
    else:
        sources = all_sources

    # Apply --discover-limit on top (total cap)
    if discover_limit > 0:
        sources = sources[:discover_limit]

    # Group sources by adapter for per-adapter progress bars
    adapter_sources: dict[str, list] = defaultdict(list)
    for s in sources:
        adapter_sources[s.adapter_name].append(s)

    # Execution order: API adapters first, browser adapters second, generic last
    _API_ADAPTERS = {"workday", "greenhouse", "workable", "ashby", "smartrecruiters", "lever", "adzuna", "reed", "google_jobs", "silkroad", "taleo"}
    _BROWSER_ADAPTERS = {"icims", "oracle", "successfactors", "eightfold", "efinancialcareers", "hibob", "selectminds"}
    _GENERIC_ADAPTERS = {"generic_site"}

    def _adapter_priority(name: str) -> int:
        if name in _API_ADAPTERS:
            return 0
        if name in _BROWSER_ADAPTERS:
            return 1
        if name in _GENERIC_ADAPTERS:
            return 2
        return 3  # anything unknown goes last

    adapter_order = sorted(adapter_sources.keys(), key=_adapter_priority)

    console.print(Panel(
        f"[bold]Sources:[/bold] {len(sources)}  |  "
        f"[bold]Adapters:[/bold] {len(adapter_order)}  |  "
        f"[bold]Snapshot every:[/bold] {snapshot_every} leads  |  "
        f"[bold]Output:[/bold] {output or 'none'}",
        title="[bold bright_white]Pipeline Run[/bold bright_white]",
        border_style="bright_blue",
    ))

    # Build progress display
    progress = Progress(
        SpinnerColumn(),
        TextColumn("[bold]{task.description}[/bold]"),
        BarColumn(bar_width=30),
        TextColumn("{task.completed}/{task.total}"),
        TextColumn("[dim]{task.fields[status]}[/dim]"),
        TimeElapsedColumn(),
        console=console,
    )

    # Overall progress bar
    overall_task = progress.add_task(
        "[bold white]Overall", total=len(sources), status="starting...",
    )

    # Per-adapter progress bars
    adapter_tasks: dict[str, TaskID] = {}
    for adapter_name in adapter_order:
        colour = _ADAPTER_COLOURS.get(adapter_name, "white")
        count = len(adapter_sources[adapter_name])
        task_id = progress.add_task(
            f"[{colour}]{adapter_name}[/{colour}]",
            total=count,
            status="waiting",
        )
        adapter_tasks[adapter_name] = task_id

    # Counters
    ok_count = 0
    fail_count = 0
    total_jobs = 0
    total_enriched = 0
    total_backfilled = 0
    total_classified = 0
    total_scored = 0
    scored_since_snapshot = 0
    scrape_log: list[tuple] = []
    log_lines: list[str] = []

    # Full per-source result log for the Excel report
    source_results: list[dict] = []

    def _log(msg: str) -> None:
        log_lines.append(msg)
        # Keep only last 8 lines for the live panel
        if len(log_lines) > 8:
            log_lines.pop(0)

    with progress:
        for adapter_name in adapter_order:
            colour = _ADAPTER_COLOURS.get(adapter_name, "white")
            task_id = adapter_tasks[adapter_name]
            progress.update(task_id, status="running")

            # ── Batch mode for generic_site: share Playwright across 15 boards ──
            if adapter_name == "generic_site":
                from vacancysoft.adapters.generic_browser import GenericBrowserAdapter

                batch_configs = []
                batch_sources = []
                for source in adapter_sources[adapter_name]:
                    config = dict(source.config_blob or {})
                    config.setdefault("company", source.employer_name)

                    def _make_page_cb(src):
                        def _cb(page_num, new_records, all_recs):
                            nonlocal total_jobs, _page_jobs_counted
                            _page_jobs_counted += len(new_records)
                            total_jobs += len(new_records)
                            import sys
                            from datetime import datetime as _dt
                            scrape_log.append((_dt.now().strftime("%Y-%m-%d %H:%M:%S"), src.employer_name, page_num, len(new_records), len(all_recs), None))
                            print(f"[{src.employer_name}] page {page_num}: +{len(new_records)} new, {len(all_recs)} total", file=sys.stderr, flush=True)
                            with SessionLocal() as sess:
                                s = sess.execute(select(Source).where(Source.source_key == src.source_key)).scalar_one()
                                persist_discovery_batch(session=sess, source=s, records=new_records, trigger="pipeline_discover")
                            progress.update(overall_task, status=f"{src.employer_name} p{page_num} | {total_jobs} jobs, {total_scored} scored")
                        return _cb

                    config["_on_page_scraped"] = _make_page_cb(source)
                    batch_configs.append(config)
                    batch_sources.append(source)

                _page_jobs_counted = 0

                def _on_board_done(config, page_result, error):
                    nonlocal ok_count, fail_count, total_jobs, total_enriched, total_backfilled, total_classified, total_scored, scored_since_snapshot, _page_jobs_counted
                    idx = batch_configs.index(config)
                    src = batch_sources[idx]

                    if error:
                        _log(f"[red]FAIL[/red] {src.source_key}: {type(error).__name__}: {error}")
                        fail_count += 1

                        if _page_jobs_counted > 0:
                            console.print(
                                f"[yellow]PARTIAL [{ok_count + fail_count}/{len(sources)}][/yellow] "
                                f"{src.employer_name:40s} | {src.adapter_name:18s} | "
                                f"{_page_jobs_counted} jobs before {type(error).__name__}, enriching..."
                            )
                            with SessionLocal() as session:
                                e = enrich_raw_jobs(session, limit=None); total_enriched += e
                            with SessionLocal() as session:
                                d = backfill_detail_for_enriched_jobs(session, limit=50, concurrency=5); total_backfilled += d
                            with SessionLocal() as session:
                                c = classify_enriched_jobs(session, limit=None); total_classified += c
                            with SessionLocal() as session:
                                s = score_enriched_jobs(session, limit=None); total_scored += s; scored_since_snapshot += s
                        else:
                            console.print(
                                f"[red]FAIL  [{ok_count + fail_count}/{len(sources)}][/red] "
                                f"{src.employer_name:40s} | {src.adapter_name:18s} | "
                                f"{type(error).__name__}: {str(error)[:80]}"
                            )

                        source_results.append({
                            "company": src.employer_name, "adapter": src.adapter_name,
                            "source_key": src.source_key,
                            "verdict": "PARTIAL" if _page_jobs_counted > 0 else "FAIL",
                            "jobs": _page_jobs_counted, "enriched": 0, "backfilled": 0,
                            "classified": 0, "scored": 0,
                            "error": f"{type(error).__name__}: {error}",
                        })
                    else:
                        job_count = len(page_result.jobs) if page_result else 0
                        remaining = job_count - _page_jobs_counted
                        if remaining > 0 and page_result and page_result.jobs:
                            total_jobs += remaining
                            with SessionLocal() as session:
                                source_obj = session.execute(select(Source).where(Source.source_key == src.source_key)).scalar_one()
                                persist_discovery_batch(session=session, source=source_obj, records=page_result.jobs, trigger="pipeline_discover")

                        ok_count += 1
                        with SessionLocal() as session:
                            e = enrich_raw_jobs(session, limit=None); total_enriched += e
                        with SessionLocal() as session:
                            d = backfill_detail_for_enriched_jobs(session, limit=50, concurrency=5); total_backfilled += d
                        with SessionLocal() as session:
                            c = classify_enriched_jobs(session, limit=None); total_classified += c
                        with SessionLocal() as session:
                            s = score_enriched_jobs(session, limit=None); total_scored += s; scored_since_snapshot += s

                        _log(f"[{colour}]OK[/{colour}] {src.employer_name}: {job_count} jobs, +{e} enriched, +{c} classified, +{s} scored")
                        source_results.append({
                            "company": src.employer_name, "adapter": src.adapter_name,
                            "source_key": src.source_key, "verdict": "ok",
                            "jobs": job_count, "enriched": e, "backfilled": d,
                            "classified": c, "scored": s, "error": "",
                        })
                        console.print(
                            f"[{colour}]OK    [{ok_count + fail_count}/{len(sources)}][/{colour}] "
                            f"{src.employer_name:40s} | {src.adapter_name:18s} | "
                            f"jobs={job_count:<4d} enriched={e:<4d} classified={c:<4d} scored={s:<4d}"
                        )

                    _page_jobs_counted = 0
                    progress.advance(task_id)
                    progress.advance(overall_task)
                    progress.update(overall_task, status=f"done {src.employer_name} | {total_jobs} jobs, {total_scored} scored")

                    if output and total_scored > 0 and scored_since_snapshot >= snapshot_every:
                        _snapshot_excel(output, export_profile, scrape_log=scrape_log, source_results=source_results)
                        scored_since_snapshot = 0
                    elif output and total_scored > 0:
                        _snapshot_excel(output, export_profile, scrape_log=scrape_log, source_results=source_results)

                # Run batched discovery
                asyncio.run(GenericBrowserAdapter().discover_batch(
                    batch_configs, batch_size=15, on_board_complete=_on_board_done,
                ))

                # Snapshot after generic finishes
                if output:
                    _snapshot_excel(output, export_profile, scrape_log=scrape_log, source_results=source_results)
                progress.update(task_id, status=f"[bold {colour}]done[/bold {colour}] — {total_jobs} jobs")
                continue

            for source in adapter_sources[adapter_name]:
                adapter_cls = ADAPTER_REGISTRY.get(source.adapter_name)
                if adapter_cls is None:
                    _log(f"[yellow]SKIP[/yellow] {source.source_key}: unknown adapter")
                    fail_count += 1
                    source_results.append({
                        "company": source.employer_name,
                        "adapter": source.adapter_name,
                        "source_key": source.source_key,
                        "verdict": "SKIP",
                        "jobs": 0,
                        "enriched": 0,
                        "backfilled": 0,
                        "classified": 0,
                        "scored": 0,
                        "error": f"unknown adapter '{source.adapter_name}'",
                    })
                    console.print(
                        f"[yellow]SKIP  [{ok_count + fail_count}/{len(sources)}][/yellow] "
                        f"{source.employer_name:40s} | {source.adapter_name:18s} | "
                        f"unknown adapter"
                    )
                    progress.advance(task_id)
                    progress.advance(overall_task)
                    continue

                config = dict(source.config_blob or {})
                config.setdefault("company", source.employer_name)

                progress.update(overall_task, status=f"{source.employer_name} | {total_jobs} jobs, {total_scored} scored")
                progress.update(task_id, status=f"{source.employer_name}")

                # --- Discover ---
                try:
                    adapter_instance = adapter_cls()

                    # Incremental callback: after every page of results from ANY
                    # adapter, persist → enrich → classify → score → write Excel.
                    _page_jobs_counted = 0  # track how many jobs the callback handled

                    def _on_page_scraped(page_num: int, new_records: list, all_recs: list) -> None:
                        nonlocal total_jobs, _page_jobs_counted
                        _page_jobs_counted += len(new_records)
                        total_jobs += len(new_records)

                        # Log entry
                        import sys
                        from datetime import datetime as _dt
                        log_entry = (_dt.now().strftime("%Y-%m-%d %H:%M:%S"), source.employer_name, page_num, len(new_records), len(all_recs), None)
                        scrape_log.append(log_entry)
                        print(f"[{source.employer_name}] page {page_num}: +{len(new_records)} new, {len(all_recs)} total", file=sys.stderr, flush=True)

                        # Persist the new page of raw jobs (fast — DB insert only)
                        with SessionLocal() as sess:
                            src = sess.execute(
                                select(Source).where(Source.source_key == source.source_key)
                            ).scalar_one()
                            persist_discovery_batch(
                                session=sess, source=src, records=new_records, trigger="pipeline_discover",
                            )

                        progress.update(overall_task, status=f"{source.employer_name} p{page_num} | {total_jobs} jobs, {total_scored} scored")

                    # Pass the callback to ALL adapters — those that paginate
                    # will fire it per-page; non-paginating ones just ignore it.
                    # Per-source timeout prevents any single board from blocking the run.
                    # Aggregators need longer — they make hundreds of API calls
                    _AGGREGATOR_ADAPTERS = {"adzuna", "reed", "efinancialcareers", "google_jobs"}
                    _SLOW_BROWSER_ADAPTERS = {"successfactors", "oracle_cloud", "icims", "generic_site"}
                    if source.adapter_name in _AGGREGATOR_ADAPTERS:
                        default_timeout = 900
                    elif source.adapter_name in _SLOW_BROWSER_ADAPTERS:
                        default_timeout = 300
                    else:
                        default_timeout = 120
                    source_timeout = float(config.get("source_timeout_seconds", default_timeout))

                    async def _discover_with_timeout():
                        return await asyncio.wait_for(
                            adapter_instance.discover(source_config=config, on_page_scraped=_on_page_scraped),
                            timeout=source_timeout,
                        )

                    page = asyncio.run(_discover_with_timeout())

                    # Count any jobs not already counted by the callback
                    remaining = len(page.jobs) - _page_jobs_counted
                    if remaining > 0:
                        total_jobs += remaining

                    job_count = len(page.jobs)

                    # Persist any jobs not already handled by the callback
                    if remaining > 0 and page.jobs:
                        with SessionLocal() as session:
                            source_obj = session.execute(
                                select(Source).where(Source.source_key == source.source_key)
                            ).scalar_one()
                            _run, raw_count = persist_discovery_batch(
                                session=session, source=source_obj, records=page.jobs, trigger="pipeline_discover",
                            )

                    ok_count += 1
                except Exception as exc:
                    _log(f"[red]FAIL[/red] {source.source_key}: {type(exc).__name__}: {exc}")
                    fail_count += 1

                    # Jobs may have been persisted by the callback before the timeout.
                    # Still run enrich/classify/score so they're not lost.
                    e = d = c = s = 0
                    if _page_jobs_counted > 0:
                        console.print(
                            f"[yellow]PARTIAL [{ok_count + fail_count}/{len(sources)}][/yellow] "
                            f"{source.employer_name:40s} | {source.adapter_name:18s} | "
                            f"{_page_jobs_counted} jobs persisted before {type(exc).__name__}, enriching..."
                        )
                        with SessionLocal() as session:
                            e = enrich_raw_jobs(session, limit=None)
                            total_enriched += e
                        with SessionLocal() as session:
                            d = backfill_detail_for_enriched_jobs(session, limit=50, concurrency=5)
                            total_backfilled += d
                        with SessionLocal() as session:
                            c = classify_enriched_jobs(session, limit=None)
                            total_classified += c
                        with SessionLocal() as session:
                            s = score_enriched_jobs(session, limit=None)
                            total_scored += s
                            scored_since_snapshot += s
                    else:
                        console.print(
                            f"[red]FAIL  [{ok_count + fail_count}/{len(sources)}][/red] "
                            f"{source.employer_name:40s} | {source.adapter_name:18s} | "
                            f"{type(exc).__name__}: {str(exc)[:80]}"
                        )

                    source_results.append({
                        "company": source.employer_name,
                        "adapter": source.adapter_name,
                        "source_key": source.source_key,
                        "verdict": "PARTIAL" if _page_jobs_counted > 0 else "FAIL",
                        "jobs": _page_jobs_counted,
                        "enriched": e,
                        "backfilled": d,
                        "classified": c,
                        "scored": s,
                        "error": f"{type(exc).__name__}: {exc}",
                    })
                    progress.advance(task_id)
                    progress.advance(overall_task)
                    continue

                # --- Enrich / Detail backfill / Classify / Score ---
                # Final pass to catch any records the callback didn't process.
                with SessionLocal() as session:
                    e = enrich_raw_jobs(session, limit=None)
                    total_enriched += e
                with SessionLocal() as session:
                    d = backfill_detail_for_enriched_jobs(session, limit=50, concurrency=5)
                    total_backfilled += d
                with SessionLocal() as session:
                    c = classify_enriched_jobs(session, limit=None)
                    total_classified += c
                with SessionLocal() as session:
                    s = score_enriched_jobs(session, limit=None)
                    total_scored += s
                    scored_since_snapshot += s

                _log(
                    f"[{colour}]OK[/{colour}] {source.employer_name}: "
                    f"{job_count} jobs, +{e} enriched, +{d} backfilled, +{c} classified, +{s} scored"
                )
                source_results.append({
                    "company": source.employer_name,
                    "adapter": source.adapter_name,
                    "source_key": source.source_key,
                    "verdict": "ok",
                    "jobs": job_count,
                    "enriched": e,
                    "backfilled": d,
                    "classified": c,
                    "scored": s,
                    "error": "",
                })
                console.print(
                    f"[{colour}]OK    [{ok_count + fail_count}/{len(sources)}][/{colour}] "
                    f"{source.employer_name:40s} | {source.adapter_name:18s} | "
                    f"jobs={job_count:<4d} enriched={e:<4d} classified={c:<4d} scored={s:<4d}"
                )

                progress.advance(task_id)
                progress.advance(overall_task)
                progress.update(overall_task, status=f"done {source.employer_name} | {total_jobs} jobs, {total_enriched} enriched, {total_scored} scored")

                # --- Incremental Excel snapshot ---
                if output and total_scored > 0 and scored_since_snapshot >= snapshot_every:
                    progress.update(overall_task, status="saving Excel snapshot...")
                    _snapshot_excel(output, export_profile, scrape_log=scrape_log, source_results=source_results)
                    _log(f"[bright_white]SNAPSHOT[/bright_white] Excel saved ({total_scored} total scored)")
                    scored_since_snapshot = 0
                elif output and total_scored > 0 and s > 0:
                    _snapshot_excel(output, export_profile, scrape_log=scrape_log, source_results=source_results)

            # Snapshot after each adapter group finishes
            if output:
                progress.update(overall_task, status=f"saving Excel ({adapter_name} done)...")
                _snapshot_excel(output, export_profile, scrape_log=scrape_log, source_results=source_results)
                _log(f"[bright_white]SNAPSHOT[/bright_white] Excel saved after {adapter_name}")
                scored_since_snapshot = 0

            # Mark adapter as done
            progress.update(task_id, status=f"[bold {colour}]done[/bold {colour}] — {total_jobs} jobs")

        progress.update(overall_task, status="exporting...")

    # --- Final export ---
    console.print()
    if output:
        _snapshot_excel(output, export_profile, scrape_log=scrape_log, source_results=source_results)
        console.print(f"[bold green]Excel written to {output}[/bold green]")

    with SessionLocal() as session:
        result = send_profile_to_webhook(
            session=session, profile_name=export_profile, limit=50000, dry_run=dry_run,
        )
    console.print(f"Webhook: {result}")

    # --- Summary table ---
    summary = Table(title="Pipeline Summary", border_style="bright_blue")
    summary.add_column("Metric", style="bold")
    summary.add_column("Value", justify="right")
    summary.add_row("Sources OK", f"[green]{ok_count}[/green]")
    summary.add_row("Sources Failed", f"[red]{fail_count}[/red]")
    summary.add_row("Raw Jobs", str(total_jobs))
    summary.add_row("Enriched", str(total_enriched))
    summary.add_row("Detail Backfilled", str(total_backfilled))
    summary.add_row("Classified", str(total_classified))
    summary.add_row("Scored", str(total_scored))
    summary.add_row("Output", output or "none")
    console.print(summary)

    # Per-adapter breakdown
    adapter_table = Table(title="Per-Adapter Breakdown", border_style="dim")
    adapter_table.add_column("Adapter", style="bold")
    adapter_table.add_column("Sources", justify="right")
    adapter_table.add_column("Colour")
    for a_name in adapter_order:
        colour = _ADAPTER_COLOURS.get(a_name, "white")
        count = len(adapter_sources[a_name])
        adapter_table.add_row(
            f"[{colour}]{a_name}[/{colour}]",
            str(count),
            f"[{colour}]{'━' * 12}[/{colour}]",
        )
    console.print(adapter_table)

    # Recent log lines
    if log_lines:
        console.print(Panel("\n".join(log_lines[-8:]), title="Recent Activity", border_style="dim"))


# ══════════════════════════════════════════════════════════════════════════
# Daily incremental run: scrape all → send only NEW leads to webhook
# ══════════════════════════════════════════════════════════════════════════

@pipeline_app.command("daily")
def daily_run(
    output: str | None = typer.Option("leads_output.xlsx", "--output"),
    dry_run: bool = typer.Option(False, "--dry-run", help="Skip webhook send"),
    export_profile: str = typer.Option("accepted_plus_review", "--export-profile"),
) -> None:
    """Daily incremental run: scrape all adapters, send only NEW leads to webhook.

    Designed for automated daily execution. Uses export_records table to track
    what's already been sent — only new leads since the last run are posted to
    the webhook.

    Typical cron usage:
        python3.13 src/vacancysoft/cli/app.py db seed-config-boards
        python3.13 src/vacancysoft/cli/app.py pipeline daily
    """
    from rich.console import Console
    from rich.panel import Panel

    console = Console()
    console.print(Panel(
        "[bold]Daily incremental pipeline[/bold]\n"
        "Scrape all adapters → Enrich → Classify → Score → Send NEW leads only",
        title="[bold bright_white]Daily Run[/bold bright_white]",
        border_style="bright_green",
    ))

    # Step 1: Run the full pipeline (no webhook — we handle that separately)
    console.print("\n[bold]Step 1:[/bold] Running full pipeline scrape...")
    from click import Context
    run_pipeline(
        adapter=None,
        source_key=None,
        discover_limit=0,
        per_adapter=0,
        export_profile=export_profile,
        output=output,
        dry_run=True,  # Always dry-run the old webhook — we use the new one below
        snapshot_every=1,
    )

    # Step 2: Send only NEW leads to webhook
    console.print("\n[bold]Step 2:[/bold] Sending new leads to webhook...")
    with SessionLocal() as session:
        result = send_new_leads_to_webhook(
            session=session,
            limit=50000,
            dry_run=dry_run,
        )

    if result.get("job_count", 0) == 0:
        console.print("[yellow]No new leads to send.[/yellow]")
    elif result.get("ok"):
        console.print(
            f"[bold green]Sent {result['job_count']} new leads to webhook[/bold green] "
            f"(batch: {result.get('batch_id', '?')})"
        )
    else:
        console.print(f"[bold red]Webhook failed:[/bold red] {result.get('error', '?')}")

    console.print(f"\nWebhook result: {result}")


@export_app.command("send-new")
def send_new(
    limit: int = typer.Option(50000, "--limit"),
    webhook_url: str | None = typer.Option(None, "--url"),
    dry_run: bool = typer.Option(False, "--dry-run"),
) -> None:
    """Send only NEW leads (not previously sent) to the webhook."""
    with SessionLocal() as session:
        result = send_new_leads_to_webhook(
            session=session,
            limit=limit,
            webhook_url=webhook_url,
            dry_run=dry_run,
        )
    typer.echo(str(result))


@export_app.command("taxonomy-preview")
def taxonomy_preview() -> None:
    with SessionLocal() as session:
        rows = list(
            session.execute(
                select(
                    EnrichedJob.title,
                    ClassificationResult.primary_taxonomy_key,
                    ClassificationResult.taxonomy_version,
                    ClassificationResult.decision,
                    ScoreResult.export_decision,
                    ScoreResult.export_eligibility_score,
                )
                .join(ClassificationResult, ClassificationResult.enriched_job_id == EnrichedJob.id)
                .join(ScoreResult, ScoreResult.enriched_job_id == EnrichedJob.id, isouter=True)
                .order_by(ClassificationResult.created_at.desc())
                .limit(10)
            )
        )
    if not rows:
        typer.echo("No classification results found")
        return
    for row in rows:
        typer.echo(
            f"title={row.title} taxonomy={row.primary_taxonomy_key} version={row.taxonomy_version} classification={row.decision} export={row.export_decision} score={row.export_eligibility_score}"
        )


@export_app.command("preview-view")
def preview_view(view_name: str = typer.Option(..., "--view"), limit: int = typer.Option(20, "--limit")) -> None:
    with SessionLocal() as session:
        if view_name == "accepted_only":
            rows = fetch_rows(session, accepted_only_query(), limit=limit)
        elif view_name == "accepted_plus_review":
            rows = fetch_rows(session, accepted_plus_review_query(), limit=limit)
        elif view_name == "grouped_by_taxonomy":
            rows = fetch_rows(session, grouped_by_taxonomy_query(), limit=limit)
        else:
            raise typer.BadParameter("Unknown view. Use accepted_only, accepted_plus_review, or grouped_by_taxonomy.")

    if not rows:
        typer.echo("No rows found")
        return
    for row in rows:
        typer.echo(str(row))


@export_app.command("preview-segment")
def preview_segment(segment_name: str = typer.Option(..., "--segment"), limit: int = typer.Option(20, "--limit")) -> None:
    config = load_exporter_config()
    with SessionLocal() as session:
        rows = fetch_rows(session, client_segment_query(segment_name, config), limit=limit)
    if not rows:
        typer.echo("No rows found")
        return
    for row in rows:
        typer.echo(str(row))


@export_app.command("json-profile")
def json_profile(profile_name: str = typer.Option(..., "--profile"), output_path: str = typer.Option(..., "--output"), limit: int = typer.Option(100, "--limit")) -> None:
    with SessionLocal() as session:
        output = export_profile_to_json(session, profile_name=profile_name, output_path=output_path, limit=limit)
    typer.echo(f"Wrote JSON export: {output}")


@export_app.command("json-segment")
def json_segment(segment_name: str = typer.Option(..., "--segment"), output_path: str = typer.Option(..., "--output"), limit: int = typer.Option(100, "--limit")) -> None:
    with SessionLocal() as session:
        output = export_segment_to_json(session, segment_name=segment_name, output_path=output_path, limit=limit)
    typer.echo(f"Wrote JSON export: {output}")


@export_app.command("excel-profile")
def excel_profile(profile_name: str = typer.Option(..., "--profile"), output_path: str = typer.Option(..., "--output"), limit: int = typer.Option(100, "--limit")) -> None:
    with SessionLocal() as session:
        output = export_profile_to_excel(session, profile_name=profile_name, output_path=output_path, limit=limit)
    typer.echo(f"Wrote Excel export: {output}")


@export_app.command("excel-segment")
def excel_segment(segment_name: str = typer.Option(..., "--segment"), output_path: str = typer.Option(..., "--output"), limit: int = typer.Option(100, "--limit")) -> None:
    with SessionLocal() as session:
        output = export_segment_to_excel(session, segment_name=segment_name, output_path=output_path, limit=limit)
    typer.echo(f"Wrote Excel export: {output}")


@export_app.command("webhook-profile")
def webhook_profile(
    profile_name: str = typer.Option(..., "--profile"),
    limit: int = typer.Option(100, "--limit"),
    webhook_url: str | None = typer.Option(None, "--url"),
    dry_run: bool = typer.Option(False, "--dry-run"),
) -> None:
    with SessionLocal() as session:
        result = send_profile_to_webhook(
            session=session,
            profile_name=profile_name,
            limit=limit,
            webhook_url=webhook_url,
            dry_run=dry_run,
        )
    typer.echo(str(result))


@export_app.command("webhook-segment")
def webhook_segment(
    segment_name: str = typer.Option(..., "--segment"),
    limit: int = typer.Option(100, "--limit"),
    webhook_url: str | None = typer.Option(None, "--url"),
    dry_run: bool = typer.Option(False, "--dry-run"),
) -> None:
    with SessionLocal() as session:
        result = send_segment_to_webhook(
            session=session,
            segment_name=segment_name,
            limit=limit,
            webhook_url=webhook_url,
            dry_run=dry_run,
        )
    typer.echo(str(result))


# ══════════════════════════════════════════════════════════════════════════
# Feature 9: Dashboard / Summary
# ══════════════════════════════════════════════════════════════════════════

@export_app.command("dashboard")
def dashboard() -> None:
    """Show pipeline summary: verdicts, categories, top employers, score distribution."""
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from sqlalchemy import func, case

    console = Console()

    with SessionLocal() as session:
        # ── Verdict counts ──
        verdict_rows = list(session.execute(
            select(ScoreResult.export_decision, func.count())
            .group_by(ScoreResult.export_decision)
        ))
        total_scored = sum(r[1] for r in verdict_rows)

        verdict_table = Table(title="Export Decisions", border_style="bright_blue")
        verdict_table.add_column("Decision", style="bold")
        verdict_table.add_column("Count", justify="right")
        verdict_table.add_column("% of Total", justify="right")
        colours = {"accepted": "green", "review": "yellow", "rejected": "red"}
        for decision, count in sorted(verdict_rows, key=lambda r: r[1], reverse=True):
            pct = f"{count / total_scored * 100:.1f}%" if total_scored else "—"
            c = colours.get(decision, "white")
            verdict_table.add_row(f"[{c}]{decision}[/{c}]", str(count), pct)
        verdict_table.add_row("[bold]Total[/bold]", f"[bold]{total_scored}[/bold]", "")
        console.print(verdict_table)

        # ── Category breakdown ──
        cat_rows = list(session.execute(
            select(
                ClassificationResult.primary_taxonomy_key,
                ScoreResult.export_decision,
                func.count(),
            )
            .join(ScoreResult, ScoreResult.enriched_job_id == ClassificationResult.enriched_job_id)
            .group_by(ClassificationResult.primary_taxonomy_key, ScoreResult.export_decision)
            .order_by(ClassificationResult.primary_taxonomy_key)
        ))
        cat_table = Table(title="Category Breakdown", border_style="cyan")
        cat_table.add_column("Category", style="bold")
        cat_table.add_column("Accepted", justify="right", style="green")
        cat_table.add_column("Review", justify="right", style="yellow")
        cat_table.add_column("Rejected", justify="right", style="red")
        cat_table.add_column("Total", justify="right", style="bold")

        from collections import defaultdict as _dd
        cat_agg: dict[str, dict[str, int]] = _dd(lambda: {"accepted": 0, "review": 0, "rejected": 0})
        for taxonomy, decision, count in cat_rows:
            cat_agg[taxonomy or "uncategorised"][decision] = count
        for cat in sorted(cat_agg.keys()):
            d = cat_agg[cat]
            cat_table.add_row(cat, str(d["accepted"]), str(d["review"]), str(d["rejected"]), str(sum(d.values())))
        console.print(cat_table)

        # ── Top employers ──
        emp_rows = list(session.execute(
            select(
                case(
                    (EnrichedJob.team.isnot(None), EnrichedJob.team),
                    else_=Source.employer_name,
                ).label("employer"),
                func.count().label("cnt"),
            )
            .join(ScoreResult, ScoreResult.enriched_job_id == EnrichedJob.id)
            .join(RawJob, EnrichedJob.raw_job_id == RawJob.id)
            .join(Source, RawJob.source_id == Source.id)
            .where(ScoreResult.export_decision.in_(["accepted", "review"]))
            .group_by("employer")
            .order_by(func.count().desc())
            .limit(20)
        ))
        emp_table = Table(title="Top 20 Employers (Accepted + Review)", border_style="green")
        emp_table.add_column("Employer", style="bold")
        emp_table.add_column("Jobs", justify="right")
        for employer, cnt in emp_rows:
            emp_table.add_row(employer or "—", str(cnt))
        console.print(emp_table)

        # ── Country breakdown ──
        geo_rows = list(session.execute(
            select(
                EnrichedJob.location_country,
                ScoreResult.export_decision,
                func.count(),
            )
            .join(ScoreResult, ScoreResult.enriched_job_id == EnrichedJob.id)
            .group_by(EnrichedJob.location_country, ScoreResult.export_decision)
            .order_by(func.count().desc())
        ))
        geo_agg: dict[str, dict[str, int]] = _dd(lambda: {"accepted": 0, "review": 0, "rejected": 0})
        for country, decision, count in geo_rows:
            geo_agg[country or "Unknown"][decision] = count
        geo_table = Table(title="Country Breakdown", border_style="magenta")
        geo_table.add_column("Country", style="bold")
        geo_table.add_column("Accepted", justify="right", style="green")
        geo_table.add_column("Review", justify="right", style="yellow")
        geo_table.add_column("Rejected", justify="right", style="red")
        geo_table.add_column("Total", justify="right", style="bold")
        for country in sorted(geo_agg.keys(), key=lambda c: sum(geo_agg[c].values()), reverse=True):
            d = geo_agg[country]
            geo_table.add_row(country, str(d["accepted"]), str(d["review"]), str(d["rejected"]), str(sum(d.values())))
        console.print(geo_table)

        # ── Score distribution ──
        score_rows = list(session.execute(
            select(ScoreResult.export_eligibility_score).order_by(ScoreResult.export_eligibility_score)
        ))
        if score_rows:
            scores = [r[0] for r in score_rows]
            buckets = {"0.0-0.2": 0, "0.2-0.4": 0, "0.4-0.6": 0, "0.6-0.8": 0, "0.8-1.0": 0}
            for s in scores:
                if s < 0.2: buckets["0.0-0.2"] += 1
                elif s < 0.4: buckets["0.2-0.4"] += 1
                elif s < 0.6: buckets["0.4-0.6"] += 1
                elif s < 0.8: buckets["0.6-0.8"] += 1
                else: buckets["0.8-1.0"] += 1
            dist_table = Table(title="Score Distribution", border_style="dim")
            dist_table.add_column("Range")
            dist_table.add_column("Count", justify="right")
            dist_table.add_column("Bar")
            max_count = max(buckets.values()) or 1
            for bucket, count in buckets.items():
                bar = "█" * int(count / max_count * 30)
                dist_table.add_row(bucket, str(count), f"[bright_blue]{bar}[/bright_blue]")
            console.print(dist_table)

        # ── Pipeline totals ──
        raw_count = session.execute(select(func.count()).select_from(RawJob)).scalar() or 0
        enriched_count = session.execute(select(func.count()).select_from(EnrichedJob)).scalar() or 0
        filtered_count = session.execute(
            select(func.count()).select_from(EnrichedJob)
            .where(EnrichedJob.detail_fetch_status.in_(["geo_filtered", "agency_filtered", "title_filtered"]))
        ).scalar() or 0
        source_count = session.execute(select(func.count()).select_from(Source).where(Source.active.is_(True))).scalar() or 0

        summary = Table(title="Pipeline Totals", border_style="bright_blue")
        summary.add_column("Metric", style="bold")
        summary.add_column("Value", justify="right")
        summary.add_row("Active Sources", str(source_count))
        summary.add_row("Raw Jobs", str(raw_count))
        summary.add_row("Enriched Jobs", str(enriched_count))
        summary.add_row("Filtered Out", f"[dim]{filtered_count}[/dim]")
        summary.add_row("Scored", str(total_scored))
        console.print(summary)


# ══════════════════════════════════════════════════════════════════════════
# Feature 11: Review Queue Management + Feedback
# ══════════════════════════════════════════════════════════════════════════

queue_app = typer.Typer(help="Review queue commands")
app.add_typer(queue_app, name="queue")


@queue_app.command("list")
def queue_list(
    status: str = typer.Option("open", "--status", help="Filter by status: open, resolved, all"),
    limit: int = typer.Option(50, "--limit"),
) -> None:
    """List review queue items."""
    from rich.console import Console
    from rich.table import Table

    console = Console()

    with SessionLocal() as session:
        stmt = (
            select(ReviewQueueItem, EnrichedJob.title, EnrichedJob.team)
            .join(EnrichedJob, ReviewQueueItem.enriched_job_id == EnrichedJob.id)
            .order_by(ReviewQueueItem.priority.desc(), ReviewQueueItem.created_at.desc())
            .limit(limit)
        )
        if status != "all":
            stmt = stmt.where(ReviewQueueItem.status == status)

        rows = list(session.execute(stmt))

    if not rows:
        typer.echo(f"No {status} queue items found.")
        return

    table = Table(title=f"Review Queue ({status})", border_style="yellow")
    table.add_column("ID", style="dim", max_width=8)
    table.add_column("Priority", justify="right")
    table.add_column("Title", max_width=40)
    table.add_column("Employer", max_width=25)
    table.add_column("Reason")
    table.add_column("Status")
    table.add_column("Resolution")

    for item, title, employer in rows:
        table.add_row(
            item.id[:8],
            str(item.priority),
            (title or "—")[:40],
            (employer or "—")[:25],
            item.reason_code,
            item.status,
            item.resolution or "—",
        )
    console.print(table)


@queue_app.command("resolve")
def queue_resolve(
    item_id: str = typer.Argument(..., help="Queue item ID (or prefix)"),
    resolution: str = typer.Option(..., "--resolution", help="accepted, rejected, or free text"),
) -> None:
    """Resolve a review queue item."""
    from datetime import datetime as _dt

    with SessionLocal() as session:
        item = session.execute(
            select(ReviewQueueItem).where(ReviewQueueItem.id.startswith(item_id))
        ).scalar_one_or_none()

        if not item:
            typer.echo(f"No queue item found matching '{item_id}'")
            raise typer.Exit(1)

        item.status = "resolved"
        item.resolution = resolution
        item.resolved_at = _dt.utcnow()
        session.commit()
        typer.echo(f"Resolved {item.id[:8]} as '{resolution}'")


@queue_app.command("add")
def queue_add(
    enriched_job_id: str = typer.Argument(..., help="Enriched job ID"),
    reason: str = typer.Option("manual_review", "--reason"),
    priority: int = typer.Option(50, "--priority"),
) -> None:
    """Add a job to the review queue."""
    with SessionLocal() as session:
        job = session.execute(
            select(EnrichedJob).where(EnrichedJob.id == enriched_job_id)
        ).scalar_one_or_none()

        if not job:
            typer.echo(f"No enriched job found with ID '{enriched_job_id}'")
            raise typer.Exit(1)

        item = ReviewQueueItem(
            enriched_job_id=enriched_job_id,
            queue_type="manual",
            priority=priority,
            reason_code=reason,
            reason_summary=f"Manual review: {reason}",
            evidence_blob={"title": job.title, "employer": job.team},
        )
        session.add(item)
        session.commit()
        typer.echo(f"Added to review queue: {item.id[:8]}")


@pipeline_app.command("feedback")
def feedback_report() -> None:
    """Show classification feedback: resolved reviews that disagreed with original scoring."""
    from rich.console import Console
    from rich.table import Table
    from sqlalchemy import func

    console = Console()

    with SessionLocal() as session:
        # Find resolved queue items with their classification
        rows = list(session.execute(
            select(
                ReviewQueueItem,
                EnrichedJob.title,
                EnrichedJob.team,
                ClassificationResult.primary_taxonomy_key,
                ClassificationResult.decision,
                ScoreResult.export_decision,
            )
            .join(EnrichedJob, ReviewQueueItem.enriched_job_id == EnrichedJob.id)
            .join(ClassificationResult, ClassificationResult.enriched_job_id == EnrichedJob.id, isouter=True)
            .join(ScoreResult, ScoreResult.enriched_job_id == EnrichedJob.id, isouter=True)
            .where(ReviewQueueItem.status == "resolved")
            .order_by(ReviewQueueItem.resolved_at.desc())
            .limit(100)
        ))

    if not rows:
        typer.echo("No resolved review items found. Use 'queue resolve' to resolve items first.")
        return

    # Summary
    agreements = 0
    overrides = 0
    override_patterns: list[dict] = []

    table = Table(title="Feedback Report: Resolved Reviews", border_style="cyan")
    table.add_column("Title", max_width=35)
    table.add_column("Employer", max_width=20)
    table.add_column("Category")
    table.add_column("Auto Decision")
    table.add_column("Human Resolution")
    table.add_column("Agreement")

    for item, title, employer, taxonomy, auto_decision, export_decision in rows:
        human = item.resolution or "—"
        auto = export_decision or auto_decision or "—"

        # Check if human agrees with auto
        human_norm = human.strip().lower()
        auto_norm = auto.strip().lower()
        agrees = human_norm == auto_norm or (
            human_norm in ("accepted", "accept", "yes", "approve") and auto_norm == "accepted"
        ) or (
            human_norm in ("rejected", "reject", "no", "deny") and auto_norm == "rejected"
        )

        if agrees:
            agreements += 1
            agreement_str = "[green]Yes[/green]"
        else:
            overrides += 1
            agreement_str = "[red]OVERRIDE[/red]"
            override_patterns.append({
                "title": title,
                "employer": employer,
                "category": taxonomy,
                "auto": auto,
                "human": human,
            })

        table.add_row(
            (title or "—")[:35],
            (employer or "—")[:20],
            taxonomy or "—",
            auto,
            human,
            agreement_str,
        )

    console.print(table)
    console.print(f"\n[bold]Agreements:[/bold] {agreements}  |  [bold]Overrides:[/bold] {overrides}")

    if override_patterns:
        console.print("\n[bold yellow]Override Patterns (candidate rule adjustments):[/bold yellow]")
        for p in override_patterns:
            console.print(
                f"  Title: [cyan]{p['title']}[/cyan]  |  "
                f"Auto: [red]{p['auto']}[/red]  →  Human: [green]{p['human']}[/green]  |  "
                f"Category: {p['category']}"
            )


# ══════════════════════════════════════════════════════════════════════════
# Intelligence commands
# ══════════════════════════════════════════════════════════════════════════

intel_app = typer.Typer(help="Intelligence dossier and campaign commands")
app.add_typer(intel_app, name="intel")


@intel_app.command("dossier")
def intel_dossier(
    job_ref: str = typer.Argument(..., help="Job reference or enriched_job_id"),
) -> None:
    """Generate an intelligence dossier for a job."""
    from rich.console import Console
    from rich.panel import Panel

    console = Console()

    with SessionLocal() as session:
        # Try to find by external_job_id first, then by enriched_job_id
        row = session.execute(
            select(EnrichedJob.id)
            .join(RawJob, EnrichedJob.raw_job_id == RawJob.id)
            .where(RawJob.external_job_id == job_ref)
            .limit(1)
        ).scalar_one_or_none()

        if not row:
            row = session.execute(
                select(EnrichedJob.id).where(EnrichedJob.id == job_ref).limit(1)
            ).scalar_one_or_none()

        if not row:
            # Try matching by canonical_job_key
            row = session.execute(
                select(EnrichedJob.id)
                .where(EnrichedJob.canonical_job_key.contains(job_ref))
                .limit(1)
            ).scalar_one_or_none()

        if not row:
            console.print(f"[red]No enriched job found for ref: {job_ref}[/red]")
            raise typer.Exit(1)

        enriched_job_id = row
        console.print(f"Generating dossier for [cyan]{enriched_job_id}[/cyan]...")

        from vacancysoft.intelligence.dossier import generate_dossier
        dossier = asyncio.run(generate_dossier(enriched_job_id, session))

        console.print()
        console.print(Panel(f"[bold]Dossier ID:[/bold] {dossier.id}", title="Generated", border_style="green"))
        console.print(f"  Category: [cyan]{dossier.category_used}[/cyan]")
        console.print(f"  Model: {dossier.model_used}")
        console.print(f"  Tokens: {dossier.tokens_used}  |  Latency: {dossier.latency_ms}ms")
        console.print(f"  Lead Score: [bold]{dossier.lead_score}[/bold]/5")
        console.print()

        if dossier.company_context:
            console.print(Panel(dossier.company_context[:500], title="Company Context", border_style="blue"))

        if dossier.hiring_managers:
            console.print("\n[bold]Hiring Managers:[/bold]")
            for hm in dossier.hiring_managers:
                if isinstance(hm, dict):
                    console.print(f"  {hm.get('name', '?')} — {hm.get('title', '?')} [{hm.get('confidence', '?')}]")


@intel_app.command("campaign")
def intel_campaign(
    dossier_id: str = typer.Argument(..., help="Intelligence dossier ID"),
) -> None:
    """Generate campaign outreach emails from an existing dossier."""
    from rich.console import Console
    from rich.panel import Panel

    console = Console()

    with SessionLocal() as session:
        from vacancysoft.intelligence.campaign import generate_campaign
        campaign = asyncio.run(generate_campaign(dossier_id, session))

        console.print()
        console.print(Panel(f"[bold]Campaign ID:[/bold] {campaign.id}", title="Generated", border_style="green"))
        console.print(f"  Model: {campaign.model_used}")
        console.print(f"  Tokens: {campaign.tokens_used}  |  Latency: {campaign.latency_ms}ms")

        emails = campaign.outreach_emails or []
        console.print(f"  Emails: {len(emails)}")
        for email in emails:
            if isinstance(email, dict):
                seq = email.get("sequence", "?")
                subj = email.get("subject", "")
                body = email.get("body", "")
                console.print()
                console.print(Panel(
                    f"[bold]Subject:[/bold] {subj}\n\n{body[:300]}{'...' if len(body) > 300 else ''}",
                    title=f"Email {seq}",
                    border_style="cyan",
                ))


@intel_app.command("prompts")
def intel_prompts() -> None:
    """List available prompt categories and their versions."""
    from rich.console import Console
    from rich.table import Table

    from vacancysoft.intelligence.prompts.category_blocks import CATEGORY_BLOCKS
    from vacancysoft.intelligence.dossier import PROMPT_VERSION

    console = Console()
    table = Table(title=f"Prompt Library (template {PROMPT_VERSION})")
    table.add_column("Category", style="cyan")
    table.add_column("Research Scope")
    table.add_column("Outreach Angle")

    for cat, blocks in CATEGORY_BLOCKS.items():
        table.add_row(
            cat,
            blocks["research_scope"][:60] + "...",
            blocks["outreach_angle"][:60] + "...",
        )

    console.print(table)


if __name__ == "__main__":
    app()

