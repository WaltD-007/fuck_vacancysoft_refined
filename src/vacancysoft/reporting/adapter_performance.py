"""Rolling per-adapter performance workbook.

Maintains a single workbook (default: ``reports/adapter_performance.xlsx``)
with one tab per adapter plus a Summary tab. Each adapter tab keeps the
last ``max_runs`` run-blocks (default 3); oldest is dropped when a new
block arrives. Summary tab has one row per (adapter, run) with counts
and deltas vs. the prior run.

Replaces the old ``leads_<adapter>_<date>.xlsx`` pattern — that produced
~6MB per file per run and repeated the full lead set every time, with
no actual per-adapter performance view. Leads data is still available
on-demand via ``prospero export excel-profile``; this workbook is
strictly about adapter health over time.

### Input contract

``write_adapter_performance`` takes the ``source_results`` list that
``cli/app.py:run_pipeline`` already builds for the legacy snapshot.
Each entry is expected to have (minimum):

    {
      "source_key":  str,   # unique per source
      "company":     str,   # employer name
      "adapter":     str,   # should equal `adapter_name` arg
      "verdict":     str,   # "ok" | "empty" | "FAIL" | "PARTIAL"
      "jobs":        int,
      "enriched":    int,
      "classified":  int,
      "scored":      int,
      "error":       str,   # empty if verdict=="ok"
    }

Any missing key is tolerated — cell falls back to a sensible default.
"""
from __future__ import annotations

from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select

from vacancysoft.db.engine import SessionLocal
from vacancysoft.db.models import Source

# Sentinel that marks the start of a run-block on an adapter tab.
_BANNER_PREFIX = "=== Run "
_BANNER_SUFFIX = " ==="

_PERF_COLUMNS = [
    "Source Key", "Employer", "Verdict",
    "Jobs", "Enriched", "Classified", "Scored",
    "Failure Type", "Error", "Board URL",
]
_SUMMARY_COLUMNS = [
    "Run Date", "Adapter", "Sources",
    "OK", "FAIL", "Empty",
    "Jobs", "Enriched", "Scored",
    "Δ OK", "Δ Jobs",
]

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", start_color="1F3864")
_BANNER_FONT = Font(name="Arial", bold=True, size=11)
_BANNER_FILL = PatternFill("solid", start_color="D9E1F2")
_ROW_FONT = Font(name="Arial", size=10)

_VERDICT_FILLS = {
    "ok":      PatternFill("solid", start_color="C6EFCE"),
    "empty":   PatternFill("solid", start_color="FFEB9C"),
    "FAIL":    PatternFill("solid", start_color="FFC7CE"),
    "PARTIAL": PatternFill("solid", start_color="FFD699"),
}


def write_adapter_performance(
    *,
    adapter_name: str,
    run_started_at: datetime,
    source_results: Iterable[dict[str, Any]],
    workbook_path: str | Path,
    max_runs: int = 3,
) -> Path:
    """Upsert this run's rows onto ``adapter_name``'s tab; refresh Summary.

    Returns the path of the saved workbook.
    """
    output = Path(workbook_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    results = [r for r in source_results if r.get("adapter") == adapter_name]
    run_label = run_started_at.strftime("%Y-%m-%d %H:%M")
    banner = f"{_BANNER_PREFIX}{run_label}{_BANNER_SUFFIX}"

    # Enrich each result with Source.base_url so the UI doesn't need
    # a second lookup to click through to the job board.
    board_urls = _lookup_board_urls(r["source_key"] for r in results if r.get("source_key"))

    if output.exists():
        wb = load_workbook(output)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    tab = wb[adapter_name] if adapter_name in wb.sheetnames else wb.create_sheet(adapter_name)
    _rewrite_adapter_tab(tab, banner, results, board_urls, max_runs)

    summary = wb["Summary"] if "Summary" in wb.sheetnames else wb.create_sheet("Summary", 0)
    if wb.index(summary) != 0:
        wb.move_sheet(summary, offset=-wb.index(summary))
    _update_summary_tab(summary, adapter_name, run_label, _metrics(results), max_runs)

    wb.save(output)
    return output


# ── Helpers ────────────────────────────────────────────────────────────────

def _lookup_board_urls(source_keys: Iterable[str]) -> dict[str, str]:
    keys = [k for k in source_keys if k]
    if not keys:
        return {}
    try:
        with SessionLocal() as session:
            rows = session.execute(
                select(Source.source_key, Source.base_url).where(Source.source_key.in_(keys))
            ).all()
        return {k: (u or "") for k, u in rows}
    except Exception:
        # Report-writing must not fail a scrape if the DB is momentarily unavailable.
        return {}


def _metrics(results: list[dict[str, Any]]) -> dict[str, int]:
    ok = empty = fail = 0
    jobs = enriched = scored = 0
    for r in results:
        v = (r.get("verdict") or "").upper()
        if v == "OK":
            ok += 1
        elif v == "EMPTY":
            empty += 1
        elif v in ("FAIL", "ERROR", "PARTIAL"):
            fail += 1
        jobs += int(r.get("jobs") or 0)
        enriched += int(r.get("enriched") or 0)
        scored += int(r.get("scored") or 0)
    return {
        "sources": len(results), "ok": ok, "fail": fail, "empty": empty,
        "jobs": jobs, "enriched": enriched, "scored": scored,
    }


def _row_from_result(r: dict[str, Any], board_urls: dict[str, str]) -> list[Any]:
    err = str(r.get("error") or "")
    failure_type = ""
    if err:
        # "ValueError: ..." → failure_type = "ValueError"
        split = err.split(":", 1)
        if len(split) == 2 and split[0].strip().isidentifier():
            failure_type = split[0].strip()
        if len(err) > 200:
            err = err[:197] + "..."
    return [
        r.get("source_key", ""),
        r.get("company", ""),
        r.get("verdict", ""),
        int(r.get("jobs") or 0),
        int(r.get("enriched") or 0),
        int(r.get("classified") or 0),
        int(r.get("scored") or 0),
        failure_type,
        err,
        board_urls.get(r.get("source_key", ""), ""),
    ]


def _parse_blocks(tab: Worksheet) -> list[list[list[Any]]]:
    """Split existing tab content into banner-delimited blocks."""
    blocks: list[list[list[Any]]] = []
    current: list[list[Any]] = []
    for row in tab.iter_rows(values_only=True):
        row_vals = list(row)
        first = row_vals[0] if row_vals else None
        if isinstance(first, str) and first.startswith(_BANNER_PREFIX):
            if current:
                blocks.append(current)
            current = [row_vals]
        elif current:
            current.append(row_vals)
    if current:
        blocks.append(current)
    return blocks


def _rewrite_adapter_tab(
    tab: Worksheet,
    banner: str,
    results: list[dict[str, Any]],
    board_urls: dict[str, str],
    max_runs: int,
) -> None:
    existing_blocks = _parse_blocks(tab)
    new_block: list[list[Any]] = [[banner]] + [list(_PERF_COLUMNS)]
    new_block.extend(_row_from_result(r, board_urls) for r in results)
    new_block.append([])  # blank separator

    kept = [new_block] + existing_blocks[: max_runs - 1]

    # Clear and rewrite
    if tab.max_row and tab.max_row > 0:
        tab.delete_rows(1, tab.max_row)

    row_num = 1
    for block in kept:
        for cells in block:
            if not cells:
                row_num += 1
                continue
            for col, val in enumerate(cells, start=1):
                tab.cell(row=row_num, column=col, value=val)
            # Styling
            first = cells[0]
            if isinstance(first, str) and first.startswith(_BANNER_PREFIX):
                c = tab.cell(row=row_num, column=1)
                c.font = _BANNER_FONT
                c.fill = _BANNER_FILL
            elif cells == _PERF_COLUMNS:
                for col in range(1, len(_PERF_COLUMNS) + 1):
                    c = tab.cell(row=row_num, column=col)
                    c.font = _HEADER_FONT
                    c.fill = _HEADER_FILL
                    c.alignment = Alignment(horizontal="center")
            else:
                for col in range(1, len(cells) + 1):
                    tab.cell(row=row_num, column=col).font = _ROW_FONT
                verdict = cells[2] if len(cells) > 2 else ""
                fill = _VERDICT_FILLS.get(verdict)
                if fill:
                    tab.cell(row=row_num, column=3).fill = fill
            row_num += 1

    widths = [36, 30, 10, 8, 10, 12, 8, 18, 44, 44]
    for i, w in enumerate(widths, start=1):
        tab.column_dimensions[tab.cell(row=1, column=i).column_letter].width = w
    tab.freeze_panes = "A2"


def _update_summary_tab(
    tab: Worksheet, adapter_name: str, run_label: str, metrics: dict[str, int], max_runs: int
) -> None:
    # Pull existing rows (skip header) into a list
    existing: list[list[Any]] = []
    for i, row in enumerate(tab.iter_rows(values_only=True)):
        if i == 0:
            continue
        row_vals = list(row)
        if not row_vals or all(v is None for v in row_vals):
            continue
        # Pad / trim to summary column count
        if len(row_vals) < len(_SUMMARY_COLUMNS):
            row_vals += [""] * (len(_SUMMARY_COLUMNS) - len(row_vals))
        existing.append(row_vals[: len(_SUMMARY_COLUMNS)])

    # Drop any prior entry for the same (adapter, run_label) combo — re-runs
    # on the same minute should overwrite, not duplicate.
    existing = [r for r in existing if not (r[1] == adapter_name and r[0] == run_label)]

    new_row: list[Any] = [
        run_label, adapter_name, metrics["sources"],
        metrics["ok"], metrics["fail"], metrics["empty"],
        metrics["jobs"], metrics["enriched"], metrics["scored"],
        "", "",
    ]

    combined = existing + [new_row]

    # Group by adapter, keep only the most-recent max_runs per adapter, recompute deltas
    by_adapter: dict[str, list[list[Any]]] = {}
    for r in combined:
        by_adapter.setdefault(str(r[1]), []).append(r)
    pruned: list[list[Any]] = []
    for rows in by_adapter.values():
        rows.sort(key=lambda r: str(r[0]), reverse=True)
        rows = rows[:max_runs]
        for i, r in enumerate(rows):
            prior = rows[i + 1] if i + 1 < len(rows) else None
            if prior:
                try:
                    r[9] = int(r[3]) - int(prior[3])   # Δ OK
                    r[10] = int(r[6]) - int(prior[6])  # Δ Jobs
                except (TypeError, ValueError):
                    r[9] = r[10] = ""
            else:
                r[9] = r[10] = ""
        pruned.extend(rows)

    pruned.sort(key=lambda r: (str(r[0]), str(r[1])), reverse=True)

    # Rewrite tab
    if tab.max_row and tab.max_row > 0:
        tab.delete_rows(1, tab.max_row)
    for col, h in enumerate(_SUMMARY_COLUMNS, start=1):
        c = tab.cell(row=1, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for ridx, r in enumerate(pruned, start=2):
        for col, val in enumerate(r, start=1):
            tab.cell(row=ridx, column=col, value=val).font = _ROW_FONT

    widths = [18, 20, 10, 8, 8, 8, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths, start=1):
        tab.column_dimensions[tab.cell(row=1, column=i).column_letter].width = w
    tab.freeze_panes = "A2"
