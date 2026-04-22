#!/usr/bin/env python3
"""Per-adapter location-extraction audit.

Samples RawJob rows for a single adapter (or a single source_id) and
writes an XLSX that surfaces, for each failing row, the raw
``listing_payload`` (JSON the adapter captured) and ``raw_text_blob``
(HTML/text the adapter scanned) — so an engineer can eyeball which
field/attribute holds the location and why the current extraction code
missed it.

Written as the investigation tool for the 2026-04-22 location-enrichment
audit that flagged two scraper-side failures:

  * phenom — 97% of rows (66/68) have location_raw=NULL
  * generic_site — 4,545 rows with location_raw=NULL

Shipped separately from the adapter fix PRs (3A phenom, 3B generic_site)
so it can be re-run after each fix to confirm the failure shape has
shrunk. Also generic enough to aim at any adapter — run it preventively
against greenhouse/workday/lever to catch a regression early.

Sheets produced:

  * Summary — run params + headline counts (total rows, failing rows,
    distinct sources in the sample, top candidate JSON fields).
  * Sample — one row per sampled RawJob. Columns: source_key, adapter,
    employer, board_url, title, location_raw, discovered_url, a
    pretty-printed ``listing_payload`` (truncated to 4,000 chars), a
    truncated ``raw_text_blob``, and a pre-computed
    ``candidate_location_fields`` column listing any top-level or
    nested JSON key whose name contains "loc"/"city"/"region"/"country"/
    "site"/"office"/"address".
  * Field Frequency — how often each candidate JSON field name appears
    across the sampled rows. A quick win here: if 90% of phenom
    payloads have a ``workLocation`` field the adapter never checks,
    one-line fix lands the recovery.
  * Source Breakdown — for --only-failing runs, how many sampled rows
    each source contributed. Tells you whether the failure is
    concentrated in a few boards (fix each) or spread wide across
    many (platform-level fix).

Usage:
    # Sample 20 failing phenom rows
    python3 scripts/audit_adapter_locations.py \\
        --adapter phenom --sample 20 --only-failing \\
        --out .data/audits/phenom-failing.xlsx

    # Compare against 5 rows that DO have a location (what's different?)
    python3 scripts/audit_adapter_locations.py \\
        --adapter phenom --sample 5 \\
        --out .data/audits/phenom-mixed.xlsx

    # Drill into one board only (after the summary sheet flags it)
    python3 scripts/audit_adapter_locations.py \\
        --adapter generic_site --source-id 123 --sample 30 \\
        --out .data/audits/generic-source-123.xlsx

Non-destructive: reads only from ``sources`` + ``raw_jobs``. Never
writes back to the DB.
"""
from __future__ import annotations

import argparse
import json
import random
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from sqlalchemy import func, select  # noqa: E402

from vacancysoft.db.engine import SessionLocal  # noqa: E402
from vacancysoft.db.models import RawJob, Source  # noqa: E402


# ── Styling ────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="FFD0D0D0", end_color="FFD0D0D0", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

# Keys worth flagging when we recursively walk the listing_payload.
# Matched by `<key-lower>.contains(needle)`, so "workLocation" and
# "primaryLocations" both match "loc".
_CANDIDATE_FIELD_NEEDLES: tuple[str, ...] = (
    "loc", "city", "region", "country", "site", "office", "address",
    "place", "area", "venue", "geo",
)

_BLOB_TRUNCATE_CHARS = 4_000


def _style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autofit(ws, col_widths: dict[int, int]) -> None:
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ── JSON walker ────────────────────────────────────────────────────


def _walk_candidate_fields(payload: Any, prefix: str = "") -> list[tuple[str, Any]]:
    """Recursively yield (dotted-path, value) for any dict key whose
    name (case-insensitively) contains one of the location needles.

    The returned value is a short string preview — dicts/lists get
    rendered compactly (keys or first item) so the XLSX column stays
    scannable.
    """
    out: list[tuple[str, Any]] = []
    if isinstance(payload, dict):
        for key, value in payload.items():
            key_str = str(key)
            key_lower = key_str.lower()
            path = f"{prefix}.{key_str}" if prefix else key_str
            if any(needle in key_lower for needle in _CANDIDATE_FIELD_NEEDLES):
                out.append((path, _preview(value)))
            # Recurse into dict / list values regardless of whether
            # the parent key matched — location often lives a level
            # below a non-matching parent ("primaryDetails" → "city").
            if isinstance(value, (dict, list)):
                out.extend(_walk_candidate_fields(value, path))
    elif isinstance(payload, list):
        # Only walk the first two items — enough to spot the pattern
        # without blowing up the preview on long arrays.
        for idx, item in enumerate(payload[:2]):
            out.extend(_walk_candidate_fields(item, f"{prefix}[{idx}]"))
    return out


def _preview(value: Any) -> str:
    """Short human-readable preview for the Sample sheet."""
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        s = str(value)
        return s if len(s) <= 120 else s[:117] + "…"
    if isinstance(value, dict):
        keys = list(value.keys())[:5]
        return "{" + ", ".join(str(k) for k in keys) + ("…" if len(value) > 5 else "") + "}"
    if isinstance(value, list):
        return f"[list len={len(value)}]"
    return repr(value)[:120]


# ── Sampler ────────────────────────────────────────────────────────


def _sample_rows(
    adapter: str | None,
    source_id: int | None,
    only_failing: bool,
    sample_size: int,
) -> tuple[list[dict], dict]:
    """Pull the sample and a headline-counts dict.

    Returns (rows, counts) where rows is a list of flat dicts ready for
    writing to the Sample sheet, and counts has keys
    ``matching_total``, ``matching_failing``, ``matching_with_loc``,
    ``sampled``, ``source_key_breakdown`` (Counter).
    """
    with SessionLocal() as s:
        # Build the base filter
        base = select(RawJob).join(Source, RawJob.source_id == Source.id)
        if adapter:
            base = base.where(Source.adapter_name == adapter)
        if source_id is not None:
            base = base.where(Source.id == source_id)
        base_failing = base.where(RawJob.location_raw.is_(None))
        base_with_loc = base.where(RawJob.location_raw.is_not(None))

        matching_total = s.execute(
            select(func.count()).select_from(base.subquery())
        ).scalar_one()
        matching_failing = s.execute(
            select(func.count()).select_from(base_failing.subquery())
        ).scalar_one()
        matching_with_loc = matching_total - matching_failing

        # Pick which pool to sample from
        sample_source = base_failing if only_failing else base
        # ORDER BY RANDOM() is SQLite-portable; for very large pools it's
        # slow but we're capping at sample_size anyway.
        sampled_rows = s.execute(
            sample_source.order_by(func.random()).limit(sample_size)
        ).scalars().all()

        # Join back to Source for the display fields
        src_ids = {r.source_id for r in sampled_rows}
        sources_by_id = {
            src.id: src
            for src in s.execute(
                select(Source).where(Source.id.in_(src_ids))
            ).scalars().all()
        }

        rows: list[dict] = []
        source_key_counter: Counter = Counter()
        for raw in sampled_rows:
            src = sources_by_id.get(raw.source_id)
            source_key_counter[src.source_key if src else f"<src#{raw.source_id}>"] += 1

            listing_payload = raw.listing_payload
            payload_preview = ""
            candidate_fields: list[tuple[str, Any]] = []
            if listing_payload is not None:
                try:
                    payload_preview = json.dumps(listing_payload, indent=2, ensure_ascii=False, default=str)
                except (TypeError, ValueError):
                    payload_preview = repr(listing_payload)
                if len(payload_preview) > _BLOB_TRUNCATE_CHARS:
                    payload_preview = payload_preview[:_BLOB_TRUNCATE_CHARS].rstrip() + "\n… (truncated)"
                candidate_fields = _walk_candidate_fields(listing_payload)

            raw_blob = raw.raw_text_blob or ""
            if len(raw_blob) > _BLOB_TRUNCATE_CHARS:
                raw_blob = raw_blob[:_BLOB_TRUNCATE_CHARS].rstrip() + "\n… (truncated)"

            candidates_column = "\n".join(f"{path}: {preview}" for path, preview in candidate_fields[:30])
            if len(candidate_fields) > 30:
                candidates_column += f"\n… ({len(candidate_fields) - 30} more)"

            rows.append({
                "source_id": raw.source_id,
                "source_key": src.source_key if src else "",
                "adapter_name": src.adapter_name if src else "",
                "employer_name": src.employer_name if src else "",
                "board_url": src.base_url if src else "",
                "title_raw": raw.title_raw or "",
                "location_raw": raw.location_raw or "",
                "discovered_url": raw.discovered_url or "",
                "candidate_location_fields": candidates_column,
                "listing_payload_json": payload_preview,
                "raw_text_blob": raw_blob,
                "_candidate_paths": [path for path, _ in candidate_fields],  # for frequency sheet
            })

        return rows, {
            "matching_total": matching_total,
            "matching_failing": matching_failing,
            "matching_with_loc": matching_with_loc,
            "sampled": len(rows),
            "source_key_breakdown": source_key_counter,
        }


# ── Sheet writers ──────────────────────────────────────────────────


def write_summary_sheet(
    wb: Workbook,
    args: argparse.Namespace,
    counts: dict,
    field_freq: Counter,
) -> None:
    """Top-level sheet — what did we run, what did we find."""
    ws = wb.active
    ws.title = "Summary"

    now = datetime.utcnow()
    top_fields = field_freq.most_common(10)

    rows: list[tuple[str, Any]] = [
        ("Report generated", now.strftime("%Y-%m-%d %H:%M UTC")),
        ("Adapter filter", args.adapter or "(all)"),
        ("Source ID filter", args.source_id if args.source_id is not None else "(all)"),
        ("Only failing (location_raw=NULL)", "Yes" if args.only_failing else "No"),
        ("Sample size requested", args.sample),
        ("", ""),
        ("Total raw_jobs matching filter", counts["matching_total"]),
        ("  With location_raw populated", counts["matching_with_loc"]),
        ("  With location_raw NULL (failing)", counts["matching_failing"]),
        ("  Failure rate", _rate(counts["matching_failing"], counts["matching_total"])),
        ("", ""),
        ("Rows actually sampled", counts["sampled"]),
        ("Distinct sources in sample", len(counts["source_key_breakdown"])),
        ("", ""),
        ("Top candidate JSON fields (across sample):", ""),
    ]
    if not top_fields:
        rows.append(("  (no JSON listing_payload in sample)", ""))
    else:
        for path, freq in top_fields:
            rows.append((f"  {path}", f"{freq} / {counts['sampled']}"))

    rows.extend([
        ("", ""),
        ("See 'Sample' sheet for per-row raw payloads.", ""),
        ("See 'Field Frequency' sheet for full field-name counts.", ""),
        ("See 'Source Breakdown' sheet for per-source sample counts.", ""),
    ])

    for row_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    _autofit(ws, {1: 48, 2: 36})


def _rate(numerator: int, denominator: int) -> str:
    if not denominator:
        return "n/a"
    pct = numerator / denominator * 100.0
    return f"{pct:.1f}%"


def write_sample_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Sample")
    headers = [
        "Source ID",
        "Source Key",
        "Adapter",
        "Employer",
        "Board URL",
        "Title",
        "Location Raw",
        "Discovered URL",
        "Candidate Location Fields (JSON paths)",
        "Listing Payload (JSON)",
        "Raw Text Blob",
    ]
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r["source_id"])
        ws.cell(row=row_idx, column=2, value=r["source_key"])
        ws.cell(row=row_idx, column=3, value=r["adapter_name"])
        ws.cell(row=row_idx, column=4, value=r["employer_name"])
        ws.cell(row=row_idx, column=5, value=r["board_url"])
        ws.cell(row=row_idx, column=6, value=r["title_raw"])
        ws.cell(row=row_idx, column=7, value=r["location_raw"])
        ws.cell(row=row_idx, column=8, value=r["discovered_url"])
        c = ws.cell(row=row_idx, column=9, value=r["candidate_location_fields"])
        c.alignment = WRAP
        c = ws.cell(row=row_idx, column=10, value=r["listing_payload_json"])
        c.alignment = WRAP
        c = ws.cell(row=row_idx, column=11, value=r["raw_text_blob"])
        c.alignment = WRAP

    _autofit(ws, {1: 9, 2: 24, 3: 14, 4: 28, 5: 40, 6: 40, 7: 24, 8: 40, 9: 40, 10: 80, 11: 60})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_field_frequency_sheet(wb: Workbook, field_freq: Counter, sample_size: int) -> None:
    ws = wb.create_sheet("Field Frequency")
    headers = ["JSON Path", "Occurrences", "% of Sample"]
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))

    # Most-frequent-first. Ties broken alphabetically so runs are
    # reproducible (Counter's tie-breaking is insertion order).
    for row_idx, (path, freq) in enumerate(
        sorted(field_freq.items(), key=lambda t: (-t[1], t[0])),
        start=2,
    ):
        ws.cell(row=row_idx, column=1, value=path)
        ws.cell(row=row_idx, column=2, value=freq)
        ws.cell(row=row_idx, column=3, value=_rate(freq, sample_size))

    _autofit(ws, {1: 60, 2: 14, 3: 14})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_source_breakdown_sheet(wb: Workbook, source_key_counter: Counter, sample_size: int) -> None:
    ws = wb.create_sheet("Source Breakdown")
    headers = ["Source Key", "Sampled Rows", "% of Sample"]
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))

    for row_idx, (key, cnt) in enumerate(
        sorted(source_key_counter.items(), key=lambda t: (-t[1], t[0])),
        start=2,
    ):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=cnt)
        ws.cell(row=row_idx, column=3, value=_rate(cnt, sample_size))

    _autofit(ws, {1: 40, 2: 14, 3: 14})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── CLI ────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0] if __doc__ else "")
    parser.add_argument(
        "--adapter",
        help="Filter by Source.adapter_name (e.g. 'phenom', 'generic_site').",
    )
    parser.add_argument(
        "--source-id",
        type=int,
        help="Filter by a single Source.id.",
    )
    parser.add_argument(
        "--sample",
        type=int,
        default=20,
        help="Max rows to sample (default 20). Rows picked at random via ORDER BY RANDOM().",
    )
    parser.add_argument(
        "--only-failing",
        action="store_true",
        help="Only sample rows where location_raw IS NULL (the failure pool).",
    )
    parser.add_argument(
        "--out",
        help="Output .xlsx path. Default: artifacts/adapter-audit-<adapter>-<YYYY-MM-DD>.xlsx",
    )
    parser.add_argument(
        "--seed",
        type=int,
        help="Optional: seed Python's random (does NOT affect SQL RANDOM()). "
             "Useful if you want reproducible post-processing.",
    )
    args = parser.parse_args()

    if not args.adapter and args.source_id is None:
        parser.error("Must pass at least --adapter or --source-id (refusing full-table audit).")

    if args.seed is not None:
        random.seed(args.seed)

    adapter_tag = args.adapter or f"src{args.source_id}"
    date_tag = datetime.utcnow().strftime("%Y-%m-%d")
    out_path = Path(args.out) if args.out else (
        PROJECT_ROOT / "artifacts" / f"adapter-audit-{adapter_tag}-{date_tag}.xlsx"
    )
    out_path = out_path.expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(
        f"Sampling raw_jobs (adapter={args.adapter or '*'} "
        f"source_id={args.source_id if args.source_id is not None else '*'} "
        f"only_failing={args.only_failing} sample={args.sample})…",
        flush=True,
    )
    rows, counts = _sample_rows(
        adapter=args.adapter,
        source_id=args.source_id,
        only_failing=args.only_failing,
        sample_size=args.sample,
    )

    # Tally candidate-field frequency across the sample
    field_freq: Counter = Counter()
    for r in rows:
        for path in r["_candidate_paths"]:
            field_freq[path] += 1

    wb = Workbook()
    write_summary_sheet(wb, args, counts, field_freq)  # becomes active sheet
    write_sample_sheet(wb, rows)
    write_field_frequency_sheet(wb, field_freq, counts["sampled"])
    write_source_breakdown_sheet(wb, counts["source_key_breakdown"], counts["sampled"])

    wb.save(str(out_path))
    print(f"Wrote {out_path}")
    print(
        f"  matching_total={counts['matching_total']} "
        f"failing={counts['matching_failing']} "
        f"with_loc={counts['matching_with_loc']} "
        f"sampled={counts['sampled']} "
        f"distinct_sources={len(counts['source_key_breakdown'])}"
    )
    if field_freq:
        top = field_freq.most_common(5)
        print("  top candidate JSON fields:")
        for path, freq in top:
            print(f"    {path}: {freq}/{counts['sampled']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
