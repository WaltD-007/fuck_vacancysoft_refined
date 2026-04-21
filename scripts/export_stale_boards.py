#!/usr/bin/env python3
"""Export a per-board and per-adapter staleness report to XLSX.

Answers the operational audit questions "when did each adapter last run"
and "when was each board last scraped" from the SourceRun history already
in the DB. Produces an XLSX with two sheets:

  - Boards   — one row per Source. Adapter, Company, Base URL, Active,
               Last Run At, Status, Error, Days Stale, Staleness.
               Sorted: most-stale boards first so the top of the list
               is the audit work queue.
  - Adapters — one row per adapter. Boards count, Most Recent Run
               (max across all boards using this adapter), Oldest Last
               Run (min across the same), Stale Board Count (>3d by
               default), Failed Board Count (last run status != success).

Default output path: ``artifacts/stale-boards-audit-<YYYY-MM-DD>.xlsx``.
``artifacts/`` is gitignored so the xlsx never lands in git.
Non-destructive: reads only from ``sources`` + ``source_runs``.

Usage:
    python3 scripts/export_stale_boards.py
    python3 scripts/export_stale_boards.py --output ~/Desktop/audit.xlsx
    python3 scripts/export_stale_boards.py --stale-hours 72
    python3 scripts/export_stale_boards.py --include-inactive

Staleness thresholds (colour / label on the Boards sheet):
    Fresh   — last run < --fresh-hours (default 24)
    Stale   — --fresh-hours <= last run < --stale-hours (default 72)
    Ancient — last run >= --stale-hours
    Never   — no SourceRun row exists for this board

Requires: openpyxl (already in main dependencies).
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from vacancysoft.db.engine import SessionLocal
from vacancysoft.db.models import Source, SourceRun


# ── Styling ────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="FFD0D0D0", end_color="FFD0D0D0", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

# Staleness row colours. Excel wants solid-fill RGBA with the alpha
# byte leading (Office convention). Kept in a named dict so a future
# "also use these for an admin page" pass can reuse them.
STALENESS_FILL: dict[str, PatternFill] = {
    "Fresh":   PatternFill(start_color="FFD0F0D8", end_color="FFD0F0D8", fill_type="solid"),  # pale green
    "Stale":   PatternFill(start_color="FFFFF1C6", end_color="FFFFF1C6", fill_type="solid"),  # pale amber
    "Ancient": PatternFill(start_color="FFF5C6C7", end_color="FFF5C6C7", fill_type="solid"),  # pale red
    "Never":   PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid"),  # grey
}


def _style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autofit(ws, col_widths: dict[int, int]) -> None:
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ── Data assembly ──────────────────────────────────────────────────


def _classify(last_run_at: datetime | None, now: datetime, fresh_cutoff: datetime, stale_cutoff: datetime) -> str:
    """Bucket a board into one of four staleness labels."""
    if last_run_at is None:
        return "Never"
    if last_run_at >= fresh_cutoff:
        return "Fresh"
    if last_run_at >= stale_cutoff:
        return "Stale"
    return "Ancient"


def _collect_board_rows(include_inactive: bool, fresh_hours: int, stale_hours: int) -> list[dict]:
    """Build the per-Source rows for the Boards sheet.

    One row per Source with its latest SourceRun attached. Sources
    with no runs at all still surface (staleness = Never) — that's
    the biggest "audit me" signal.

    N+1-ish but fine at current scale (~1000 sources). The alternative
    (LATERAL JOIN / DISTINCT ON) optimises a use case we don't have.
    """
    now = datetime.utcnow()
    fresh_cutoff = now - timedelta(hours=fresh_hours)
    stale_cutoff = now - timedelta(hours=stale_hours)

    with SessionLocal() as s:
        stmt = select(Source)
        if not include_inactive:
            stmt = stmt.where(Source.active.is_(True))
        sources = s.execute(stmt.order_by(Source.adapter_name, Source.employer_name)).scalars().all()

        rows: list[dict] = []
        for src in sources:
            last_run = s.execute(
                select(SourceRun)
                .where(SourceRun.source_id == src.id)
                .order_by(SourceRun.created_at.desc())
                .limit(1)
            ).scalar_one_or_none()

            if last_run is None:
                last_run_at = None
                status = ""
                error = ""
                jobs_last_run = None
            else:
                # Prefer finished_at when populated (actual end of the
                # run), fall back to started_at (scheduler time) or
                # created_at (row insert time). In practice finished_at
                # is set after a successful run and is the most useful
                # "when did this board last produce data" signal.
                last_run_at = last_run.finished_at or last_run.started_at or last_run.created_at
                status = last_run.status or ""
                diag = last_run.diagnostics_blob or {}
                error = diag.get("error") or ""
                jobs_last_run = last_run.raw_jobs_created

            days_stale: str | float = ""
            if last_run_at is not None:
                delta = now - last_run_at
                days_stale = round(delta.total_seconds() / 86400.0, 1)

            rows.append({
                "adapter": src.adapter_name or "",
                "company": src.employer_name or "",
                "source_id": src.id,
                "base_url": src.base_url or "",
                "active": "Y" if src.active else "N",
                "last_run_at": last_run_at,
                "status": status,
                "error": (error[:300] + "…") if error and len(error) > 300 else error,
                "jobs_last_run": jobs_last_run,
                "days_stale": days_stale,
                "staleness": _classify(last_run_at, now, fresh_cutoff, stale_cutoff),
            })
        return rows


def _summarise_adapters(board_rows: list[dict], stale_hours: int) -> list[dict]:
    """Aggregate per-adapter from the board rows."""
    by_adapter: dict[str, list[dict]] = {}
    for row in board_rows:
        by_adapter.setdefault(row["adapter"], []).append(row)

    out: list[dict] = []
    for adapter in sorted(by_adapter.keys()):
        group = by_adapter[adapter]
        run_times = [r["last_run_at"] for r in group if r["last_run_at"] is not None]
        stale_count = sum(1 for r in group if r["staleness"] == "Ancient")
        never_count = sum(1 for r in group if r["staleness"] == "Never")
        failed_count = sum(1 for r in group if r["status"] and r["status"].lower() not in ("success", ""))
        out.append({
            "adapter": adapter,
            "boards_count": len(group),
            "most_recent_run": max(run_times) if run_times else None,
            "oldest_last_run": min(run_times) if run_times else None,
            "stale_count": stale_count,
            "never_scraped_count": never_count,
            "failed_last_run_count": failed_count,
        })
    # Push adapters with the most urgent audit signal to the top: never-
    # scraped boards > ancient boards > failed boards > just sort by
    # oldest-last-run ascending (so silent adapters surface first).
    out.sort(key=lambda r: (
        -r["never_scraped_count"],
        -r["stale_count"],
        -r["failed_last_run_count"],
        (r["oldest_last_run"] or datetime.min),
    ))
    return out


# ── Sheet writers ──────────────────────────────────────────────────


def write_boards_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Boards")
    headers = [
        "Adapter",
        "Company",
        "Source ID",
        "Base URL",
        "Active",
        "Last Run At",
        "Days Stale",
        "Staleness",
        "Last Run Status",
        "Jobs (last run)",
        "Last Run Error",
    ]
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))

    # Sort: audit-urgent first. Never-scraped → Ancient → Stale → Fresh.
    # Inside a bucket: oldest last-run first (or alphabetical if the
    # bucket carries no timestamps, e.g. Never).
    bucket_order = {"Never": 0, "Ancient": 1, "Stale": 2, "Fresh": 3}
    sorted_rows = sorted(
        rows,
        key=lambda r: (
            bucket_order.get(r["staleness"], 9),
            r["last_run_at"] or datetime.max,
            r["adapter"],
            r["company"],
        ),
    )

    for row_idx, r in enumerate(sorted_rows, start=2):
        ws.cell(row=row_idx, column=1, value=r["adapter"])
        ws.cell(row=row_idx, column=2, value=r["company"])
        ws.cell(row=row_idx, column=3, value=r["source_id"])
        ws.cell(row=row_idx, column=4, value=r["base_url"])
        ws.cell(row=row_idx, column=5, value=r["active"])
        # Excel is happiest with native datetimes for sort/filter UX.
        if r["last_run_at"]:
            c = ws.cell(row=row_idx, column=6, value=r["last_run_at"])
            c.number_format = "yyyy-mm-dd hh:mm"
        else:
            ws.cell(row=row_idx, column=6, value="")
        ws.cell(row=row_idx, column=7, value=r["days_stale"])
        ws.cell(row=row_idx, column=8, value=r["staleness"])
        ws.cell(row=row_idx, column=9, value=r["status"])
        ws.cell(row=row_idx, column=10, value=r["jobs_last_run"])
        ws.cell(row=row_idx, column=11, value=r["error"])

        # Row-level tint so a visual scan of the sheet matches the
        # sort order — the eye picks up red rows at the top before
        # reading the label column.
        fill = STALENESS_FILL.get(r["staleness"])
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    _autofit(ws, {1: 16, 2: 34, 3: 10, 4: 50, 5: 8, 6: 18, 7: 11, 8: 11, 9: 14, 10: 12, 11: 60})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_adapters_sheet(wb: Workbook, summaries: list[dict]) -> None:
    ws = wb.create_sheet("Adapters")
    headers = [
        "Adapter",
        "Boards",
        "Most Recent Run",
        "Oldest Last Run",
        "Ancient Boards",
        "Never Scraped",
        "Failed Last Run",
    ]
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))

    for row_idx, r in enumerate(summaries, start=2):
        ws.cell(row=row_idx, column=1, value=r["adapter"])
        ws.cell(row=row_idx, column=2, value=r["boards_count"])
        if r["most_recent_run"]:
            c = ws.cell(row=row_idx, column=3, value=r["most_recent_run"])
            c.number_format = "yyyy-mm-dd hh:mm"
        if r["oldest_last_run"]:
            c = ws.cell(row=row_idx, column=4, value=r["oldest_last_run"])
            c.number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=row_idx, column=5, value=r["stale_count"])
        ws.cell(row=row_idx, column=6, value=r["never_scraped_count"])
        ws.cell(row=row_idx, column=7, value=r["failed_last_run_count"])

    _autofit(ws, {1: 18, 2: 10, 3: 20, 4: 20, 5: 14, 6: 14, 7: 16})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_summary_sheet(wb: Workbook, board_rows: list[dict], args) -> None:
    """Top-level sheet with run parameters + headline counts so the
    xlsx is self-explanatory when shared around."""
    ws = wb.active
    ws.title = "Summary"

    now = datetime.utcnow()
    counts = {"Fresh": 0, "Stale": 0, "Ancient": 0, "Never": 0}
    for r in board_rows:
        counts[r["staleness"]] = counts.get(r["staleness"], 0) + 1

    rows = [
        ("Report generated", now.strftime("%Y-%m-%d %H:%M UTC")),
        ("Active boards included", "Yes (only)" if not args.include_inactive else "Yes + inactive"),
        ("Fresh threshold (hours)", args.fresh_hours),
        ("Stale threshold (hours)", args.stale_hours),
        ("", ""),
        ("Total boards audited", len(board_rows)),
        ("  Fresh (< %dh)" % args.fresh_hours, counts["Fresh"]),
        ("  Stale (%dh–%dh)" % (args.fresh_hours, args.stale_hours), counts["Stale"]),
        ("  Ancient (> %dh)" % args.stale_hours, counts["Ancient"]),
        ("  Never scraped", counts["Never"]),
        ("", ""),
        ("See 'Boards' sheet for per-board detail (most-stale first).", ""),
        ("See 'Adapters' sheet for per-adapter rollup (most-urgent first).", ""),
    ]
    for row_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    _autofit(ws, {1: 48, 2: 26})


# ── CLI ────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0] if __doc__ else "")
    default_output = (
        PROJECT_ROOT
        / "artifacts"
        / f"stale-boards-audit-{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    )
    parser.add_argument(
        "--output",
        default=str(default_output),
        help=f"Output .xlsx path (default: {default_output})",
    )
    parser.add_argument(
        "--fresh-hours",
        type=int,
        default=24,
        help="Boards scraped within this many hours are labelled 'Fresh' (default 24).",
    )
    parser.add_argument(
        "--stale-hours",
        type=int,
        default=72,
        help="Boards not scraped in this many hours are labelled 'Ancient' (default 72).",
    )
    parser.add_argument(
        "--include-inactive",
        action="store_true",
        help="Include sources with active=False in the report (default: active only).",
    )
    args = parser.parse_args()

    out_path = Path(args.output).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Querying Source + SourceRun…", flush=True)
    board_rows = _collect_board_rows(args.include_inactive, args.fresh_hours, args.stale_hours)
    adapter_summaries = _summarise_adapters(board_rows, args.stale_hours)

    wb = Workbook()
    write_summary_sheet(wb, board_rows, args)   # becomes the active sheet
    write_adapters_sheet(wb, adapter_summaries)
    write_boards_sheet(wb, board_rows)

    wb.save(str(out_path))
    print(f"Wrote {out_path}")
    print(
        f"  {len(board_rows)} boards · {len(adapter_summaries)} adapters · "
        f"fresh={sum(1 for r in board_rows if r['staleness'] == 'Fresh')} "
        f"stale={sum(1 for r in board_rows if r['staleness'] == 'Stale')} "
        f"ancient={sum(1 for r in board_rows if r['staleness'] == 'Ancient')} "
        f"never={sum(1 for r in board_rows if r['staleness'] == 'Never')}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
