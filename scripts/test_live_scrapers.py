from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook


def _load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            os.environ.setdefault(key, value)


def _progress(message: str) -> None:
    print(message, file=sys.stderr, flush=True)


def _normalise_url(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or not text.startswith(("http://", "https://")):
        return None
    return text


KNOWN_SKIP_HOST_FRAGMENTS = {"linkedin.com", "indeed.com", "glassdoor.", "jobs.google.com", "google.com"}
API_ADAPTERS = {"greenhouse", "workable", "workday", "google_jobs", "ashby", "smartrecruiters", "lever"}
BROWSER_ADAPTERS = {"eightfold", "generic_site", "icims"}
ALL_ADAPTERS = ["greenhouse", "workable", "workday", "ashby", "smartrecruiters", "lever", "icims", "eightfold", "generic_site", "google_jobs"]


def _looks_like_board(url: str) -> bool:
    lowered = url.lower()
    return not any(fragment in lowered for fragment in KNOWN_SKIP_HOST_FRAGMENTS)


def _classify_url(url: str) -> str:
    lowered = url.lower()
    host = urlparse(url).netloc.lower()
    if "greenhouse.io" in host:
        return "greenhouse"
    if "apply.workable.com" in host:
        return "workable"
    if "myworkdayjobs.com" in host:
        return "workday"
    if "ashbyhq.com" in host:
        return "ashby"
    if "smartrecruiters.com" in host:
        return "smartrecruiters"
    if "jobs.lever.co" in host:
        return "lever"
    if ".icims.com" in host:
        return "icims"
    if "eightfold.ai" in lowered or "eightfold" in lowered:
        return "eightfold"
    return "generic_site"


def _extract_greenhouse_slug(url: str) -> str | None:
    parts = [part for part in urlparse(url).path.split("/") if part]
    if not parts:
        return None
    if parts[0] == "boards" and len(parts) >= 2:
        return parts[1]
    return parts[0]


def _extract_workable_slug(url: str) -> str | None:
    parsed = urlparse(url)
    parts = [part for part in parsed.path.split("/") if part]
    return parts[0] if parsed.netloc.lower() == "apply.workable.com" and parts else None


def _extract_tail_slug(url: str) -> str | None:
    parts = [part for part in urlparse(url).path.split("/") if part]
    return parts[-1] if parts else None


def load_board_urls(xlsx_path: Path) -> dict[str, str]:
    workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    found: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if len(row) < 2:
            continue
        url = _normalise_url(row[1])
        if not url or not _looks_like_board(url):
            continue
        found.setdefault(_classify_url(url), url)
    return found


def add_repo_to_path(repo_root: Path) -> None:
    src_path = repo_root / "src"
    if str(src_path) not in sys.path:
        sys.path.insert(0, str(src_path))


def serialise_record(record: Any) -> dict[str, Any]:
    return {
        "title": getattr(record, "title_raw", None),
        "company": getattr(record, "provenance", {}).get("company") if getattr(record, "provenance", None) else None,
        "location": getattr(record, "location_raw", None),
        "posted_at": getattr(record, "posted_at_raw", None),
        "url": getattr(record, "discovered_url", None),
        "apply_url": getattr(record, "apply_url", None),
        "adapter": getattr(record, "provenance", {}).get("adapter") if getattr(record, "provenance", None) else None,
    }


def _should_run(name: str, args: argparse.Namespace) -> bool:
    if args.adapter:
        return name == args.adapter
    if args.only:
        return name in set(args.only)
    if args.no_browser and name in BROWSER_ADAPTERS:
        return False
    if args.only_browser:
        return name in BROWSER_ADAPTERS
    if args.only_api:
        return name in API_ADAPTERS
    return True


async def run_smoke_tests(args: argparse.Namespace) -> dict[str, Any]:
    from vacancysoft.adapters import (
        AshbyAdapter,
        EightfoldAdapter,
        GenericBrowserAdapter,
        GoogleJobsAdapter,
        GreenhouseAdapter,
        IcimsAdapter,
        LeverAdapter,
        SmartRecruitersAdapter,
        WorkableAdapter,
        WorkdayAdapter,
    )
    from vacancysoft.source_registry.legacy_board_mappings import lookup_company

    if args.board_url:
        adapter_name = args.adapter or _classify_url(args.board_url)
        board_urls = {adapter_name: args.board_url}
        _progress(f"Using direct board override: {board_urls}")
    else:
        _progress(f"Loading workbook: {args.xlsx}")
        board_urls = load_board_urls(Path(args.xlsx))
    _progress(f"Found source URLs: {board_urls}")
    results: dict[str, Any] = {"source_urls": board_urls, "runs": {}}

    async def capture(name: str, coro: Any) -> None:
        _progress(f"Starting {name}...")
        try:
            page = await coro
            _progress(f"Finished {name}: ok, jobs={len(page.jobs)}")
            results["runs"][name] = {
                "ok": True,
                "job_count": len(page.jobs),
                "sample": [serialise_record(job) for job in page.jobs[: args.limit]],
                "diagnostics": {
                    "counters": dict(getattr(page.diagnostics, "counters", {})),
                    "warnings": list(getattr(page.diagnostics, "warnings", [])),
                    "errors": list(getattr(page.diagnostics, "errors", [])),
                    "metadata": dict(getattr(page.diagnostics, "metadata", {})),
                },
            }
        except Exception as exc:
            _progress(f"Finished {name}: failed with {type(exc).__name__}: {exc}")
            results["runs"][name] = {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

    def canonical_company(adapter_name: str, board_url: str | None, slug: str | None = None) -> str | None:
        return lookup_company(adapter_name, board_url=board_url, slug=slug)

    greenhouse_url = board_urls.get("greenhouse")
    if _should_run("greenhouse", args):
        if greenhouse_url:
            slug = _extract_greenhouse_slug(greenhouse_url)
            await capture("greenhouse", GreenhouseAdapter().discover({"slug": slug, "company": canonical_company("greenhouse", greenhouse_url, slug), "job_board_url": greenhouse_url, "timeout_seconds": args.timeout_seconds}))
        else:
            results["runs"]["greenhouse"] = {"ok": False, "error": "No Greenhouse URL found"}

    workable_url = board_urls.get("workable")
    if _should_run("workable", args):
        if workable_url:
            slug = _extract_workable_slug(workable_url)
            await capture("workable", WorkableAdapter().discover({"slug": slug, "company": canonical_company("workable", workable_url, slug), "job_board_url": workable_url, "timeout_seconds": args.timeout_seconds}))
        else:
            results["runs"]["workable"] = {"ok": False, "error": "No Workable URL found"}

    workday_url = board_urls.get("workday")
    if _should_run("workday", args):
        if workday_url:
            adapter = WorkdayAdapter()
            _progress("Starting workday...")
            try:
                _endpoint, page = await adapter.discover_from_board_url(job_board_url=workday_url, limit=max(args.limit, 2))
                _progress(f"Finished workday: ok, jobs={len(page.jobs)}")
                results["runs"]["workday"] = {"ok": True, "job_count": len(page.jobs), "sample": [serialise_record(job) for job in page.jobs[: args.limit]], "diagnostics": {"counters": dict(getattr(page.diagnostics, "counters", {})), "warnings": list(getattr(page.diagnostics, "warnings", [])), "errors": list(getattr(page.diagnostics, "errors", [])), "metadata": dict(getattr(page.diagnostics, "metadata", {}))}}
            except Exception as exc:
                _progress(f"Finished workday: failed with {type(exc).__name__}: {exc}")
                results["runs"]["workday"] = {"ok": False, "error": f"{type(exc).__name__}: {exc}"}
        else:
            results["runs"]["workday"] = {"ok": False, "error": "No Workday URL found"}

    ashby_url = board_urls.get("ashby")
    if _should_run("ashby", args):
        if ashby_url:
            slug = _extract_tail_slug(ashby_url)
            await capture("ashby", AshbyAdapter().discover({"slug": slug, "company": canonical_company("ashby", ashby_url, slug), "job_board_url": ashby_url, "timeout_seconds": args.timeout_seconds}))
        else:
            results["runs"]["ashby"] = {"ok": False, "error": "No Ashby URL found"}

    smart_url = board_urls.get("smartrecruiters")
    if _should_run("smartrecruiters", args):
        if smart_url:
            slug = _extract_tail_slug(smart_url)
            await capture("smartrecruiters", SmartRecruitersAdapter().discover({"slug": slug, "company": canonical_company("smartrecruiters", smart_url, slug), "job_board_url": smart_url, "timeout_seconds": args.timeout_seconds, "search_terms": args.search_terms}))
        else:
            results["runs"]["smartrecruiters"] = {"ok": False, "error": "No SmartRecruiters URL found"}

    lever_url = board_urls.get("lever")
    if _should_run("lever", args):
        if lever_url:
            slug = _extract_tail_slug(lever_url)
            await capture("lever", LeverAdapter().discover({"slug": slug, "company": canonical_company("lever", lever_url, slug), "job_board_url": lever_url, "timeout_seconds": args.timeout_seconds}))
        else:
            results["runs"]["lever"] = {"ok": False, "error": "No Lever URL found"}

    icims_url = board_urls.get("icims")
    if _should_run("icims", args):
        if icims_url:
            await capture("icims", IcimsAdapter().discover({"job_board_url": icims_url, "company": canonical_company("icims", icims_url), "page_timeout_ms": args.page_timeout_ms, "search_terms": args.search_terms}))
        else:
            results["runs"]["icims"] = {"ok": False, "error": "No iCIMS URL found"}

    eightfold_url = board_urls.get("eightfold")
    if _should_run("eightfold", args):
        if eightfold_url:
            await capture("eightfold", EightfoldAdapter().discover({"job_board_url": eightfold_url, "company": canonical_company("eightfold", eightfold_url), "search_terms": args.search_terms, "page_timeout_ms": args.page_timeout_ms, "search_settle_ms": args.search_settle_ms}))
        else:
            results["runs"]["eightfold"] = {"ok": False, "error": "No Eightfold URL found"}

    generic_url = board_urls.get("generic_site")
    if _should_run("generic_site", args):
        if generic_url:
            await capture("generic_site", GenericBrowserAdapter().discover({"job_board_url": generic_url, "company": canonical_company("generic_site", generic_url), "search_terms": args.search_terms, "page_timeout_ms": args.page_timeout_ms, "wait_after_nav_ms": args.search_settle_ms}))
        else:
            results["runs"]["generic_site"] = {"ok": False, "error": "No generic-site URL found"}

    if _should_run("google_jobs", args):
        await capture("google_jobs", GoogleJobsAdapter().discover({"search_terms": args.google_queries, "timeout_seconds": args.timeout_seconds, "max_pages_per_query": 1}))

    return results


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Smoke test live scraper adapters against live job boards")
    parser.add_argument("--repo-root", default=".")
    parser.add_argument("--xlsx", default="Jobs_2026_03_31_clean.xlsx")
    parser.add_argument("--env-file", default=".env")
    parser.add_argument("--limit", type=int, default=2)
    parser.add_argument("--timeout-seconds", type=float, default=20.0)
    parser.add_argument("--page-timeout-ms", type=int, default=30000)
    parser.add_argument("--search-settle-ms", type=int, default=2000)
    parser.add_argument("--no-browser", action="store_true")
    parser.add_argument("--only-browser", action="store_true")
    parser.add_argument("--only-api", action="store_true")
    parser.add_argument("--only", action="append", choices=ALL_ADAPTERS)
    parser.add_argument("--adapter", choices=[name for name in ALL_ADAPTERS if name != "google_jobs"])
    parser.add_argument("--board-url")
    parser.add_argument("--search-term", dest="search_terms", action="append", default=None)
    parser.add_argument("--google-query", dest="google_queries", action="append", default=None)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.board_url and not args.adapter:
        args.adapter = _classify_url(args.board_url)
    if args.adapter == "google_jobs" and args.board_url:
        raise SystemExit("--board-url is not supported for google_jobs")
    repo_root = Path(args.repo_root).resolve()
    add_repo_to_path(repo_root)
    env_path = Path(args.env_file)
    if not env_path.is_absolute():
        env_path = repo_root / env_path
    _load_env_file(env_path)
    _progress(f"Loaded env file: {env_path}")
    if args.search_terms is None:
        args.search_terms = ["risk", "quant"]
    if args.google_queries is None:
        args.google_queries = ["risk manager finance", "quantitative analyst finance"]
    if not args.board_url:
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.is_absolute():
            xlsx_path = repo_root / xlsx_path
        args.xlsx = str(xlsx_path)
        if not xlsx_path.exists():
            raise SystemExit(f"Workbook not found: {xlsx_path}")
    print(json.dumps(asyncio.run(run_smoke_tests(args)), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
