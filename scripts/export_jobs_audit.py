#!/usr/bin/env python3
"""Export every RawJob in the DB with audit tags to an XLSX.

Answers "which jobs got enriched, which adapter found them, and
which ones were also picked up by an aggregator" — the three tags
an operator needs to audit pipeline coverage. No schema change,
read-only.

Output: ``artifacts/jobs-audit-<YYYY-MM-DD>.xlsx`` by default.

Sheets:

  - Summary   — totals + per-adapter breakdown (jobs, % enriched,
                % with location, % also on aggregator)
  - Jobs      — one row per RawJob. Excel autofilter + colour
                tints per pipeline stage so an operator can slice
                by adapter / pipeline stage / aggregator overlap
                right in Excel without the script running a dozen
                filtered re-exports.

Aggregator overlap (Option B, operator-chosen 2026-04-22):
Match is fuzzy on ``(lower(employer_norm), lower(title_raw))``
— employer is resolved to the canonical name via
``EnrichedJob.team`` (for aggregator rows that get enriched)
falling back to ``Source.employer_name`` (direct sources). Two
raw_jobs sharing the same `(employer, title)` pair are
considered "the same job". A direct-source job is flagged as
"Also on aggregator: X, Y" when any matching row belongs to an
aggregator adapter.

Deliberately includes all RawJob rows — even ones that never
enriched or got filtered out by the geo / agency / title
filters. The pipeline-stage column tells you why each row is
where it is. Use Excel's autofilter to slice if you only want
the enriched subset.

Usage:
    python3 scripts/export_jobs_audit.py
    python3 scripts/export_jobs_audit.py --output ~/Desktop/jobs.xlsx
    python3 scripts/export_jobs_audit.py --include-inactive

Requires: openpyxl (already in main dependencies).
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from vacancysoft.db.engine import SessionLocal
from vacancysoft.db.models import EnrichedJob, RawJob, Source


# Must match api/ledger.py::_AGGREGATOR_ADAPTERS exactly. Duplicated
# here so the script can run standalone without importing API modules
# (which would pull FastAPI in at import time). Keep in sync.
AGGREGATOR_ADAPTERS: frozenset[str] = frozenset(
    {"adzuna", "reed", "efinancialcareers", "google_jobs", "coresignal"}
)


# ── Styling ────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="FFD0D0D0", end_color="FFD0D0D0", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

# Per-pipeline-stage row tint so a visual scan of the Jobs sheet
# surfaces clusters (e.g. a big block of "geo_filtered" rows under
# one adapter → the adapter is pulling cross-border jobs).
STAGE_FILL: dict[str, PatternFill] = {
    "enriched":        PatternFill(start_color="FFD5F0D8", end_color="FFD5F0D8", fill_type="solid"),  # green
    "detail_fetched":  PatternFill(start_color="FFD5F0D8", end_color="FFD5F0D8", fill_type="solid"),  # green (alias)
    "pending":         PatternFill(start_color="FFFFF7D6", end_color="FFFFF7D6", fill_type="solid"),  # pale yellow
    "geo_filtered":    PatternFill(start_color="FFF5C6C7", end_color="FFF5C6C7", fill_type="solid"),  # red
    "agency_filtered": PatternFill(start_color="FFF5C6C7", end_color="FFF5C6C7", fill_type="solid"),  # red
    "title_filtered":  PatternFill(start_color="FFF5C6C7", end_color="FFF5C6C7", fill_type="solid"),  # red
    "detail_failed":   PatternFill(start_color="FFFFE0B3", end_color="FFFFE0B3", fill_type="solid"),  # amber
    "raw_only":        PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid"),  # grey
    "dead_at_source":  PatternFill(start_color="FFC0C0C0", end_color="FFC0C0C0", fill_type="solid"),  # darker grey
}


def _style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autofit(ws, col_widths: dict[int, int]) -> None:
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ── Data assembly ──────────────────────────────────────────────────


def _normalise(s: str | None) -> str:
    return (s or "").strip().lower()


def _pipeline_stage(raw: RawJob, enriched: EnrichedJob | None) -> str:
    """Derive a single human-readable pipeline-stage label.

    Priority: dead-at-source > enriched stage > raw-only. The
    detail_fetch_status column carries the authoritative per-stage
    outcome when the job was touched by enrichment
    (``detail_fetched`` / ``geo_filtered`` / ``agency_filtered`` /
    ``title_filtered`` / ``detail_failed``). When no EnrichedJob
    exists, the row never left the raw discovery stage.
    """
    if getattr(raw, "is_deleted_at_source", False):
        return "dead_at_source"
    if enriched is None:
        return "raw_only"
    return enriched.detail_fetch_status or "pending"


def _canonical_employer(
    source_employer: str,
    enriched_team: str | None,
    is_aggregator: bool,
) -> str:
    """Resolve the 'real employer' for cross-source matching.

    Direct sources — ``Source.employer_name`` IS the real employer
    (Barclays / HSBC / etc.).
    Aggregator sources — ``Source.employer_name`` is the aggregator
    name (Reed / Adzuna). The real employer comes out of the
    listing payload during enrichment and ends up on
    ``EnrichedJob.team``. We prefer that when populated.

    Returns an empty string when we genuinely can't tell — for
    example an aggregator row that never got enriched so
    ``team`` is None and ``employer_name`` is just "Reed".
    """
    team = _normalise(enriched_team)
    if team:
        return team
    if not is_aggregator:
        return _normalise(source_employer)
    # Aggregator without enrichment — can't identify the employer.
    return ""


def _collect_rows(include_inactive: bool) -> tuple[list[dict], dict[str, Any]]:
    """Build the per-row dict list + summary stats.

    Single SELECT joining RawJob, Source, optional EnrichedJob.
    Everything else (canonical employer, aggregator-overlap index,
    per-row overlap lookup) happens in Python — cheap at current
    scale (< 20k jobs).
    """
    with SessionLocal() as s:
        stmt = (
            select(RawJob, Source, EnrichedJob)
            .join(Source, RawJob.source_id == Source.id)
            .outerjoin(EnrichedJob, EnrichedJob.raw_job_id == RawJob.id)
            .order_by(Source.adapter_name, Source.employer_name, RawJob.first_seen_at.desc())
        )
        if not include_inactive:
            stmt = stmt.where(Source.active.is_(True))
        raw_rows = s.execute(stmt).all()

    # Phase 1: build the overlap index. Key = (employer_norm,
    # title_norm). Value = list of (adapter, source_id,
    # is_aggregator) tuples. Then a direct-source row looks up
    # its own key and collects any aggregator adapters in the
    # bucket that aren't its own source.
    overlap_index: dict[tuple[str, str], list[tuple[str, int, bool]]] = defaultdict(list)
    for raw, source, enriched in raw_rows:
        is_agg = (source.adapter_name or "") in AGGREGATOR_ADAPTERS
        canon_empl = _canonical_employer(
            source.employer_name or "",
            enriched.team if enriched else None,
            is_agg,
        )
        title_norm = _normalise(raw.title_raw)
        if not canon_empl or not title_norm:
            # Can't match this row — skip it for index purposes,
            # but it still goes into the output with "(unmatchable)"
            # in the aggregator column below.
            continue
        overlap_index[(canon_empl, title_norm)].append(
            (source.adapter_name or "", source.id, is_agg)
        )

    # Phase 2: build the output rows.
    out_rows: list[dict] = []
    stage_counts: dict[str, int] = defaultdict(int)
    per_adapter: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    total_enriched = 0
    total_with_location = 0
    total_on_aggregator = 0

    for raw, source, enriched in raw_rows:
        is_agg = (source.adapter_name or "") in AGGREGATOR_ADAPTERS
        canon_empl = _canonical_employer(
            source.employer_name or "",
            enriched.team if enriched else None,
            is_agg,
        )
        title_norm = _normalise(raw.title_raw)
        stage = _pipeline_stage(raw, enriched)

        # Aggregator-overlap lookup. Returns distinct aggregator
        # adapters matching this row, excluding the row's own
        # source — so a Reed row cross-matching itself doesn't
        # self-flag.
        overlap_aggregators: set[str] = set()
        if canon_empl and title_norm:
            for (adapter, src_id, was_agg) in overlap_index.get((canon_empl, title_norm), []):
                if was_agg and src_id != source.id:
                    overlap_aggregators.add(adapter)

        display_company = (
            (enriched.team if enriched and enriched.team else None)
            or source.employer_name
            or ""
        )
        enriched_country = (enriched.location_country if enriched else None) or ""
        enriched_city = (enriched.location_city if enriched else None) or ""
        loc_enriched = bool(enriched_country)

        out_rows.append({
            "adapter": source.adapter_name or "",
            "source_type": "aggregator" if is_agg else "direct",
            "source_id": source.id,
            "company": display_company,
            "title": raw.title_raw or "",
            "raw_location": raw.location_raw or "",
            "enriched_city": enriched_city,
            "enriched_country": enriched_country,
            "location_enriched": "Y" if loc_enriched else ("N" if enriched else "—"),
            "pipeline_stage": stage,
            "dead_at_source": "Y" if raw.is_deleted_at_source else "N",
            "also_on_aggregator": ", ".join(sorted(overlap_aggregators)),
            "first_seen_at": raw.first_seen_at,
            "discovered_url": raw.discovered_url or "",
        })

        # Stats
        stage_counts[stage] += 1
        per_adapter[source.adapter_name or ""]["total"] += 1
        if enriched:
            total_enriched += 1
            per_adapter[source.adapter_name or ""]["enriched"] += 1
            if enriched_country:
                total_with_location += 1
                per_adapter[source.adapter_name or ""]["with_location"] += 1
        if overlap_aggregators:
            total_on_aggregator += 1
            per_adapter[source.adapter_name or ""]["on_aggregator"] += 1

    summary = {
        "total": len(out_rows),
        "enriched": total_enriched,
        "with_location": total_with_location,
        "on_aggregator": total_on_aggregator,
        "stage_counts": dict(stage_counts),
        "per_adapter": {
            adapter: dict(counts) for adapter, counts in per_adapter.items()
        },
    }
    return out_rows, summary


# ── Sheet writers ──────────────────────────────────────────────────


def write_jobs_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Jobs")
    headers = [
        "Adapter",
        "Source Type",
        "Source ID",
        "Company",
        "Title",
        "Raw Location",
        "Enriched City",
        "Enriched Country",
        "Loc Enriched?",
        "Pipeline Stage",
        "Dead At Source",
        "Also On Aggregator",
        "First Seen At",
        "Discovered URL",
    ]
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r["adapter"])
        ws.cell(row=row_idx, column=2, value=r["source_type"])
        ws.cell(row=row_idx, column=3, value=r["source_id"])
        ws.cell(row=row_idx, column=4, value=r["company"])
        ws.cell(row=row_idx, column=5, value=r["title"])
        ws.cell(row=row_idx, column=6, value=r["raw_location"])
        ws.cell(row=row_idx, column=7, value=r["enriched_city"])
        ws.cell(row=row_idx, column=8, value=r["enriched_country"])
        ws.cell(row=row_idx, column=9, value=r["location_enriched"])
        ws.cell(row=row_idx, column=10, value=r["pipeline_stage"])
        ws.cell(row=row_idx, column=11, value=r["dead_at_source"])
        ws.cell(row=row_idx, column=12, value=r["also_on_aggregator"])
        if r["first_seen_at"]:
            c = ws.cell(row=row_idx, column=13, value=r["first_seen_at"])
            c.number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=row_idx, column=14, value=r["discovered_url"])

        # Pipeline-stage row tint.
        fill = STAGE_FILL.get(r["pipeline_stage"])
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    _autofit(ws, {1: 16, 2: 12, 3: 10, 4: 34, 5: 40, 6: 28, 7: 16, 8: 16,
                  9: 14, 10: 16, 11: 14, 12: 26, 13: 18, 14: 70})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_summary_sheet(wb: Workbook, summary: dict, args) -> None:
    ws = wb.active
    ws.title = "Summary"

    total = summary["total"] or 1  # avoid div-by-zero in percentages

    def pct(n: int) -> str:
        return f"{(100 * n / total):.1f}%"

    now = datetime.utcnow()
    rows: list[tuple[Any, Any]] = [
        ("Jobs audit report", ""),
        ("Report generated", now.strftime("%Y-%m-%d %H:%M UTC")),
        ("Inactive sources included?", "Yes" if args.include_inactive else "No (active only)"),
        ("", ""),
        ("Total raw jobs", summary["total"]),
        ("  Enriched (has EnrichedJob row)", f"{summary['enriched']} ({pct(summary['enriched'])})"),
        ("  With location (post-enrichment)", f"{summary['with_location']} ({pct(summary['with_location'])})"),
        ("  Also picked up by ≥1 aggregator", f"{summary['on_aggregator']} ({pct(summary['on_aggregator'])})"),
        ("", ""),
        ("Pipeline-stage breakdown", ""),
    ]
    for stage, count in sorted(summary["stage_counts"].items(), key=lambda kv: -kv[1]):
        rows.append((f"  {stage}", f"{count} ({pct(count)})"))
    rows.append(("", ""))
    rows.append(("Per-adapter breakdown", ""))
    rows.append(("  adapter · total · enriched · with-loc · on-aggregator", ""))
    for adapter, counts in sorted(summary["per_adapter"].items()):
        t = counts.get("total", 0)
        e = counts.get("enriched", 0)
        loc = counts.get("with_location", 0)
        agg = counts.get("on_aggregator", 0)
        if t == 0:
            continue
        rows.append((
            f"  {adapter}",
            f"{t} · {e} ({100*e/t:.0f}%) · {loc} ({100*loc/t:.0f}%) · {agg} ({100*agg/t:.0f}%)",
        ))
    rows.append(("", ""))
    rows.append(("See 'Jobs' sheet for per-row detail.", ""))
    rows.append(("Aggregator-overlap method", "Title + employer fuzzy match (Option B)"))

    for row_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    _autofit(ws, {1: 60, 2: 30})


# ── CLI ────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0] if __doc__ else "")
    default_output = (
        PROJECT_ROOT
        / "artifacts"
        / f"jobs-audit-{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    )
    parser.add_argument(
        "--output",
        default=str(default_output),
        help=f"Output .xlsx path (default: {default_output})",
    )
    parser.add_argument(
        "--include-inactive",
        action="store_true",
        help="Include sources with active=False (default: active only).",
    )
    args = parser.parse_args()

    out_path = Path(args.output).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print("Querying RawJob + Source + EnrichedJob…", flush=True)
    rows, summary = _collect_rows(args.include_inactive)

    print(
        f"  {summary['total']} rows · {summary['enriched']} enriched · "
        f"{summary['with_location']} with location · "
        f"{summary['on_aggregator']} also on aggregator",
        flush=True,
    )

    wb = Workbook()
    write_summary_sheet(wb, summary, args)
    write_jobs_sheet(wb, rows)

    wb.save(str(out_path))
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
