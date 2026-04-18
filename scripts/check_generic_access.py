from __future__ import annotations

import asyncio
import importlib.util
import sys
from pathlib import Path

from rich.console import Console
from rich.table import Table
from playwright.async_api import async_playwright

console = Console()


def add_repo_to_path(repo_root: Path) -> None:
    src_path = repo_root / "src"
    if str(src_path) not in sys.path:
        sys.path.insert(0, str(src_path))


def load_generic_browser_boards(repo_root: Path) -> list[dict]:
    config_path = repo_root / "configs" / "config.py"
    if not config_path.exists():
        raise FileNotFoundError(f"Could not find config file: {config_path}")

    spec = importlib.util.spec_from_file_location("project_config", config_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load config module from: {config_path}")

    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)

    boards = getattr(module, "GENERIC_BROWSER_BOARDS", None)
    if boards is None:
        raise AttributeError(f"GENERIC_BROWSER_BOARDS not found in: {config_path}")
    return boards


def clean(text: str | None, limit: int = 120) -> str:
    if not text:
        return ""
    text = " ".join(text.split())
    return text[: limit - 3] + "..." if len(text) > limit else text


def classify_status(status_code: int | None, title: str, body_text: str) -> str:
    lowered = f"{title}\n{body_text}".lower()
    if status_code == 403 or "403" in lowered or "forbidden" in lowered or "access is denied" in lowered:
        return "403"
    if status_code == 401 or "unauthor" in lowered:
        return "401"
    if "access denied" in lowered or "blocked" in lowered or "captcha" in lowered:
        return "blocked"
    return "ok"


async def inspect_one(playwright, url: str, company: str):
    from vacancysoft.browser.session import browser_session

    async with browser_session(playwright, headless=True) as (_browser, context):
        page = await context.new_page()
        main_status = None

        def handle_response(response):
            nonlocal main_status
            if response.url.rstrip("/") == url.rstrip("/") and main_status is None:
                main_status = response.status

        page.on("response", handle_response)

        try:
            response = await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await page.wait_for_timeout(3000)

            title = await page.title()
            final_url = page.url
            body_text = await page.evaluate("() => (document.body?.innerText || '').slice(0, 2000)")

            status_code = response.status if response is not None else main_status
            verdict = classify_status(status_code, title, body_text)

            return {
                "company": company,
                "url": url,
                "final_url": final_url,
                "status": status_code,
                "title": title,
                "verdict": verdict,
                "body_sample": clean(body_text, 200),
            }
        except Exception as exc:
            return {
                "company": company,
                "url": url,
                "final_url": "",
                "status": None,
                "title": "",
                "verdict": "error",
                "body_sample": f"{type(exc).__name__}: {exc}",
            }
        finally:
            await page.close()


async def main() -> None:
    repo_root = Path(".").resolve()
    add_repo_to_path(repo_root)

    all_boards = load_generic_browser_boards(repo_root)
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else 20
    boards = all_boards[:limit]
    results = []

    async with async_playwright() as playwright:
        for board in boards:
            company = board["company"]
            url = board["url"]
            console.print(f"[dim]Checking {company}[/dim]")
            result = await inspect_one(playwright, url, company)
            results.append(result)

    table = Table(title=f"Generic board access check ({len(results)} sites)")
    table.add_column("Company")
    table.add_column("Verdict")
    table.add_column("HTTP")
    table.add_column("Title")
    table.add_column("Final URL")
    table.add_column("Body sample")

    for row in results:
        verdict_colour = {
            "ok": "green",
            "403": "red",
            "401": "red",
            "blocked": "yellow",
            "error": "magenta",
        }.get(row["verdict"], "white")

        table.add_row(
            row["company"],
            f"[{verdict_colour}]{row['verdict']}[/{verdict_colour}]",
            str(row["status"] or ""),
            clean(row["title"], 60),
            clean(row["final_url"], 80),
            row["body_sample"],
        )

    console.print()
    console.print(table)

    counts = {}
    for row in results:
        counts[row["verdict"]] = counts.get(row["verdict"], 0) + 1

    console.print(f"\n[bold]Verdict counts:[/bold] {counts}")

    # Export to Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "Access Check"

        headers = ["Company", "Verdict", "HTTP Status", "Title", "URL", "Final URL", "Body Sample"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        verdict_fills = {
            "ok": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "403": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "401": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "blocked": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "error": PatternFill(start_color="E4AAFF", end_color="E4AAFF", fill_type="solid"),
        }

        for row_idx, row in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=row["company"])
            verdict_cell = ws.cell(row=row_idx, column=2, value=row["verdict"])
            verdict_cell.fill = verdict_fills.get(row["verdict"], PatternFill())
            ws.cell(row=row_idx, column=3, value=row["status"])
            ws.cell(row=row_idx, column=4, value=row["title"])
            ws.cell(row=row_idx, column=5, value=row["url"])
            ws.cell(row=row_idx, column=6, value=row["final_url"])
            ws.cell(row=row_idx, column=7, value=row["body_sample"])

        # Auto-fit column widths (approximate)
        col_widths = {"A": 30, "B": 12, "C": 12, "D": 50, "E": 60, "F": 60, "G": 80}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Add verdict summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.cell(row=1, column=1, value="Verdict").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="Count").font = Font(bold=True)
        for i, (verdict, count) in enumerate(sorted(counts.items()), 2):
            ws2.cell(row=i, column=1, value=verdict)
            ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=len(counts) + 3, column=1, value="Total").font = Font(bold=True)
        ws2.cell(row=len(counts) + 3, column=2, value=len(results)).font = Font(bold=True)

        # Add autofilter to main sheet
        ws.auto_filter.ref = f"A1:G{len(results) + 1}"

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = repo_root / f"generic_access_check_{timestamp}.xlsx"
        wb.save(out_path)
        console.print(f"\n[bold green]Report saved to:[/bold green] {out_path}")

    except ImportError:
        console.print("\n[yellow]openpyxl not installed — skipping Excel export. pip install openpyxl[/yellow]")


if __name__ == "__main__":
    asyncio.run(main())