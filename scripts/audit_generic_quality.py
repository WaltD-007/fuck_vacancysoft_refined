from __future__ import annotations

import asyncio
import importlib.util
import json
import sys
from datetime import datetime
from pathlib import Path

from rich.console import Console
from rich.table import Table

console = Console()


def add_repo_to_path(repo_root: Path) -> None:
    src_path = repo_root / "src"
    if str(src_path) not in sys.path:
        sys.path.insert(0, str(src_path))


def load_generic_browser_boards(repo_root: Path) -> list[dict]:
    config_path = repo_root / "configs" / "config.py"
    if not config_path.exists():
        raise FileNotFoundError(f"Could not find config file: {config_path}")

    module_name = "project_config_for_audit"
    spec = importlib.util.spec_from_file_location(module_name, config_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load config module from: {config_path}")

    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)

    boards = getattr(module, "GENERIC_BROWSER_BOARDS", None)
    if boards is None:
        raise AttributeError(f"GENERIC_BROWSER_BOARDS not found in: {config_path}")
    return boards


def clean(text: str | None, limit: int = 80) -> str:
    if not text:
        return ""
    text = " ".join(text.split())
    return text[: limit - 3] + "..." if len(text) > limit else text


def sanitise(text: str | None) -> str:
    """Strip characters illegal in Excel XML."""
    if not text:
        return ""
    return "".join(
        ch for ch in text
        if ch in ("\n", "\r", "\t") or ord(ch) >= 32
    )


async def main() -> None:
    repo_root = Path(".").resolve()
    add_repo_to_path(repo_root)

    from vacancysoft.adapters import GenericBrowserAdapter

    all_boards = load_generic_browser_boards(repo_root)
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else len(all_boards)
    boards = all_boards[:limit]

    results = []

    for idx, board in enumerate(boards, 1):
        company = board["company"]
        url = board["url"]
        try:
            page = await GenericBrowserAdapter().discover(
                {
                    "job_board_url": url,
                    "company": company,
                    "page_timeout_ms": 30000,
                    "wait_after_nav_ms": 2000,
                }
            )

            diagnostics = getattr(page, "diagnostics", None)
            metadata = dict(getattr(diagnostics, "metadata", {})) if diagnostics else {}
            counters = dict(getattr(diagnostics, "counters", {})) if diagnostics else {}
            warnings = list(getattr(diagnostics, "warnings", [])) if diagnostics else []
            errors = list(getattr(diagnostics, "errors", [])) if diagnostics else []
            timings = dict(getattr(diagnostics, "timings_ms", {})) if diagnostics else {}

            sample_titles = [getattr(job, "title_raw", None) for job in page.jobs[:5]]
            sample_urls = [getattr(job, "discovered_url", None) for job in page.jobs[:5]]

            verdict = "ok"
            if len(page.jobs) == 0:
                verdict = "empty"
            elif sample_titles and any(
                t and any(
                    bad in t.lower()
                    for bad in ("saved jobs", "talent community", "job search", "clear all", "view all jobs")
                )
                for t in sample_titles
            ):
                verdict = "noisy"

            row = {
                "company": company,
                "url": url,
                "verdict": verdict,
                "job_count": len(page.jobs),
                "listings_seen": counters.get("listings_seen", 0),
                "unique_urls": counters.get("unique_urls", 0),
                "pages_scraped": counters.get("pages_scraped", 0),
                "duplicate_urls": counters.get("duplicate_urls", 0),
                "last_selector_used": metadata.get("last_selector_used", ""),
                "final_url": metadata.get("final_url", ""),
                "http_status": counters.get("http_status", ""),
                "elapsed_ms": timings.get("discover", 0),
                "sample_titles": sample_titles,
                "sample_urls": sample_urls,
                "warnings": warnings,
                "errors": errors,
            }
            results.append(row)

            colour = {"ok": "green", "empty": "yellow", "noisy": "magenta"}.get(verdict, "white")
            top_title = clean(sample_titles[0], 50) if sample_titles and sample_titles[0] else "—"
            console.print(
                f"[{colour}][{idx}/{len(boards)}] {verdict:6s}[/{colour}] "
                f"| {company:40s} | jobs={len(page.jobs):<4d} "
                f"| {top_title}"
            )

        except Exception as exc:
            row = {
                "company": company,
                "url": url,
                "verdict": "error",
                "job_count": 0,
                "listings_seen": 0,
                "unique_urls": 0,
                "pages_scraped": 0,
                "duplicate_urls": 0,
                "last_selector_used": "",
                "final_url": "",
                "http_status": "",
                "elapsed_ms": 0,
                "sample_titles": [f"{type(exc).__name__}: {exc}"],
                "sample_urls": [],
                "warnings": [],
                "errors": [str(exc)],
            }
            results.append(row)

            console.print(
                f"[red][{idx}/{len(boards)}] error [/red] "
                f"| {company:40s} | {type(exc).__name__}: {str(exc)[:80]}"
            )

    # -----------------------------------------------------------------------
    # Console summary
    # -----------------------------------------------------------------------
    table = Table(title=f"Generic audit ({len(results)} sites)")
    table.add_column("Company")
    table.add_column("Verdict")
    table.add_column("Jobs", justify="right")
    table.add_column("Listings", justify="right")
    table.add_column("Unique", justify="right")
    table.add_column("Selector")
    table.add_column("Sample")

    for row in results:
        colour = {
            "ok": "green",
            "empty": "yellow",
            "noisy": "magenta",
            "error": "red",
        }.get(row["verdict"], "white")

        table.add_row(
            row["company"],
            f"[{colour}]{row['verdict']}[/{colour}]",
            str(row["job_count"]),
            str(row["listings_seen"]),
            str(row["unique_urls"]),
            clean(row["last_selector_used"], 30),
            clean(" | ".join(t for t in row["sample_titles"] if t), 100),
        )

    console.print()
    console.print(table)

    counts = {}
    for row in results:
        counts[row["verdict"]] = counts.get(row["verdict"], 0) + 1
    console.print(f"\n[bold]Verdict counts:[/bold] {counts}")

    # -----------------------------------------------------------------------
    # JSON output
    # -----------------------------------------------------------------------
    json_path = repo_root / "generic_audit_results.json"
    json_path.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
    console.print(f"[dim]Wrote JSON to {json_path}[/dim]")

    # -----------------------------------------------------------------------
    # Excel output — one row per firm
    # -----------------------------------------------------------------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Results"

        headers = [
            "Company", "URL", "Verdict", "Jobs Found", "Listings Seen",
            "Unique URLs", "Pages Scraped", "Duplicates", "HTTP Status",
            "Elapsed (ms)", "Selector Used", "Final URL",
            "Sample Title 1", "Sample Title 2", "Sample Title 3",
            "Sample URL 1", "Sample URL 2", "Sample URL 3",
            "Warnings", "Errors",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        verdict_fills = {
            "ok": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "empty": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "noisy": PatternFill(start_color="E4AAFF", end_color="E4AAFF", fill_type="solid"),
            "error": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        }

        for row_idx, r in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=r["company"])
            ws.cell(row=row_idx, column=2, value=r["url"])
            vc = ws.cell(row=row_idx, column=3, value=r["verdict"])
            vc.fill = verdict_fills.get(r["verdict"], PatternFill())
            ws.cell(row=row_idx, column=4, value=r["job_count"])
            ws.cell(row=row_idx, column=5, value=r["listings_seen"])
            ws.cell(row=row_idx, column=6, value=r["unique_urls"])
            ws.cell(row=row_idx, column=7, value=r["pages_scraped"])
            ws.cell(row=row_idx, column=8, value=r["duplicate_urls"])
            ws.cell(row=row_idx, column=9, value=r.get("http_status", ""))
            ws.cell(row=row_idx, column=10, value=r["elapsed_ms"])
            ws.cell(row=row_idx, column=11, value=r["last_selector_used"])
            ws.cell(row=row_idx, column=12, value=r["final_url"])

            titles = r.get("sample_titles", [])
            for i in range(3):
                val = sanitise(titles[i]) if i < len(titles) else ""
                ws.cell(row=row_idx, column=13 + i, value=val)

            urls = r.get("sample_urls", [])
            for i in range(3):
                val = sanitise(urls[i]) if i < len(urls) else ""
                ws.cell(row=row_idx, column=16 + i, value=val)

            ws.cell(row=row_idx, column=19, value=sanitise("; ".join(r.get("warnings", [])))[:500])
            ws.cell(row=row_idx, column=20, value=sanitise("; ".join(r.get("errors", [])))[:500])

        # Autofilter
        ws.auto_filter.ref = f"A1:T{len(results) + 1}"

        # Column widths
        col_widths = {
            "A": 30, "B": 55, "C": 10, "D": 10, "E": 12, "F": 10,
            "G": 12, "H": 10, "I": 10, "J": 12, "K": 25, "L": 55,
            "M": 40, "N": 40, "O": 40, "P": 55, "Q": 55, "R": 55,
            "S": 40, "T": 40,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.cell(row=1, column=1, value="Verdict").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="Count").font = Font(bold=True)
        ws2.cell(row=1, column=3, value="% of Total").font = Font(bold=True)
        for i, (verdict, count) in enumerate(sorted(counts.items()), 2):
            ws2.cell(row=i, column=1, value=verdict)
            ws2.cell(row=i, column=2, value=count)
            ws2.cell(row=i, column=3, value=f"{count / len(results) * 100:.1f}%")
            vc = ws2.cell(row=i, column=1)
            vc.fill = verdict_fills.get(verdict, PatternFill())
        ws2.cell(row=len(counts) + 3, column=1, value="Total").font = Font(bold=True)
        ws2.cell(row=len(counts) + 3, column=2, value=len(results)).font = Font(bold=True)

        total_jobs = sum(r["job_count"] for r in results)
        ws2.cell(row=len(counts) + 4, column=1, value="Total Jobs Found").font = Font(bold=True)
        ws2.cell(row=len(counts) + 4, column=2, value=total_jobs).font = Font(bold=True)

        for col, w in {"A": 16, "B": 10, "C": 12}.items():
            ws2.column_dimensions[col].width = w

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = repo_root / f"generic_audit_results_{timestamp}.xlsx"
        wb.save(xlsx_path)
        console.print(f"\n[bold green]Excel report saved to:[/bold green] {xlsx_path}")

    except ImportError:
        console.print("\n[yellow]openpyxl not installed — skipping Excel export. pip install openpyxl[/yellow]")


if __name__ == "__main__":
    asyncio.run(main())
