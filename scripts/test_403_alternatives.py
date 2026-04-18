"""
Test alternative scraping strategies for the 19 sites that returned 403.

Strategies tested:
  1. httpx GET with realistic browser headers (no Playwright overhead)
  2. Playwright headless with latest Chrome user-agent
  3. Playwright headless with stealth settings (webdriver flag removed, etc.)
  4. Playwright non-headless (headed) mode
  5. httpx GET following redirects to discover the real careers URL
  6. Google cache / alternate URL discovery via redirect chain

Usage:
    python3.13 scripts/test_403_alternatives.py
"""

from __future__ import annotations

import asyncio
import sys
import time
from datetime import datetime
from pathlib import Path

from rich.console import Console
from rich.table import Table

console = Console()

# ---------------------------------------------------------------------------
# The 19 sites that returned 403 (deduplicated by URL)
# ---------------------------------------------------------------------------
SITES_403: list[dict[str, str]] = [
    {"company": "Arbuthnot Latham", "url": "https://careers.arbuthnotlatham.co.uk/vacancies"},
    {"company": "Bloomberg", "url": "https://www.bloomberg.com/company/careers/"},
    {"company": "BNP Paribas", "url": "https://group.bnpparibas/en/careers/job-offers"},
    {"company": "CQS", "url": "https://www.cqs.com/careers"},
    {"company": "Ecotricity", "url": "https://jobs.ecotricity.co.uk/vacancies"},
    {"company": "GAMCO Investors", "url": "https://www.gabelli.com/careers"},
    {"company": "Manulife Investment Management", "url": "https://careers.manulife.com/"},
    {"company": "Market Harborough Building Society", "url": "https://mhbs.co.uk/vacancies/"},
    {"company": "Neuberger Berman", "url": "https://www.nb.com/en/global/careers"},
    {"company": "New York Life", "url": "https://www.newyorklife.com/careers"},
    {"company": "NFU Mutual", "url": "https://emea3.recruitmentplatform.com/syndicated/lay/jsoutputinitrapido.cfm"},
    {"company": "Rabobank", "url": "https://www.rabobank.jobs/en/"},
    {"company": "Schonfeld", "url": "https://www.schonfeld.com/careers"},
    {"company": "SoFi", "url": "https://www.sofi.com/careers/"},
    {"company": "Virtu Financial", "url": "https://www.virtu.com/careers/"},
    {"company": "Walleye Capital", "url": "https://walleye.com/careers"},
]

# ---------------------------------------------------------------------------
# User-agent strings
# ---------------------------------------------------------------------------
UA_CHROME_LATEST = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/131.0.0.0 Safari/537.36"
)
UA_FIREFOX = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:132.0) "
    "Gecko/20100101 Firefox/132.0"
)
UA_SAFARI = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) "
    "Version/18.1 Safari/605.1.15"
)

REALISTIC_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-GB,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "DNT": "1",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Cache-Control": "max-age=0",
}

TIMEOUT_MS = 30_000


# ---------------------------------------------------------------------------
# Strategy 1: Plain httpx with realistic headers
# ---------------------------------------------------------------------------
async def strategy_httpx(url: str) -> dict:
    import httpx

    for ua_label, ua in [("chrome", UA_CHROME_LATEST), ("firefox", UA_FIREFOX), ("safari", UA_SAFARI)]:
        headers = {**REALISTIC_HEADERS, "User-Agent": ua}
        try:
            async with httpx.AsyncClient(timeout=20, follow_redirects=True, max_redirects=10) as client:
                resp = await client.get(url, headers=headers)
                if resp.status_code < 400:
                    body_sample = resp.text[:300].replace("\n", " ").strip()
                    return {
                        "status": resp.status_code,
                        "verdict": "ok",
                        "detail": f"ua={ua_label}",
                        "body_sample": body_sample,
                        "final_url": str(resp.url),
                    }
        except Exception as exc:
            continue

    # All UAs failed — return last attempt info
    try:
        headers = {**REALISTIC_HEADERS, "User-Agent": UA_CHROME_LATEST}
        async with httpx.AsyncClient(timeout=20, follow_redirects=True) as client:
            resp = await client.get(url, headers=headers)
            return {
                "status": resp.status_code,
                "verdict": "fail",
                "detail": "all UAs failed",
                "body_sample": resp.text[:200].replace("\n", " "),
                "final_url": str(resp.url),
            }
    except Exception as exc:
        return {"status": None, "verdict": "error", "detail": str(exc)[:150], "body_sample": "", "final_url": ""}


# ---------------------------------------------------------------------------
# Strategy 2: Playwright headless with latest UA
# ---------------------------------------------------------------------------
async def strategy_pw_headless(pw, url: str) -> dict:
    browser = await pw.chromium.launch(headless=True)
    try:
        ctx = await browser.new_context(
            user_agent=UA_CHROME_LATEST,
            viewport={"width": 1280, "height": 900},
            extra_http_headers=REALISTIC_HEADERS,
        )
        page = await ctx.new_page()
        try:
            resp = await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT_MS)
            await page.wait_for_timeout(2000)
            title = await page.title()
            body = await page.evaluate("() => (document.body?.innerText || '').slice(0, 300)")
            status = resp.status if resp else None
            verdict = "ok" if status and status < 400 else "fail"
            return {
                "status": status,
                "verdict": verdict,
                "detail": f"title={title[:60]}",
                "body_sample": body.replace("\n", " ")[:200],
                "final_url": page.url,
            }
        except Exception as exc:
            return {"status": None, "verdict": "error", "detail": str(exc)[:150], "body_sample": "", "final_url": ""}
        finally:
            await page.close()
            await ctx.close()
    finally:
        await browser.close()


# ---------------------------------------------------------------------------
# Strategy 3: Playwright headless + stealth (remove webdriver flag)
# ---------------------------------------------------------------------------
STEALTH_JS = """
Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
Object.defineProperty(navigator, 'languages', { get: () => ['en-GB', 'en'] });
Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
window.chrome = { runtime: {} };
"""


async def strategy_pw_stealth(pw, url: str) -> dict:
    browser = await pw.chromium.launch(headless=True)
    try:
        ctx = await browser.new_context(
            user_agent=UA_CHROME_LATEST,
            viewport={"width": 1280, "height": 900},
            locale="en-GB",
            timezone_id="Europe/London",
            extra_http_headers=REALISTIC_HEADERS,
        )
        await ctx.add_init_script(STEALTH_JS)
        page = await ctx.new_page()
        try:
            resp = await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT_MS)
            await page.wait_for_timeout(3000)
            title = await page.title()
            body = await page.evaluate("() => (document.body?.innerText || '').slice(0, 300)")
            status = resp.status if resp else None
            verdict = "ok" if status and status < 400 else "fail"
            return {
                "status": status,
                "verdict": verdict,
                "detail": f"stealth, title={title[:50]}",
                "body_sample": body.replace("\n", " ")[:200],
                "final_url": page.url,
            }
        except Exception as exc:
            return {"status": None, "verdict": "error", "detail": str(exc)[:150], "body_sample": "", "final_url": ""}
        finally:
            await page.close()
            await ctx.close()
    finally:
        await browser.close()


# ---------------------------------------------------------------------------
# Strategy 4: Playwright headed (non-headless) — most like a real browser
# ---------------------------------------------------------------------------
async def strategy_pw_headed(pw, url: str) -> dict:
    browser = await pw.chromium.launch(headless=False)
    try:
        ctx = await browser.new_context(
            user_agent=UA_CHROME_LATEST,
            viewport={"width": 1280, "height": 900},
            locale="en-GB",
            timezone_id="Europe/London",
            extra_http_headers=REALISTIC_HEADERS,
        )
        await ctx.add_init_script(STEALTH_JS)
        page = await ctx.new_page()
        try:
            resp = await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT_MS)
            await page.wait_for_timeout(3000)
            title = await page.title()
            body = await page.evaluate("() => (document.body?.innerText || '').slice(0, 300)")
            status = resp.status if resp else None
            verdict = "ok" if status and status < 400 else "fail"
            return {
                "status": status,
                "verdict": verdict,
                "detail": f"headed, title={title[:50]}",
                "body_sample": body.replace("\n", " ")[:200],
                "final_url": page.url,
            }
        except Exception as exc:
            return {"status": None, "verdict": "error", "detail": str(exc)[:150], "body_sample": "", "final_url": ""}
        finally:
            await page.close()
            await ctx.close()
    finally:
        await browser.close()


# ---------------------------------------------------------------------------
# Strategy 5: Playwright Firefox (different engine entirely)
# ---------------------------------------------------------------------------
async def strategy_pw_firefox(pw, url: str) -> dict:
    try:
        browser = await pw.firefox.launch(headless=True)
    except Exception as exc:
        return {"status": None, "verdict": "skip", "detail": f"Firefox not installed: {str(exc)[:80]}", "body_sample": "", "final_url": ""}
    try:
        ctx = await browser.new_context(
            user_agent=UA_FIREFOX,
            viewport={"width": 1280, "height": 900},
            locale="en-GB",
            timezone_id="Europe/London",
        )
        page = await ctx.new_page()
        try:
            resp = await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT_MS)
            await page.wait_for_timeout(2000)
            title = await page.title()
            body = await page.evaluate("() => (document.body?.innerText || '').slice(0, 300)")
            status = resp.status if resp else None
            verdict = "ok" if status and status < 400 else "fail"
            return {
                "status": status,
                "verdict": verdict,
                "detail": f"firefox, title={title[:50]}",
                "body_sample": body.replace("\n", " ")[:200],
                "final_url": page.url,
            }
        except Exception as exc:
            return {"status": None, "verdict": "error", "detail": str(exc)[:150], "body_sample": "", "final_url": ""}
        finally:
            await page.close()
            await ctx.close()
    finally:
        await browser.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
STRATEGIES = [
    ("httpx_realistic", None),     # no pw needed
    ("pw_headless",     None),
    ("pw_stealth",      None),
    ("pw_headed",       None),
    ("pw_firefox",      None),
]


async def run_strategy(name: str, pw, url: str) -> dict:
    if name == "httpx_realistic":
        return await strategy_httpx(url)
    elif name == "pw_headless":
        return await strategy_pw_headless(pw, url)
    elif name == "pw_stealth":
        return await strategy_pw_stealth(pw, url)
    elif name == "pw_headed":
        return await strategy_pw_headed(pw, url)
    elif name == "pw_firefox":
        return await strategy_pw_firefox(pw, url)
    return {"status": None, "verdict": "error", "detail": "unknown strategy", "body_sample": "", "final_url": ""}


async def main() -> None:
    from playwright.async_api import async_playwright

    all_results: list[dict] = []

    async with async_playwright() as pw:
        for site in SITES_403:
            company = site["company"]
            url = site["url"]
            console.print(f"\n[bold cyan]{'=' * 70}[/bold cyan]")
            console.print(f"[bold]{company}[/bold] — {url}")

            for strat_name, _ in STRATEGIES:
                console.print(f"  [dim]→ {strat_name}...[/dim]", end=" ")
                t0 = time.monotonic()
                result = await run_strategy(strat_name, pw, url)
                elapsed = round((time.monotonic() - t0) * 1000)
                result["company"] = company
                result["url"] = url
                result["strategy"] = strat_name
                result["elapsed_ms"] = elapsed

                colour = "green" if result["verdict"] == "ok" else ("red" if result["verdict"] == "fail" else "magenta")
                console.print(
                    f"[{colour}]{result['verdict']}[/{colour}] "
                    f"(HTTP {result.get('status', '?')}, {elapsed}ms) "
                    f"{result.get('detail', '')[:80]}"
                )
                all_results.append(result)

                # If this strategy worked, no need to try more
                if result["verdict"] == "ok":
                    # Still run remaining strategies for comparison, but mark this as the winner
                    pass

    # -----------------------------------------------------------------------
    # Console summary table
    # -----------------------------------------------------------------------
    console.print(f"\n\n[bold]{'=' * 70}[/bold]")
    console.print("[bold]SUMMARY: Best strategy per site[/bold]\n")

    summary_table = Table(title="403 Sites — Best Working Strategy")
    summary_table.add_column("Company", style="bold")
    summary_table.add_column("Best Strategy", style="cyan")
    summary_table.add_column("HTTP")
    summary_table.add_column("Final URL")
    summary_table.add_column("Still 403?", style="red")

    site_best: list[dict] = []
    for site in SITES_403:
        site_results = [r for r in all_results if r["url"] == site["url"]]
        ok_results = [r for r in site_results if r["verdict"] == "ok"]
        if ok_results:
            best = ok_results[0]
            summary_table.add_row(
                site["company"],
                best["strategy"],
                str(best.get("status", "")),
                (best.get("final_url") or "")[:70],
                "No",
            )
            site_best.append({**site, "best_strategy": best["strategy"], "still_403": False, **best})
        else:
            summary_table.add_row(site["company"], "—", "—", "—", "[bold red]Yes[/bold red]")
            site_best.append({**site, "best_strategy": None, "still_403": True})

    console.print(summary_table)

    ok_count = sum(1 for s in site_best if not s["still_403"])
    console.print(f"\n[bold green]{ok_count}/{len(SITES_403)} sites accessible with alternative strategies[/bold green]")
    console.print(f"[bold red]{len(SITES_403) - ok_count}/{len(SITES_403)} remain inaccessible[/bold red]")

    # -----------------------------------------------------------------------
    # Export to Excel
    # -----------------------------------------------------------------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Sheet 1: All results (every strategy × every site)
        ws1 = wb.active
        ws1.title = "All Results"
        headers = ["Company", "URL", "Strategy", "Verdict", "HTTP Status", "Elapsed (ms)", "Detail", "Final URL", "Body Sample"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        verdict_fills = {
            "ok": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "fail": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "error": PatternFill(start_color="E4AAFF", end_color="E4AAFF", fill_type="solid"),
        }
        for row_idx, r in enumerate(all_results, 2):
            ws1.cell(row=row_idx, column=1, value=r.get("company", ""))
            ws1.cell(row=row_idx, column=2, value=r.get("url", ""))
            ws1.cell(row=row_idx, column=3, value=r.get("strategy", ""))
            vc = ws1.cell(row=row_idx, column=4, value=r.get("verdict", ""))
            vc.fill = verdict_fills.get(r.get("verdict", ""), PatternFill())
            ws1.cell(row=row_idx, column=5, value=r.get("status"))
            ws1.cell(row=row_idx, column=6, value=r.get("elapsed_ms"))
            ws1.cell(row=row_idx, column=7, value=r.get("detail", "")[:200])
            ws1.cell(row=row_idx, column=8, value=r.get("final_url", "")[:200])
            # Strip illegal XML characters that crash openpyxl
            body_safe = "".join(ch for ch in (r.get("body_sample", "") or "") if ch == "\n" or ch == "\r" or ch == "\t" or ord(ch) >= 32)
            ws1.cell(row=row_idx, column=9, value=body_safe[:300])

        ws1.auto_filter.ref = f"A1:I{len(all_results) + 1}"
        for col, w in {"A": 30, "B": 55, "C": 18, "D": 10, "E": 12, "F": 12, "G": 50, "H": 60, "I": 80}.items():
            ws1.column_dimensions[col].width = w

        # Sheet 2: Summary — best strategy per site
        ws2 = wb.create_sheet("Best Strategy")
        headers2 = ["Company", "URL", "Best Strategy", "Still 403?", "HTTP Status", "Final URL"]
        for col_idx, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, s in enumerate(site_best, 2):
            ws2.cell(row=row_idx, column=1, value=s["company"])
            ws2.cell(row=row_idx, column=2, value=s["url"])
            ws2.cell(row=row_idx, column=3, value=s.get("best_strategy") or "None")
            still = ws2.cell(row=row_idx, column=4, value="Yes" if s["still_403"] else "No")
            still.fill = verdict_fills["fail"] if s["still_403"] else verdict_fills["ok"]
            ws2.cell(row=row_idx, column=5, value=s.get("status"))
            ws2.cell(row=row_idx, column=6, value=s.get("final_url", "")[:200])

        for col, w in {"A": 30, "B": 55, "C": 18, "D": 12, "E": 12, "F": 60}.items():
            ws2.column_dimensions[col].width = w

        # Sheet 3: Strategy effectiveness
        ws3 = wb.create_sheet("Strategy Stats")
        ws3.cell(row=1, column=1, value="Strategy").font = Font(bold=True)
        ws3.cell(row=1, column=2, value="OK").font = Font(bold=True)
        ws3.cell(row=1, column=3, value="Fail").font = Font(bold=True)
        ws3.cell(row=1, column=4, value="Error").font = Font(bold=True)
        ws3.cell(row=1, column=5, value="Success Rate").font = Font(bold=True)

        for row_idx, (strat_name, _) in enumerate(STRATEGIES, 2):
            strat_results = [r for r in all_results if r["strategy"] == strat_name]
            ok = sum(1 for r in strat_results if r["verdict"] == "ok")
            fail = sum(1 for r in strat_results if r["verdict"] == "fail")
            error = sum(1 for r in strat_results if r["verdict"] == "error")
            total = len(strat_results)
            ws3.cell(row=row_idx, column=1, value=strat_name)
            ws3.cell(row=row_idx, column=2, value=ok)
            ws3.cell(row=row_idx, column=3, value=fail)
            ws3.cell(row=row_idx, column=4, value=error)
            ws3.cell(row=row_idx, column=5, value=f"{ok / total * 100:.0f}%" if total else "0%")

        for col, w in {"A": 20, "B": 8, "C": 8, "D": 8, "E": 14}.items():
            ws3.column_dimensions[col].width = w

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = f"test_403_alternatives_{timestamp}.xlsx"
        wb.save(out_path)
        console.print(f"\n[bold green]Report saved to:[/bold green] {out_path}")

    except ImportError:
        console.print("\n[yellow]openpyxl not installed — skipping Excel export[/yellow]")


if __name__ == "__main__":
    asyncio.run(main())
