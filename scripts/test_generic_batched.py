"""
Resumable batched test for generic browser boards.

Tests boards in batches, saves results to a single xlsx that accumulates
across runs. Each run picks up where the last one left off.

Usage:
    python3.13 scripts/test_generic_batched.py              # Test next 10 untested boards
    python3.13 scripts/test_generic_batched.py --batch 20   # Test next 20
    python3.13 scripts/test_generic_batched.py --status      # Show progress without testing
    python3.13 scripts/test_generic_batched.py --reset       # Clear progress and start over
    python3.13 scripts/test_generic_batched.py --timeout 45  # Custom timeout per board
    python3.13 scripts/test_generic_batched.py --retest-failures  # Re-run timeout/error boards
"""

from __future__ import annotations

import argparse
import asyncio
import importlib.util
import os
import sys
import time
from datetime import datetime
from pathlib import Path

from rich.console import Console

console = Console()

OUTPUT_FILE = "generic_board_tests.xlsx"


# ---------------------------------------------------------------------------
# Board loading
# ---------------------------------------------------------------------------

def add_repo_to_path(repo_root: Path) -> None:
    for p in [repo_root / "src", repo_root]:
        if str(p) not in sys.path:
            sys.path.insert(0, str(p))


def load_generic_boards(repo_root: Path) -> list[dict]:
    config_path = repo_root / "configs" / "config.py"
    module_name = "project_config_batched"
    spec = importlib.util.spec_from_file_location(module_name, config_path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return getattr(module, "GENERIC_BROWSER_BOARDS", [])


# ---------------------------------------------------------------------------
# Excel read/write
# ---------------------------------------------------------------------------

def sanitise(text: str | None) -> str:
    if not text:
        return ""
    return "".join(ch for ch in text if ch in ("\n", "\r", "\t") or ord(ch) >= 32)


def load_existing_results(path: Path) -> list[dict]:
    """Load previously tested results from the xlsx."""
    if not path.exists():
        return []
    from openpyxl import load_workbook
    wb = load_workbook(path)
    if "Test Results" not in wb.sheetnames:
        return []
    ws = wb["Test Results"]
    headers = [cell.value for cell in ws[1]]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        results.append({
            "company": str(row[0] or ""),
            "url": str(row[1] or ""),
            "verdict": str(row[2] or ""),
            "jobs": int(row[3] or 0),
            "elapsed": float(row[4] or 0),
            "selector": str(row[5] or ""),
            "sample_titles": [str(row[6] or ""), str(row[7] or "")],
            "sample_urls": [str(row[8] or ""), str(row[9] or "")],
            "error": str(row[10] or ""),
        })
    return results


def save_results(path: Path, results: list[dict], total_boards: int) -> None:
    """Write all results to xlsx with Test Results + Summary sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    headers = ["Company", "URL", "Verdict", "Jobs", "Elapsed (s)", "Selector",
               "Sample Title 1", "Sample Title 2", "Sample URL 1", "Sample URL 2", "Error"]
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill

    verdict_fills = {
        "ok": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "empty": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "timeout": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "error": PatternFill(start_color="E4AAFF", end_color="E4AAFF", fill_type="solid"),
    }

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["company"])
        ws.cell(row=row_idx, column=2, value=r["url"])
        vc = ws.cell(row=row_idx, column=3, value=r["verdict"])
        vc.fill = verdict_fills.get(r["verdict"], PatternFill())
        ws.cell(row=row_idx, column=4, value=r["jobs"])
        ws.cell(row=row_idx, column=5, value=r["elapsed"])
        ws.cell(row=row_idx, column=6, value=r["selector"])
        titles = r.get("sample_titles", [])
        ws.cell(row=row_idx, column=7, value=sanitise(titles[0]) if len(titles) > 0 else "")
        ws.cell(row=row_idx, column=8, value=sanitise(titles[1]) if len(titles) > 1 else "")
        urls = r.get("sample_urls", [])
        ws.cell(row=row_idx, column=9, value=sanitise(urls[0]) if len(urls) > 0 else "")
        ws.cell(row=row_idx, column=10, value=sanitise(urls[1]) if len(urls) > 1 else "")
        ws.cell(row=row_idx, column=11, value=sanitise(r.get("error", ""))[:300])

    ws.auto_filter.ref = f"A1:K{len(results) + 1}"
    for col, w in {"A": 35, "B": 55, "C": 10, "D": 8, "E": 10, "F": 20, "G": 40, "H": 40, "I": 55, "J": 55, "K": 50}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # Summary sheet
    from collections import Counter
    ws2 = wb.create_sheet("Summary")
    counts = Counter(r["verdict"] for r in results)
    total_jobs = sum(r["jobs"] for r in results)
    tested = len(results)
    remaining = total_boards - tested

    ws2.cell(row=1, column=1, value="Metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Value").font = Font(bold=True)
    summary_rows = [
        ("Total Boards", total_boards),
        ("Tested", tested),
        ("Remaining", remaining),
        ("Progress", f"{tested / total_boards * 100:.1f}%" if total_boards else "0%"),
        ("", ""),
        ("OK (found jobs)", counts.get("ok", 0)),
        ("Empty (no jobs)", counts.get("empty", 0)),
        ("Timeout", counts.get("timeout", 0)),
        ("Error", counts.get("error", 0)),
        ("", ""),
        ("Total Jobs Found", total_jobs),
        ("Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (metric, value) in enumerate(summary_rows, 2):
        ws2.cell(row=i, column=1, value=metric)
        ws2.cell(row=i, column=2, value=value)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15

    wb.save(path)


# ---------------------------------------------------------------------------
# Test runner (reused from test_generic_boards.py)
# ---------------------------------------------------------------------------

async def test_one_board(board: dict, timeout_seconds: int = 60) -> dict:
    from vacancysoft.adapters import GenericBrowserAdapter

    company = board["company"]
    url = board["url"]
    t0 = time.monotonic()

    try:
        page = await asyncio.wait_for(
            GenericBrowserAdapter().discover({
                "job_board_url": url,
                "company": company,
                "page_timeout_ms": 30000,
                "wait_after_nav_ms": 2000,
            }),
            timeout=timeout_seconds,
        )

        elapsed = round(time.monotonic() - t0, 1)
        diagnostics = getattr(page, "diagnostics", None)
        metadata = dict(getattr(diagnostics, "metadata", {})) if diagnostics else {}

        sample_titles = [getattr(j, "title_raw", None) for j in page.jobs[:2]]
        sample_urls = [getattr(j, "discovered_url", None) for j in page.jobs[:2]]

        return {
            "company": company,
            "url": url,
            "verdict": "ok" if len(page.jobs) > 0 else "empty",
            "jobs": len(page.jobs),
            "elapsed": elapsed,
            "selector": metadata.get("last_selector_used", ""),
            "sample_titles": sample_titles,
            "sample_urls": sample_urls,
            "error": "",
        }
    except asyncio.TimeoutError:
        return {
            "company": company, "url": url, "verdict": "timeout",
            "jobs": 0, "elapsed": round(time.monotonic() - t0, 1), "selector": "",
            "sample_titles": [], "sample_urls": [],
            "error": f"Timed out after {timeout_seconds}s",
        }
    except Exception as exc:
        return {
            "company": company, "url": url, "verdict": "error",
            "jobs": 0, "elapsed": round(time.monotonic() - t0, 1), "selector": "",
            "sample_titles": [], "sample_urls": [],
            "error": f"{type(exc).__name__}: {str(exc)[:150]}",
        }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def print_status(existing: list[dict], total: int) -> None:
    from collections import Counter
    tested = len(existing)
    remaining = total - tested
    counts = Counter(r["verdict"] for r in existing)
    total_jobs = sum(r["jobs"] for r in existing)

    console.print(f"\n[bold]Progress: {tested}/{total} tested ({tested / total * 100:.1f}%), {remaining} remaining[/bold]")
    console.print(f"  OK: [green]{counts.get('ok', 0)}[/green]  |  "
                  f"Empty: [yellow]{counts.get('empty', 0)}[/yellow]  |  "
                  f"Timeout: [red]{counts.get('timeout', 0)}[/red]  |  "
                  f"Error: [magenta]{counts.get('error', 0)}[/magenta]")
    console.print(f"  Total jobs found: {total_jobs}\n")


async def main() -> None:
    parser = argparse.ArgumentParser(description="Resumable batched generic board tester")
    parser.add_argument("--batch", type=int, default=10, help="Number of boards to test per run (default: 10)")
    parser.add_argument("--timeout", type=int, default=60, help="Per-board timeout in seconds (default: 60)")
    parser.add_argument("--status", action="store_true", help="Show progress without testing")
    parser.add_argument("--reset", action="store_true", help="Clear progress and start over")
    parser.add_argument("--retest-failures", action="store_true", help="Re-run timeout and error boards")
    args = parser.parse_args()

    repo_root = Path(".").resolve()
    add_repo_to_path(repo_root)
    output_path = repo_root / OUTPUT_FILE

    all_boards = load_generic_boards(repo_root)
    total = len(all_boards)

    # --reset
    if args.reset:
        if output_path.exists():
            output_path.unlink()
            console.print(f"[bold red]Reset:[/bold red] Deleted {OUTPUT_FILE}")
        else:
            console.print("Nothing to reset — no existing results file.")
        return

    # Load existing results
    existing = load_existing_results(output_path)
    tested_keys = {(r["company"], r["url"]) for r in existing}

    # --status
    if args.status:
        print_status(existing, total)
        return

    # --retest-failures: remove timeout/error results so they get retested
    if args.retest_failures:
        failures = [r for r in existing if r["verdict"] in ("timeout", "error")]
        if not failures:
            console.print("No failures to retest.")
            return
        console.print(f"[bold]Retesting {len(failures)} failed boards...[/bold]\n")
        # Remove failures from existing results
        existing = [r for r in existing if r["verdict"] not in ("timeout", "error")]
        tested_keys = {(r["company"], r["url"]) for r in existing}
        # The failures become the batch to test
        boards_to_test = [{"company": r["company"], "url": r["url"]} for r in failures]
    else:
        # Find untested boards
        untested = [b for b in all_boards if (b["company"], b["url"]) not in tested_keys]
        if not untested:
            console.print("[bold green]All boards have been tested![/bold green]")
            print_status(existing, total)
            return
        boards_to_test = untested[:args.batch]
        console.print(f"[bold]Testing batch of {len(boards_to_test)} boards "
                      f"({len(existing)}/{total} already done, {len(untested)} remaining)[/bold]\n")

    # Run tests
    new_results = []
    for idx, board in enumerate(boards_to_test, 1):
        console.print(f"[dim][{idx}/{len(boards_to_test)}] {board['company']}...[/dim]", end=" ")
        result = await test_one_board(board, timeout_seconds=args.timeout)
        new_results.append(result)

        colour = {"ok": "green", "empty": "yellow", "timeout": "red", "error": "magenta"}.get(result["verdict"], "white")
        top_title = (result["sample_titles"][0] or "—")[:40] if result["sample_titles"] else "—"
        console.print(
            f"[{colour}]{result['verdict']:7s}[/{colour}] | "
            f"jobs={result['jobs']:<4d} | {result['elapsed']:5.1f}s | {top_title}"
        )

    # Merge and save
    all_results = existing + new_results
    save_results(output_path, all_results, total)

    # Print status
    print_status(all_results, total)
    console.print(f"[bold green]Results saved to {OUTPUT_FILE}[/bold green]")


if __name__ == "__main__":
    asyncio.run(main())
