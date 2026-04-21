#!/usr/bin/env python3
"""Export every LLM prompt / category block into a multi-sheet XLSX.

One workbook, one sheet per "prompt surface":

  - Index       — overview + link to each other sheet
  - Dossier     — DOSSIER_SYSTEM + DOSSIER_TEMPLATE (base template)
  - HM search   — the category-specific HM search prompt builder
  - Campaign    — CAMPAIGN_SYSTEM + CAMPAIGN_TEMPLATE
  - Categories  — one row per category (risk / quant / compliance / …)
                  × one column per block field (research_scope,
                  market_context_guidance, outreach_angle, etc.)
  - HM queries  — the _*_HM_SEARCHES string templates, one per row

Writes to artifacts/prompts-review.xlsx (artifacts/ is gitignored per
.gitignore so the export doesn't land in git). Non-destructive: just
reads from the live Python modules, so running it at any time gives
the current state of all prompts.

Usage:
    python3 scripts/export_prompts_xlsx.py
    python3 scripts/export_prompts_xlsx.py --output /some/other/path.xlsx

Requires: openpyxl (already in main dependencies).
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vacancysoft.intelligence.prompts.base_dossier import (
    DOSSIER_SYSTEM,
    DOSSIER_TEMPLATE,
)
from vacancysoft.intelligence.prompts.base_campaign import (
    CAMPAIGN_SYSTEM,
    CAMPAIGN_TEMPLATE_V1,
    CAMPAIGN_TEMPLATE_V2,
)
from vacancysoft.intelligence.prompts.category_blocks import (
    CATEGORY_BLOCKS,
    DEFAULT_CATEGORY,
)

# The HM-search prompt is built at call time by dossier.py::_build_hm_prompt.
# Import the function so we can render a realistic example against every
# category block.
from vacancysoft.intelligence.dossier import _build_hm_prompt as build_hm_prompt


HEADER_FILL = PatternFill(start_color="FFD0D0D0", end_color="FFD0D0D0", fill_type="solid")
HEADER_FONT = Font(bold=True)
MONO_FONT = Font(name="Menlo", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def _style_prompt_cell(cell) -> None:
    cell.alignment = WRAP
    cell.font = MONO_FONT


def _autofit(ws, col_widths: dict[int, int]) -> None:
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_index_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Index"
    rows = [
        ("Sheet", "What it covers", "Where in code"),
        ("Dossier", "The master dossier prompt (system + user template)",
         "src/vacancysoft/intelligence/prompts/base_dossier.py"),
        ("HM search", "Per-category hiring-manager search prompt, as the "
                      "resolver actually builds it at call time",
         "src/vacancysoft/intelligence/dossier.py::_build_hm_prompt + "
         "prompts/category_blocks.py::_*_HM_SEARCHES"),
        ("Campaign", "Campaign email-sequence template "
                     "(5 sequences × 6 tones). Row 3 is V2 (live, default); "
                     "row 4 is V1 (legacy rollback target — not in use).",
         "src/vacancysoft/intelligence/prompts/base_campaign.py::"
         "CAMPAIGN_TEMPLATE_V2 / CAMPAIGN_TEMPLATE_V1"),
        ("Categories", "Per-category blocks: research scope, market context, "
                       "outreach angle, HM function guidance (injected into the templates above)",
         "src/vacancysoft/intelligence/prompts/category_blocks.py::CATEGORY_BLOCKS"),
        ("HM queries", "Per-category LinkedIn search query templates "
                       "(one sheet row per category)",
         "src/vacancysoft/intelligence/prompts/category_blocks.py::_*_HM_SEARCHES"),
    ]
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                _style_header(cell)
            else:
                cell.alignment = WRAP
    _autofit(ws, {1: 22, 2: 60, 3: 60})
    ws.freeze_panes = "A2"


def write_dossier_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Dossier")
    ws.cell(row=1, column=1, value="Field")
    ws.cell(row=1, column=2, value="Content")
    for c in (ws.cell(row=1, column=1), ws.cell(row=1, column=2)):
        _style_header(c)
    ws.cell(row=2, column=1, value="DOSSIER_SYSTEM (system message)")
    ws.cell(row=2, column=2, value=DOSSIER_SYSTEM)
    _style_prompt_cell(ws.cell(row=2, column=2))
    ws.cell(row=3, column=1, value="DOSSIER_TEMPLATE (user message, with placeholders)")
    ws.cell(row=3, column=2, value=DOSSIER_TEMPLATE)
    _style_prompt_cell(ws.cell(row=3, column=2))
    ws.row_dimensions[3].height = 500
    _autofit(ws, {1: 30, 2: 120})
    ws.freeze_panes = "A2"


def write_hm_search_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("HM search")

    # Header
    cols = ("Category", "System message", "User message (rendered for a sample job)")
    for i, h in enumerate(cols, start=1):
        _style_header(ws.cell(row=1, column=i, value=h))

    sample_job = {
        "company": "EXAMPLE BANK PLC",
        "title": "Head of Credit Risk",
        "location": "London, United Kingdom",
        "description": "[job description text would appear here]",
    }

    row = 2
    for cat in sorted(CATEGORY_BLOCKS.keys()):
        messages = build_hm_prompt(sample_job, cat)
        system = next((m["content"] for m in messages if m["role"] == "system"), "")
        user = next((m["content"] for m in messages if m["role"] == "user"), "")
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=system)
        ws.cell(row=row, column=3, value=user)
        _style_prompt_cell(ws.cell(row=row, column=2))
        _style_prompt_cell(ws.cell(row=row, column=3))
        ws.row_dimensions[row].height = 300
        row += 1
    _autofit(ws, {1: 15, 2: 40, 3: 100})
    ws.freeze_panes = "A2"


def write_campaign_sheet(wb: Workbook) -> None:
    """Campaign sheet — v2 (live, default) at the top, v1 (rollback target)
    archived below. Operators almost always want v2; v1 is kept in the
    workbook so a rollback review has the legacy text to hand."""
    ws = wb.create_sheet("Campaign")
    ws.cell(row=1, column=1, value="Field")
    ws.cell(row=1, column=2, value="Content")
    for c in (ws.cell(row=1, column=1), ws.cell(row=1, column=2)):
        _style_header(c)

    ws.cell(row=2, column=1, value="CAMPAIGN_SYSTEM (system message)")
    ws.cell(row=2, column=2, value=CAMPAIGN_SYSTEM)
    _style_prompt_cell(ws.cell(row=2, column=2))

    ws.cell(row=3, column=1, value="CAMPAIGN_TEMPLATE_V2 (live — user message, with placeholders)")
    ws.cell(row=3, column=2, value=CAMPAIGN_TEMPLATE_V2)
    _style_prompt_cell(ws.cell(row=3, column=2))
    ws.row_dimensions[3].height = 750

    ws.cell(row=4, column=1, value="CAMPAIGN_TEMPLATE_V1 (legacy / rollback target)")
    ws.cell(row=4, column=2, value=CAMPAIGN_TEMPLATE_V1)
    _style_prompt_cell(ws.cell(row=4, column=2))
    ws.row_dimensions[4].height = 500

    _autofit(ws, {1: 48, 2: 120})
    ws.freeze_panes = "A2"


def write_categories_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Categories")
    # Discover all keys across all categories
    all_keys = sorted({k for block in CATEGORY_BLOCKS.values() for k in block.keys()})
    # Put the most-common fields first
    priority = ["research_scope", "market_context_guidance", "search_boolean_guidance",
                "outreach_angle", "hm_function_guidance", "hm_search_queries"]
    ordered_keys = [k for k in priority if k in all_keys] + [k for k in all_keys if k not in priority]

    # Header: Category + one column per block field
    ws.cell(row=1, column=1, value="Category")
    _style_header(ws.cell(row=1, column=1))
    for i, key in enumerate(ordered_keys, start=2):
        c = ws.cell(row=1, column=i, value=key)
        _style_header(c)

    row = 2
    for cat in sorted(CATEGORY_BLOCKS.keys()):
        block = CATEGORY_BLOCKS[cat]
        ws.cell(row=row, column=1, value=cat + (" (DEFAULT)" if cat == DEFAULT_CATEGORY else ""))
        for i, key in enumerate(ordered_keys, start=2):
            val = block.get(key, "")
            c = ws.cell(row=row, column=i, value=val)
            _style_prompt_cell(c)
        ws.row_dimensions[row].height = 220
        row += 1

    widths = {1: 18}
    for i, _ in enumerate(ordered_keys, start=2):
        widths[i] = 60
    _autofit(ws, widths)
    ws.freeze_panes = "B2"


def write_hm_queries_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("HM queries")
    ws.cell(row=1, column=1, value="Category")
    ws.cell(row=1, column=2, value="hm_search_queries template")
    for c in (ws.cell(row=1, column=1), ws.cell(row=1, column=2)):
        _style_header(c)
    row = 2
    for cat in sorted(CATEGORY_BLOCKS.keys()):
        queries = CATEGORY_BLOCKS[cat].get("hm_search_queries", "")
        ws.cell(row=row, column=1, value=cat)
        c = ws.cell(row=row, column=2, value=queries)
        _style_prompt_cell(c)
        ws.row_dimensions[row].height = 180
        row += 1
    _autofit(ws, {1: 15, 2: 120})
    ws.freeze_panes = "A2"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0] if __doc__ else "")
    parser.add_argument(
        "--output",
        default=str(PROJECT_ROOT / "artifacts" / "prompts-review.xlsx"),
        help="Destination .xlsx path (default: artifacts/prompts-review.xlsx)",
    )
    args = parser.parse_args()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    write_index_sheet(wb)
    write_dossier_sheet(wb)
    write_hm_search_sheet(wb)
    write_campaign_sheet(wb)
    write_categories_sheet(wb)
    write_hm_queries_sheet(wb)

    wb.save(output_path)
    print(f"Wrote {output_path}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  Size:   {output_path.stat().st_size:,} bytes")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
