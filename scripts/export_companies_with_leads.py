#!/usr/bin/env python3
"""Export every company in the Sources 'with leads' view to XLSX.

Admin-only script. Answers "which employers are contributing leads,
how many, and in what shape?" — the per-employer aggregation the
Sources page's "with leads" view shows, downloadable for audit.

Reuses the live ``_get_cached_ledger`` helper from
``src/vacancysoft/api/ledger.py`` so the export matches what the UI
shows byte-for-byte. No new DB queries, no schema change. "With
leads" = at least one enriched job contributed to the card, which
is exactly what the Sources page filters on.

Output: ``artifacts/companies-with-leads-<YYYY-MM-DD>.xlsx``.

Sheets:
  * Summary   — totals + per-category rollup across all companies
                with leads. Quick sanity check: "Risk accounts for
                X% of our lead volume, split across N companies."
  * Companies — one row per employer card. Columns:
                Company, Adapter, Base URL, Source Type, Leads,
                Raw Jobs Scraped, Conversion %, Risk, Quant,
                Compliance, Audit, Cyber, Legal, Front Office,
                Top Countries, Aggregator Hits, Employment Types,
                Last Run Status, Last Run Error.
                Excel autofilter on; sorted by lead count desc so
                the biggest contributors surface at the top.

Deliberately includes aggregator-only cards (source_type=aggregator,
Adapter column empty because there's no direct Source row — the
employer was discovered via Reed / Adzuna / etc. but isn't a
first-party board in our scrape list). Flag with ``--direct-only``
to filter them out.

Usage:
    python3 scripts/export_companies_with_leads.py
    python3 scripts/export_companies_with_leads.py --output ~/Desktop/companies.xlsx
    python3 scripts/export_companies_with_leads.py --direct-only
    python3 scripts/export_companies_with_leads.py --country "United Kingdom"

Requires: openpyxl (already in main dependencies).
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vacancysoft.api.ledger import _get_cached_ledger


# Canonical category order mirrors the Sources page's pill row so
# the xlsx columns read the same way an operator scans the UI.
CATEGORY_ORDER = [
    "Risk", "Quant", "Compliance", "Audit", "Cyber", "Legal", "Front Office",
]


# ── Styling ────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="FFD0D0D0", end_color="FFD0D0D0", fill_type="solid")
HEADER_FONT = Font(bold=True)


def _style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autofit(ws, col_widths: dict[int, int]) -> None:
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ── Data assembly ──────────────────────────────────────────────────


def _top_countries_text(categories_by_country: dict[str, dict[str, int]]) -> str:
    """Flatten {country: {cat: count}} into a "UK: 20, USA: 8" string
    of the top contributing countries. Top-3 to keep the cell
    readable."""
    totals: list[tuple[str, int]] = []
    for country, cat_counts in (categories_by_country or {}).items():
        if not country:
            continue
        total = sum(cat_counts.values())
        if total > 0:
            totals.append((country, total))
    totals.sort(key=lambda kv: kv[1], reverse=True)
    return ", ".join(f"{c}: {n}" for c, n in totals[:3])


def _aggregator_hits_text(hits: dict[str, int]) -> str:
    """Render aggregator_hits as "reed: 2, adzuna: 5"."""
    if not hits:
        return ""
    return ", ".join(f"{a}: {n}" for a, n in sorted(hits.items(), key=lambda kv: kv[1], reverse=True))


def _employment_types_text(types: dict[str, int]) -> str:
    if not types:
        return ""
    return ", ".join(f"{t}: {n}" for t, n in sorted(types.items(), key=lambda kv: kv[1], reverse=True))


def _conversion_pct(leads: int, raw_jobs: int) -> float | str:
    """leads / raw_jobs as percentage. When raw_jobs is 0 (happens
    for aggregator-only cards where the employer has no direct
    source), return an empty string — percentage is meaningless."""
    if raw_jobs <= 0:
        return ""
    return round(100 * leads / raw_jobs, 1)


def _collect_rows(country: str | None, direct_only: bool) -> tuple[list[dict], dict]:
    """Pull from the live ledger. ``_get_cached_ledger(country=)``
    returns exactly the card list the Sources page renders, already
    aggregated per-employer with cross-source dedupe. We just filter
    to 'with leads' and shape for XLSX."""
    ledger = _get_cached_ledger(country=country)

    rows: list[dict] = []
    total_leads_by_category: dict[str, int] = {}
    total_companies_by_category: dict[str, int] = {}

    for card in ledger:
        lead_count = len(card.get("lead_ids") or [])
        if lead_count == 0:
            continue
        # The "with leads" filter corresponds to SourceView='leads' in
        # the UI — cards that contributed at least one enriched job.

        adapter = card.get("adapter_name") or ""
        is_direct = bool(adapter) and card.get("seed_type") != "aggregator"
        if direct_only and not is_direct:
            continue

        categories = card.get("categories") or {}
        row = {
            "company": card.get("employer_display") or "",
            "adapter": adapter,
            "base_url": card.get("base_url") or "",
            "source_type": "direct" if is_direct else "aggregator",
            "leads": lead_count,
            "raw_jobs_scraped": int(card.get("raw_jobs_count") or 0),
            "conversion_pct": _conversion_pct(
                lead_count, int(card.get("raw_jobs_count") or 0)
            ),
            "top_countries": _top_countries_text(card.get("categories_by_country") or {}),
            "aggregator_hits": _aggregator_hits_text(card.get("aggregator_hits") or {}),
            "employment_types": _employment_types_text(card.get("employment_types") or {}),
            "last_run_status": card.get("last_run_status") or "",
            "last_run_error": (card.get("last_run_error") or "")[:300],
        }
        for cat in CATEGORY_ORDER:
            row[f"cat_{cat}"] = int(categories.get(cat, 0))
            if categories.get(cat, 0) > 0:
                total_leads_by_category[cat] = total_leads_by_category.get(cat, 0) + categories[cat]
                total_companies_by_category[cat] = total_companies_by_category.get(cat, 0) + 1

        rows.append(row)

    # Sort: biggest contributors first — matches how an operator thinks.
    rows.sort(key=lambda r: (-r["leads"], r["company"].lower()))

    summary = {
        "total_companies": len(rows),
        "total_leads": sum(r["leads"] for r in rows),
        "total_raw_jobs": sum(r["raw_jobs_scraped"] for r in rows),
        "direct_count": sum(1 for r in rows if r["source_type"] == "direct"),
        "aggregator_only_count": sum(1 for r in rows if r["source_type"] == "aggregator"),
        "leads_by_category": total_leads_by_category,
        "companies_by_category": total_companies_by_category,
    }
    return rows, summary


# ── Sheet writers ──────────────────────────────────────────────────


def write_companies_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Companies")
    headers = [
        "Company",
        "Adapter",
        "Base URL",
        "Source Type",
        "Leads",
        "Raw Jobs Scraped",
        "Conversion %",
    ] + CATEGORY_ORDER + [
        "Top Countries",
        "Aggregator Hits",
        "Employment Types",
        "Last Run Status",
        "Last Run Error",
    ]
    for col_idx, h in enumerate(headers, start=1):
        _style_header(ws.cell(row=1, column=col_idx, value=h))

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r["company"])
        ws.cell(row=row_idx, column=2, value=r["adapter"])
        ws.cell(row=row_idx, column=3, value=r["base_url"])
        ws.cell(row=row_idx, column=4, value=r["source_type"])
        ws.cell(row=row_idx, column=5, value=r["leads"])
        ws.cell(row=row_idx, column=6, value=r["raw_jobs_scraped"])
        ws.cell(row=row_idx, column=7, value=r["conversion_pct"])
        for i, cat in enumerate(CATEGORY_ORDER, start=8):
            ws.cell(row=row_idx, column=i, value=r[f"cat_{cat}"])
        base = 7 + len(CATEGORY_ORDER)
        ws.cell(row=row_idx, column=base + 1, value=r["top_countries"])
        ws.cell(row=row_idx, column=base + 2, value=r["aggregator_hits"])
        ws.cell(row=row_idx, column=base + 3, value=r["employment_types"])
        ws.cell(row=row_idx, column=base + 4, value=r["last_run_status"])
        ws.cell(row=row_idx, column=base + 5, value=r["last_run_error"])

    widths = {1: 32, 2: 16, 3: 42, 4: 12, 5: 8, 6: 12, 7: 12}
    for i in range(len(CATEGORY_ORDER)):
        widths[8 + i] = 10
    base = 7 + len(CATEGORY_ORDER)
    widths[base + 1] = 28
    widths[base + 2] = 24
    widths[base + 3] = 22
    widths[base + 4] = 14
    widths[base + 5] = 60
    _autofit(ws, widths)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_summary_sheet(wb: Workbook, summary: dict, args) -> None:
    ws = wb.active
    ws.title = "Summary"

    now = datetime.utcnow()
    rows: list[tuple[str, object]] = [
        ("Companies with leads — audit report", ""),
        ("Report generated", now.strftime("%Y-%m-%d %H:%M UTC")),
        ("Country filter", args.country or "(all countries)"),
        ("Aggregator-only cards", "Excluded" if args.direct_only else "Included"),
        ("", ""),
        ("Total companies with leads", summary["total_companies"]),
        ("  with a direct source", summary["direct_count"]),
        ("  aggregator-only (no direct board)", summary["aggregator_only_count"]),
        ("", ""),
        ("Total leads across all companies", summary["total_leads"]),
        ("Total raw jobs scraped (direct sources only)", summary["total_raw_jobs"]),
        ("", ""),
        ("Per-category rollup", ""),
        ("  Category / Leads / Companies contributing", ""),
    ]
    for cat in CATEGORY_ORDER:
        leads = summary["leads_by_category"].get(cat, 0)
        comps = summary["companies_by_category"].get(cat, 0)
        rows.append((f"  {cat}", f"{leads} leads across {comps} compan{'y' if comps == 1 else 'ies'}"))
    rows.append(("", ""))
    rows.append(("See 'Companies' sheet — one row per employer, Excel autofilter on every column.", ""))

    for row_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    _autofit(ws, {1: 56, 2: 42})


# ── CLI ────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0] if __doc__ else "")
    default_output = (
        PROJECT_ROOT
        / "artifacts"
        / f"companies-with-leads-{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    )
    parser.add_argument(
        "--output",
        default=str(default_output),
        help=f"Output .xlsx path (default: {default_output})",
    )
    parser.add_argument(
        "--direct-only",
        action="store_true",
        help="Exclude aggregator-only cards (employers with no direct source).",
    )
    parser.add_argument(
        "--country",
        default=None,
        help="Country filter passed to the ledger (matches the Sources page country dropdown).",
    )
    args = parser.parse_args()

    out_path = Path(args.output).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print("Loading ledger…", flush=True)
    rows, summary = _collect_rows(args.country, args.direct_only)

    print(
        f"  {summary['total_companies']} companies with leads "
        f"({summary['direct_count']} direct · "
        f"{summary['aggregator_only_count']} aggregator-only) · "
        f"{summary['total_leads']} total leads · "
        f"{summary['total_raw_jobs']} raw jobs scraped",
        flush=True,
    )

    wb = Workbook()
    write_summary_sheet(wb, summary, args)
    write_companies_sheet(wb, rows)

    wb.save(str(out_path))
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
