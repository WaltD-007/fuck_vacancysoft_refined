"""Export the full scraping + hiring-manager taxonomy to an .xlsx workbook.

Sheets produced:
  1. Overview               — one-line description of every other sheet
  2. Scraping — Rules        — weighted classifier rules (category, keyword, weight)
  3. Scraping — Relevance    — HIGH_RELEVANCE_PHRASES + MEDIUM_RELEVANCE_WORDS gate
  4. Scraping — Blocklist    — regex alternates in _TITLE_BLOCKLIST
  5. Scraping — Legacy YAML  — segments + aliases from configs/review/title_taxonomy.yaml
  6. HM — Category Context   — per-category research scope / guidance / outreach angle
  7. HM — LinkedIn Searches  — per-category Boolean search templates
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = REPO_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

from vacancysoft.classifiers.taxonomy import _TAXONOMY_RULES, _TITLE_BLOCKLIST  # noqa: E402
from vacancysoft.classifiers.title_rules import (  # noqa: E402
    HIGH_RELEVANCE_PHRASES,
    MEDIUM_RELEVANCE_WORDS,
)
from vacancysoft.intelligence.prompts.category_blocks import CATEGORY_BLOCKS  # noqa: E402


# ── Styling helpers ────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
TITLE_FONT = Font(name="Arial", bold=True, size=14)


def _write_header(sheet, headers: list[str]) -> None:
    for col, label in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "A2"


def _style_body(sheet, start_row: int = 2) -> None:
    for row in sheet.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value is None:
                continue
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN


def _autosize(sheet, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


# ── Sheet builders ─────────────────────────────────────────────────────────

def build_overview(wb: Workbook) -> None:
    sheet = wb.active
    sheet.title = "Overview"
    sheet["A1"] = "Prospero taxonomy export"
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells("A1:C1")

    sheet["A3"] = "Export date"
    from datetime import date
    sheet["B3"] = date.today().isoformat()
    sheet["A3"].font = Font(name="Arial", bold=True, size=10)

    rows = [
        ("Sheet", "What it contains", "Source file"),
        (
            "Scraping — Rules",
            "Weighted keyword/phrase rules used to classify scraped job titles into one of 7 categories. "
            "Higher weight = stronger signal. Phrases are matched with word boundaries, longest first.",
            "src/vacancysoft/classifiers/taxonomy.py (_TAXONOMY_RULES)",
        ),
        (
            "Scraping — Relevance",
            "Flat list of phrases and single words used by the 'is this title relevant at all' gate. "
            "HIGH phrases score 0.95, MEDIUM words score 0.80. Miss = 0.15.",
            "src/vacancysoft/classifiers/title_rules.py (HIGH_RELEVANCE_PHRASES, MEDIUM_RELEVANCE_WORDS)",
        ),
        (
            "Scraping — Blocklist",
            "Regex alternates that force a title to 'not relevant' even when it contains taxonomy keywords. "
            "Used to kill retail/non-financial jobs that happen to mention words like 'trading' or 'audit'.",
            "src/vacancysoft/classifiers/taxonomy.py (_TITLE_BLOCKLIST)",
        ),
        (
            "Scraping — Legacy YAML",
            "Legacy v1 segment keys + human labels + coarse aliases. Used as the taxonomy_version identifier "
            "and as a display label source.",
            "configs/review/title_taxonomy.yaml",
        ),
        (
            "HM — Category Context",
            "Per-category research/market/boolean/outreach guidance injected into the dossier prompt. "
            "Defines the 'voice' for the hiring-manager search by category.",
            "src/vacancysoft/intelligence/prompts/category_blocks.py (CATEGORY_BLOCKS, non-search fields)",
        ),
        (
            "HM — LinkedIn Searches",
            "Per-category 7-step LinkedIn search recipes used by the HM search LLM call. "
            "These are templates — [company name], [function], and [asset class] are filled in at runtime.",
            "src/vacancysoft/intelligence/prompts/category_blocks.py (_*_HM_SEARCHES strings)",
        ),
    ]
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 5:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
            else:
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN
    sheet.row_dimensions[5].height = 24
    _autosize(sheet, {1: 28, 2: 80, 3: 55})
    for r in range(6, 5 + len(rows)):
        sheet.row_dimensions[r].height = 55


def build_scraping_rules(wb: Workbook) -> None:
    sheet = wb.create_sheet("Scraping — Rules")
    _write_header(sheet, ["Category", "Keyword / Phrase", "Weight", "Token count"])
    row = 2
    for category, rules in _TAXONOMY_RULES.items():
        # Sort by weight desc, then alphabetically
        sorted_rules = sorted(rules, key=lambda r: (-r[1], r[0]))
        for phrase, weight in sorted_rules:
            sheet.cell(row=row, column=1, value=category)
            sheet.cell(row=row, column=2, value=phrase)
            sheet.cell(row=row, column=3, value=weight)
            sheet.cell(row=row, column=4, value=len(phrase.split()))
            row += 1
    _style_body(sheet)
    _autosize(sheet, {1: 14, 2: 40, 3: 8, 4: 12})


def build_relevance(wb: Workbook) -> None:
    sheet = wb.create_sheet("Scraping — Relevance")
    _write_header(sheet, ["Tier", "Entry", "Score on match", "Token count"])
    row = 2
    for phrase in HIGH_RELEVANCE_PHRASES:
        sheet.cell(row=row, column=1, value="HIGH phrase")
        sheet.cell(row=row, column=2, value=phrase)
        sheet.cell(row=row, column=3, value=0.95)
        sheet.cell(row=row, column=4, value=len(phrase.split()))
        row += 1
    for word in MEDIUM_RELEVANCE_WORDS:
        sheet.cell(row=row, column=1, value="MEDIUM word")
        sheet.cell(row=row, column=2, value=word)
        sheet.cell(row=row, column=3, value=0.80)
        sheet.cell(row=row, column=4, value=len(word.split()))
        row += 1
    _style_body(sheet)
    _autosize(sheet, {1: 16, 2: 40, 3: 16, 4: 12})


def build_blocklist(wb: Workbook) -> None:
    sheet = wb.create_sheet("Scraping — Blocklist")
    _write_header(sheet, ["Alternate #", "Regex fragment", "Readable intent"])
    # Pull the top-level alternates out of the compiled regex pattern.
    # _TITLE_BLOCKLIST.pattern is the raw string including \b(...)\b wrapper.
    raw = _TITLE_BLOCKLIST.pattern
    # Strip the outer \b(...)\b
    match = re.match(r"\\b\((.*)\)\\b", raw, re.DOTALL)
    inside = match.group(1) if match else raw

    # Split on top-level | (not inside parentheses). Minimal depth tracker.
    alternates = []
    buf: list[str] = []
    depth = 0
    i = 0
    while i < len(inside):
        ch = inside[i]
        if ch == "\\" and i + 1 < len(inside):
            buf.append(inside[i:i + 2])
            i += 2
            continue
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if ch == "|" and depth == 0:
            alternates.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
        i += 1
    if buf:
        alternates.append("".join(buf).strip())

    intent_hints = {
        "trading assistant": "Retail store 'trading assistant' (not markets trader)",
        "estore trading|store trading|retail trading": "Retail-store variants of 'trading'",
        "shelf stacker|stock replenish|merchandis": "Retail shop-floor jobs",
        "checkout|cashier|store manager|shop manager|retail assistant": "Retail-store roles",
        "customer assistant|customer team": "Retail / customer-facing non-finance",
        "warehouse operative|delivery driver": "Logistics non-finance",
        "actuar": "Insurance actuarial — excluded unless explicitly in taxonomy",
        "underwrit": "Insurance underwriting — excluded unless credit-underwriting",
    }

    for idx, alt in enumerate(alternates, start=1):
        sheet.cell(row=idx + 1, column=1, value=idx)
        sheet.cell(row=idx + 1, column=2, value=alt)
        # Find a hint by substring match
        hint = ""
        for key, val in intent_hints.items():
            if key in alt:
                hint = val
                break
        sheet.cell(row=idx + 1, column=3, value=hint)
    _style_body(sheet)
    _autosize(sheet, {1: 10, 2: 80, 3: 50})


def build_legacy_yaml(wb: Workbook) -> None:
    sheet = wb.create_sheet("Scraping — Legacy YAML")
    yaml_path = REPO_ROOT / "configs" / "review" / "title_taxonomy.yaml"
    data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
    version = data.get("taxonomy_version", "")
    _write_header(sheet, ["Key", "Label", "Aliases", "Taxonomy version"])
    row = 2
    for segment in data.get("segments", []):
        sheet.cell(row=row, column=1, value=segment.get("key"))
        sheet.cell(row=row, column=2, value=segment.get("label"))
        sheet.cell(row=row, column=3, value=", ".join(segment.get("aliases", [])))
        sheet.cell(row=row, column=4, value=version)
        row += 1
    _style_body(sheet)
    _autosize(sheet, {1: 16, 2: 20, 3: 60, 4: 16})


def build_hm_context(wb: Workbook) -> None:
    sheet = wb.create_sheet("HM — Category Context")
    _write_header(
        sheet,
        [
            "Category",
            "Research scope",
            "Market context guidance",
            "Boolean search guidance",
            "HM function guidance",
            "Outreach angle",
        ],
    )
    row = 2
    for category, block in CATEGORY_BLOCKS.items():
        sheet.cell(row=row, column=1, value=category)
        sheet.cell(row=row, column=2, value=block.get("research_scope", ""))
        sheet.cell(row=row, column=3, value=block.get("market_context_guidance", ""))
        sheet.cell(row=row, column=4, value=block.get("search_boolean_guidance", ""))
        sheet.cell(row=row, column=5, value=block.get("hm_function_guidance", ""))
        sheet.cell(row=row, column=6, value=block.get("outreach_angle", ""))
        row += 1
    _style_body(sheet)
    _autosize(sheet, {1: 14, 2: 40, 3: 55, 4: 55, 5: 45, 6: 55})
    for r in range(2, row):
        sheet.row_dimensions[r].height = 100


def build_hm_searches(wb: Workbook) -> None:
    sheet = wb.create_sheet("HM — LinkedIn Searches")
    _write_header(sheet, ["Category", "Search #", "Query template"])
    row = 2
    for category, block in CATEGORY_BLOCKS.items():
        raw = block.get("hm_search_queries", "")
        for line in raw.splitlines():
            line = line.strip()
            if not line:
                continue
            # Format: "Search N: query"
            m = re.match(r"^Search\s+(\d+):\s*(.+)$", line)
            if m:
                search_num = int(m.group(1))
                query = m.group(2)
            else:
                search_num = None
                query = line
            sheet.cell(row=row, column=1, value=category)
            sheet.cell(row=row, column=2, value=search_num)
            sheet.cell(row=row, column=3, value=query)
            row += 1
    _style_body(sheet)
    _autosize(sheet, {1: 14, 2: 10, 3: 90})


# ── Entrypoint ─────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description="Export Prospero taxonomy to xlsx.")
    parser.add_argument(
        "--output",
        type=Path,
        default=REPO_ROOT / "taxonomy_export.xlsx",
        help="Output .xlsx path (default: repo-root/taxonomy_export.xlsx).",
    )
    args = parser.parse_args()

    wb = Workbook()
    build_overview(wb)
    build_scraping_rules(wb)
    build_relevance(wb)
    build_blocklist(wb)
    build_legacy_yaml(wb)
    build_hm_context(wb)
    build_hm_searches(wb)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
