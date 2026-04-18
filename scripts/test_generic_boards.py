"""
Test the updated generic boards — scrape each one and report results.

Focuses on:
  - The 14 URL-changed boards
  - The 5 scroll-to-bottom boards
  - The 2 missed-jobs boards
  - Any board you pass via --company filter

Usage:
    python3.13 scripts/test_generic_boards.py                    # Test all changed boards
    python3.13 scripts/test_generic_boards.py --all              # Test ALL generic boards
    python3.13 scripts/test_generic_boards.py --company "Allianz"  # Test one company
    python3.13 scripts/test_generic_boards.py --limit 20         # Test first N generic boards
"""

from __future__ import annotations

import asyncio
import importlib.util
import sys
import time
from datetime import datetime
from pathlib import Path

from rich.console import Console
from rich.table import Table

console = Console()

# Boards that were changed and need validation
CHANGED_BOARDS = {
    # URL changes
    "Affirm", "Aioi Nissay Dowa Insurance UK Limited", "Aldermore Bank",
    "motonovo finance", "Allied Irish Banks", "Ally Invest", "Alpaca",
    "Alphadyne", "Aon", "Apex Fintech Solutions",
    "AQR Capital Management (Europe) LLP", "Ardian", "Aurora Energy Research",
    # Scroll-to-bottom boards
    "Accelerant Insurance UK Limited", "Accredited Insurance (UK) Limited",
    "Addepar", "Alan Boswell Group", "Atrium",
    # Missed jobs
    "Allianz", "Atradius",
    # B&CE fix
    "B&CE",
}


def add_repo_to_path(repo_root: Path) -> None:
    src_path = repo_root / "src"
    if str(src_path) not in sys.path:
        sys.path.insert(0, str(src_path))
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))


def load_generic_boards(repo_root: Path) -> list[dict]:
    config_path = repo_root / "configs" / "config.py"
    module_name = "project_config_test"
    spec = importlib.util.spec_from_file_location(module_name, config_path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return getattr(module, "GENERIC_BROWSER_BOARDS", [])


def sanitise(text: str | None) -> str:
    if not text:
        return ""
    return "".join(ch for ch in text if ch in ("\n", "\r", "\t") or ord(ch) >= 32)


async def test_one_board(board: dict, timeout_seconds: int = 60) -> dict:
    from vacancysoft.adapters import GenericBrowserAdapter

    company = board["company"]
    url = board["url"]
    t0 = time.monotonic()

    try:
        page = await asyncio.wait_for(
            GenericBrowserAdapter().discover({
                "job_board_url": url,
                "company": company,
                "page_timeout_ms": 30000,
                "wait_after_nav_ms": 2000,
            }),
            timeout=timeout_seconds,
        )

        elapsed = round(time.monotonic() - t0, 1)
        diagnostics = getattr(page, "diagnostics", None)
        metadata = dict(getattr(diagnostics, "metadata", {})) if diagnostics else {}
        counters = dict(getattr(diagnostics, "counters", {})) if diagnostics else {}

        sample_titles = [getattr(j, "title_raw", None) for j in page.jobs[:3]]
        sample_urls = [getattr(j, "discovered_url", None) for j in page.jobs[:3]]

        verdict = "ok" if len(page.jobs) > 0 else "empty"

        return {
            "company": company,
            "url": url,
            "verdict": verdict,
            "jobs": len(page.jobs),
            "elapsed": elapsed,
            "selector": metadata.get("last_selector_used", ""),
            "sample_titles": sample_titles,
            "sample_urls": sample_urls,
            "error": "",
        }
    except asyncio.TimeoutError:
        elapsed = round(time.monotonic() - t0, 1)
        return {
            "company": company, "url": url, "verdict": "timeout",
            "jobs": 0, "elapsed": elapsed, "selector": "",
            "sample_titles": [], "sample_urls": [],
            "error": f"Timed out after {timeout_seconds}s",
        }
    except Exception as exc:
        elapsed = round(time.monotonic() - t0, 1)
        return {
            "company": company, "url": url, "verdict": "error",
            "jobs": 0, "elapsed": elapsed, "selector": "",
            "sample_titles": [], "sample_urls": [],
            "error": f"{type(exc).__name__}: {str(exc)[:150]}",
        }


async def main() -> None:
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--all", action="store_true", help="Test ALL generic boards")
    parser.add_argument("--company", type=str, help="Test a specific company")
    parser.add_argument("--limit", type=int, default=0, help="Test first N boards")
    parser.add_argument("--timeout", type=int, default=60, help="Per-board timeout in seconds")
    args = parser.parse_args()

    repo_root = Path(".").resolve()
    add_repo_to_path(repo_root)

    all_boards = load_generic_boards(repo_root)

    if args.company:
        boards = [b for b in all_boards if args.company.lower() in b["company"].lower()]
        if not boards:
            console.print(f"[red]No boards matching '{args.company}'[/red]")
            return
    elif args.all:
        boards = all_boards
    elif args.limit > 0:
        boards = all_boards[:args.limit]
    else:
        boards = [b for b in all_boards if b["company"] in CHANGED_BOARDS]

    console.print(f"[bold]Testing {len(boards)} generic boards (timeout={args.timeout}s per board)[/bold]\n")

    results = []
    for idx, board in enumerate(boards, 1):
        console.print(f"[dim][{idx}/{len(boards)}] {board['company']}...[/dim]", end=" ")
        result = await test_one_board(board, timeout_seconds=args.timeout)
        results.append(result)

        colour = {"ok": "green", "empty": "yellow", "timeout": "red", "error": "magenta"}.get(result["verdict"], "white")
        top_title = (result["sample_titles"][0] or "—")[:40] if result["sample_titles"] else "—"
        console.print(
            f"[{colour}]{result['verdict']:7s}[/{colour}] | "
            f"jobs={result['jobs']:<4d} | {result['elapsed']:5.1f}s | {top_title}"
        )

    # Summary table
    console.print()
    table = Table(title=f"Generic Board Test ({len(results)} boards)")
    table.add_column("Company", style="bold", max_width=35)
    table.add_column("Verdict")
    table.add_column("Jobs", justify="right")
    table.add_column("Time", justify="right")
    table.add_column("Selector", max_width=20)
    table.add_column("Sample Title", max_width=40)
    table.add_column("Error", max_width=50)

    for r in results:
        colour = {"ok": "green", "empty": "yellow", "timeout": "red", "error": "magenta"}.get(r["verdict"], "white")
        title = sanitise((r["sample_titles"][0] or ""))[:40] if r["sample_titles"] else ""
        table.add_row(
            r["company"][:35],
            f"[{colour}]{r['verdict']}[/{colour}]",
            str(r["jobs"]),
            f"{r['elapsed']:.1f}s",
            r["selector"][:20],
            title,
            sanitise(r["error"])[:50],
        )
    console.print(table)

    # Verdict summary
    from collections import Counter
    counts = Counter(r["verdict"] for r in results)
    console.print(f"\n[bold]Verdicts:[/bold] {dict(counts)}")
    total_jobs = sum(r["jobs"] for r in results)
    console.print(f"[bold]Total jobs found:[/bold] {total_jobs}")

    # Export to Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"

        headers = ["Company", "URL", "Verdict", "Jobs", "Elapsed (s)", "Selector",
                   "Sample Title 1", "Sample Title 2", "Sample URL 1", "Sample URL 2", "Error"]
        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hfont
            cell.fill = hfill

        verdict_fills = {
            "ok": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "empty": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "timeout": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "error": PatternFill(start_color="E4AAFF", end_color="E4AAFF", fill_type="solid"),
        }

        for row_idx, r in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=r["company"])
            ws.cell(row=row_idx, column=2, value=r["url"])
            vc = ws.cell(row=row_idx, column=3, value=r["verdict"])
            vc.fill = verdict_fills.get(r["verdict"], PatternFill())
            ws.cell(row=row_idx, column=4, value=r["jobs"])
            ws.cell(row=row_idx, column=5, value=r["elapsed"])
            ws.cell(row=row_idx, column=6, value=r["selector"])
            titles = r.get("sample_titles", [])
            ws.cell(row=row_idx, column=7, value=sanitise(titles[0]) if len(titles) > 0 and titles[0] else "")
            ws.cell(row=row_idx, column=8, value=sanitise(titles[1]) if len(titles) > 1 and titles[1] else "")
            urls = r.get("sample_urls", [])
            ws.cell(row=row_idx, column=9, value=sanitise(urls[0]) if len(urls) > 0 and urls[0] else "")
            ws.cell(row=row_idx, column=10, value=sanitise(urls[1]) if len(urls) > 1 and urls[1] else "")
            ws.cell(row=row_idx, column=11, value=sanitise(r.get("error", ""))[:300])

        ws.auto_filter.ref = f"A1:K{len(results) + 1}"
        for col, w in {"A": 35, "B": 55, "C": 10, "D": 8, "E": 10, "F": 20, "G": 40, "H": 40, "I": 55, "J": 55, "K": 50}.items():
            ws.column_dimensions[col].width = w

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = f"test_generic_boards_{timestamp}.xlsx"
        wb.save(out_path)
        console.print(f"\n[bold green]Report saved to:[/bold green] {out_path}")
    except ImportError:
        console.print("[yellow]openpyxl not installed — skipping Excel export[/yellow]")


if __name__ == "__main__":
    asyncio.run(main())
