"""
Quick test — scrape ~100 leads from the fastest API adapters and write to test_output.xlsx.
Run: python3 test_quick.py
"""
from __future__ import annotations

import asyncio
import os
import sys
import time
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))
sys.path.insert(0, str(PROJECT_ROOT))

# Load .env
_env_path = PROJECT_ROOT / ".env"
if _env_path.exists():
    for _line in _env_path.read_text(encoding="utf-8").splitlines():
        _line = _line.strip()
        if not _line or _line.startswith("#") or "=" not in _line:
            continue
        k, v = _line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

from vacancysoft.adapters import (
    AshbyAdapter,
    GreenhouseAdapter,
    LeverAdapter,
    SmartRecruitersAdapter,
    WorkableAdapter,
)
from vacancysoft.db.base import Base
from vacancysoft.db.engine import build_engine
from vacancysoft.db.models import Source
from vacancysoft.db.session import SessionLocal
from vacancysoft.exporters.excel_exporter import export_profile_to_excel
from vacancysoft.pipelines.classification_persistence import classify_enriched_jobs
from vacancysoft.pipelines.enrichment_persistence import enrich_raw_jobs
from vacancysoft.pipelines.persistence import persist_discovery_batch
from vacancysoft.pipelines.scoring_persistence import score_enriched_jobs

MAX_LEADS = 100

# Fast, reliable API-only boards
QUICK_SOURCES = [
    {
        "adapter": "greenhouse",
        "adapter_cls": GreenhouseAdapter,
        "company": "Man Group",
        "config": {"slug": "mangroup", "job_board_url": "https://job-boards.eu.greenhouse.io/mangroup/jobs"},
    },
    {
        "adapter": "greenhouse",
        "adapter_cls": GreenhouseAdapter,
        "company": "Robinhood",
        "config": {"slug": "robinhood", "job_board_url": "https://boards.greenhouse.io/robinhood"},
    },
    {
        "adapter": "greenhouse",
        "adapter_cls": GreenhouseAdapter,
        "company": "Stripe",
        "config": {"slug": "stripe", "job_board_url": "https://boards.greenhouse.io/stripe"},
    },
    {
        "adapter": "workable",
        "adapter_cls": WorkableAdapter,
        "company": "Hayfin",
        "config": {"slug": "hayfin-capital-management", "job_board_url": "https://apply.workable.com/hayfin-capital-management"},
    },
    {
        "adapter": "ashby",
        "adapter_cls": AshbyAdapter,
        "company": "Allica Bank",
        "config": {"slug": "allica-bank", "job_board_url": "https://jobs.ashbyhq.com/allica-bank"},
    },
    {
        "adapter": "lever",
        "adapter_cls": LeverAdapter,
        "company": "Plaid",
        "config": {"slug": "plaid", "job_board_url": "https://jobs.lever.co/plaid"},
    },
    {
        "adapter": "smartrecruiters",
        "adapter_cls": SmartRecruitersAdapter,
        "company": "AJ Bell",
        "config": {"slug": "AJBell1", "job_board_url": "https://jobs.smartrecruiters.com/AJBell1"},
    },
]


def _ensure_source(session, source_def: dict) -> Source:
    from sqlalchemy import select
    from urllib.parse import urlparse

    source_key = f"quick_{source_def['adapter']}_{source_def['company'].lower().replace(' ', '_')}"
    existing = session.execute(select(Source).where(Source.source_key == source_key)).scalar_one_or_none()
    if existing:
        return existing

    url = source_def["config"].get("job_board_url") or "https://example.com"
    hostname = urlparse(url).hostname or "example.com"

    source = Source(
        source_key=source_key,
        employer_name=source_def["company"],
        board_name=source_def["adapter"],
        base_url=url,
        hostname=hostname,
        source_type="quick_test",
        ats_family=source_def["adapter"],
        adapter_name=source_def["adapter"],
        active=True,
        seed_type="quick_seed",
        discovery_method="test_quick",
        fingerprint=f"{hostname}|quick",
        config_blob=source_def["config"],
        capability_blob={},
    )
    session.add(source)
    session.flush()
    return source


def main() -> None:
    print("\n=== Quick Pipeline Test ===\n")

    # Ensure DB schema
    engine = build_engine()
    Base.metadata.create_all(bind=engine)

    # --- Discovery ---
    total = 0
    for src in QUICK_SOURCES:
        label = f"{src['adapter']} / {src['company']}"
        config = dict(src["config"])
        config.setdefault("company", src["company"])
        t0 = time.time()
        try:
            page = asyncio.run(src["adapter_cls"]().discover(source_config=config))
            count = len(page.jobs)
            total += count
            if count:
                with SessionLocal() as session:
                    source_obj = _ensure_source(session, src)
                    persist_discovery_batch(session=session, source=source_obj, records=page.jobs, trigger="test_quick")
            print(f"  {label}: {count} jobs ({time.time() - t0:.1f}s)")
        except Exception as exc:
            print(f"  {label}: FAILED — {exc} ({time.time() - t0:.1f}s)")

        if total >= MAX_LEADS * 3:
            break

    print(f"\nRaw discovered: {total}")

    # --- Enrich ---
    with SessionLocal() as session:
        enriched = enrich_raw_jobs(session, limit=MAX_LEADS * 2)
    print(f"Enriched: {enriched}")

    # --- Classify ---
    with SessionLocal() as session:
        classified = classify_enriched_jobs(session, limit=MAX_LEADS * 2)
    print(f"Classified: {classified}")

    # --- Score ---
    with SessionLocal() as session:
        scored = score_enriched_jobs(session, limit=MAX_LEADS * 2)
    print(f"Scored: {scored}")

    # --- Export ---
    out = str(PROJECT_ROOT / "test_output.xlsx")
    try:
        with SessionLocal() as session:
            path = export_profile_to_excel(session, profile_name="accepted_plus_review", output_path=out, limit=MAX_LEADS)
        print(f"\nWrote {MAX_LEADS} leads -> {path}")
    except Exception as exc:
        print(f"\nExport failed: {exc}")
        return

    # --- Webhook (generic) ---
    webhook_url = os.getenv("WEBHOOK_URL")
    if webhook_url:
        import httpx
        import openpyxl

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [dict(zip(headers, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]
        try:
            resp = httpx.post(webhook_url, json=rows, timeout=60)
            resp.raise_for_status()
            print(f"Posted {len(rows)} leads to webhook")
        except Exception as exc:
            print(f"Webhook post failed: {exc}")
    else:
        print("WEBHOOK_URL not set — skipping webhook")

    print()


if __name__ == "__main__":
    main()
